#!/usr/bin/env python3
"""
Surenka įrašų statistiką iš nurodyto aplanko .
Kiekvienas įrašas turi turėti JSON metaduomenų failą ir atitinkamą duomenų failą.
Duomenų failai gali būti .bin / trijų skaitmenų plėtiniai (size//4 imčių) arba .npy (np.load(...).size).
---
Handles data files with .bin / three-digit extensions (size//4 samples) or .npy (np.load(...).size).

Difference from analyze_records_2026.py: Not using full_list_path Excel any more, just scans all JSON files in data_dir. 

"""
from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, Tuple, Union, Any, Callable, Optional, List
from dataclasses import asdict, dataclass, field
import json

import numpy as np
import pandas as pd

from openpyxl.utils import get_column_letter


JsonLikeRef = Union[str, Path]

# -----------------------------
# Data models
# -----------------------------
@dataclass
class RecordSummary:
    sequenceId: str
    recordingId: Optional[str]
    # basename: str    
    file_name: str
    channel_count: Optional[int]
    user_id: Optional[str]

    bin_bytes: int
    samples: int
    duration_s: Optional[float]

    flags_count: int
    flags: List[str]

    rpeaks_count: int
    ann_n_count: int
    ann_s_count: int
    ann_v_count: int
    ann_u_count: int

    json_ok: bool

    noises_count: int
    noises_fraction: Optional[float]
    
    annotated_noises_count: int
    annotated_noises_fraction: Optional[float]

    has_comment: bool
    json_keys_correct: bool

# def read_filenames_from_excel(xlsx_path: Path) -> tuple[list[str], pd.DataFrame]:
#     """Skaito Excel, grąžina sąrašą `.npy` failų ir visą DF (meta duomenims paimti)."""
#     print("\nSkaitomas Excel: %s", xlsx_path)
#     df = pd.read_excel(xlsx_path, dtype=str)  # saugu tolimesnėms konversijoms
#     if "filename" not in df.columns or "tag" not in df.columns:
#         raise ValueError("Excel must contain columns: 'filename' and 'tag'")

#     filtered = df[df["tag"] != "9999"]
#     names = []
#     for s in filtered["filename"].dropna():
#         name = str(s).strip()
#         if not name.endswith(".npy"):
#             name += ".npy"
#         names.append(name)

#     print("Įrašų sąraše:", len(names))
#     return names, df


def _flatten_flags(flags: Any) -> List[str]:
    if flags is None:
        return []
    if isinstance(flags, list):
        return [str(x) for x in flags]
    # unexpected type
    return [str(flags)]

def _summarize_rpeaks(rpeaks: Any) -> Tuple[int, Optional[int], Optional[int], Dict[str, int]]:
    """
    rpeaks: list of {"sampleIndex": int, "annotationValue": str}
    Returns (count, first_idx, last_idx, annotation_counts)
    """
    if not isinstance(rpeaks, list) or len(rpeaks) == 0:
        return 0, None, None, {}
    idxs: List[int] = []
    ann: Dict[str, int] = {}
    for rp in rpeaks:
        if not isinstance(rp, dict):
            continue
        si = rp.get("sampleIndex")
        av = rp.get("annotationValue")
        if isinstance(si, int):
            idxs.append(si)
        if isinstance(av, str):
            ann[av] = ann.get(av, 0) + 1
        elif av is not None:
            ann[str(av)] = ann.get(str(av), 0) + 1
    if not idxs:
        return 0, None, None, ann
    idxs.sort()
    return len(idxs), idxs[0], idxs[-1], ann

def _sum_noise_samples(noises: Any) -> Tuple[int, int]:
    """
    Returns (interval_count, total_samples_covered) for noises list.
    Accepts items like {"startIndex": x, "endIndex": y}.
    """
    if not isinstance(noises, list):
        return 0, 0
    cnt = 0
    total = 0
    for it in noises:
        if not isinstance(it, dict):
            continue
        a = it.get("startIndex")
        b = it.get("endIndex")
        if isinstance(a, int) and isinstance(b, int) and b > a:
            cnt += 1
            total += (b - a)
    return cnt, total


def _safe_json_load_path(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def summarize_record(
    sequenceId: str,
    # recordingId: str,
    # basename: str,
    file_name: str,
    json_ref: JsonLikeRef,
    data_ref: JsonLikeRef,
    data_ext: str,
    *,
    load_json_fn: Callable[[JsonLikeRef], Any],
    file_size_fn: Callable[[JsonLikeRef], int],
    sample_count_fn: Callable[[JsonLikeRef, str], int],
    fs: int,
) -> Tuple[RecordSummary, Dict[str, Any]]:
    """
    Compute per-record statistics and return both RecordSummary and its dict form.

    This helper is designed to be reusable from other scripts: provide your own
    load_json_fn / file_size_fn / sample_count_fn depending on whether data lives
    in a ZIP or filesystem and what the data extension is.
    """
    try:
        meta = load_json_fn(json_ref)
        json_ok = True
    except Exception as exc:
        print(f"WARN: failed to load JSON for {sequenceId}/{file_name}: {exc}")
        meta = {}
        json_ok = False

    bsz = file_size_fn(data_ref)
    samples = sample_count_fn(data_ref, data_ext)
    # print("meta:", meta)
    channel_count = meta.get("channelCount") if isinstance(meta, dict) else None
    user_id = meta.get("userId") if isinstance(meta, dict) else None
    recordingId = meta.get("recordingId") if isinstance(meta, dict) else None
    flags_list = _flatten_flags(meta.get("flags") if isinstance(meta, dict) else None)

    rpeaks = meta.get("rpeaks") if isinstance(meta, dict) else None
    rpk_cnt, _rpk_first, _rpk_last, rpk_ann = _summarize_rpeaks(rpeaks)

    ann_counts = meta.get("rpeakAnnotationCounts") if isinstance(meta, dict) else None
    ann_counts_clean: Dict[str, int] = {}
    if isinstance(ann_counts, dict):
        for k, v in ann_counts.items():
            if k == "__info":
                continue
            if isinstance(v, int):
                ann_counts_clean[str(k)] = v
    else:
        ann_counts_clean = rpk_ann

    ann_n = int(ann_counts_clean.get("N", 0))
    ann_s = int(ann_counts_clean.get("S", 0))
    ann_v = int(ann_counts_clean.get("V", 0))
    ann_u = int(ann_counts_clean.get("U", 0))

    ann_noises = meta.get("noises_annotated") if isinstance(meta, dict) else None
    ann_nz_cnt, ann_nz_samples = _sum_noise_samples(ann_noises)
    noises = meta.get("noises") if isinstance(meta, dict) else None
    nz_cnt, nz_samples = _sum_noise_samples(noises)

    duration_s = (samples / fs) if (fs is not None and fs > 0 and samples > 0) else None
    noises_fraction = (nz_samples / samples)*100. if (samples > 0) else None
    annotated_noises_fraction = (ann_nz_samples / samples)*100. if (samples > 0) else None

    allowed_keys = {
        "channelCount",
        "comment",
        "flags",
        "noises",
        "noises_annotated",
        "recordingId",
        "rpeakAnnotationCounts",
        "rpeaks",
        "userId",
    }
    meta_keys = set(meta.keys()) if isinstance(meta, dict) else set()
    json_keys_correct = meta_keys.issubset(allowed_keys) if meta_keys else False

    rec = RecordSummary(
        sequenceId=sequenceId,
        recordingId=recordingId,
        # basename=basename,
        file_name=Path(data_ref).name,
        channel_count=int(channel_count) if isinstance(channel_count, int) else None,
        user_id=str(user_id) if isinstance(user_id, str) else None,
        bin_bytes=int(bsz),
        samples=int(samples),
        duration_s=float(duration_s) if duration_s is not None else None,
        flags_count=len(flags_list),
        flags=flags_list,
        rpeaks_count=int(rpk_cnt),
        ann_n_count=ann_n,
        ann_s_count=ann_s,
        ann_v_count=ann_v,
        ann_u_count=ann_u,
        json_ok=json_ok,
        noises_count=int(nz_cnt),
        noises_fraction=float(noises_fraction) if noises_fraction is not None else None,
        annotated_noises_count=int(ann_nz_cnt),
        annotated_noises_fraction=float(annotated_noises_fraction) if annotated_noises_fraction is not None else None,
        has_comment=bool(meta.get("comment")) if isinstance(meta, dict) else False,
        json_keys_correct=json_keys_correct,
    )
    rec_dict = asdict(rec)
    rec_dict["_annotated_noises_samples"] = nz_samples  # helper key for higher-level aggregation
    return rec, rec_dict



def find_data_file(json_path: Path) -> Tuple[Path, str] | None:
    """Return matching data file and extension for a given JSON.

    Supports these layouts (same folder):
      - <id>.json  -> <id>.npy / <id>.bin / <id>.<ddd>
      - <id>.<ddd>.json -> <id>.<ddd>   (three-digit extension kept, then '.json' appended)
    """
    stem = json_path.stem
    rec_dir = json_path.parent

    # Special case: JSON is '<id>.<ddd>.json' and data file is '<id>.<ddd>'
    exact = rec_dir / stem
    if exact.exists() and exact.is_file() and exact.suffix[1:].isdigit() and len(exact.suffix) == 4:
        return exact, exact.suffix

    # priority: .npy, .bin, then any three-digit extension
    candidates = [
        rec_dir / f"{stem}.npy",
        rec_dir / f"{stem}.bin",
    ]
    candidates += sorted(
        [p for p in rec_dir.glob(f"{stem}.*") if p.suffix[1:].isdigit() and len(p.suffix) == 4]
    )
    for p in candidates:
        if p.exists() and p.is_file():
            return p, p.suffix
    return None
def sample_count(path: Path, ext: str) -> int:
    ext = ext.lower()
    if ext == ".npy":
        arr = np.load(path, mmap_mode="r", allow_pickle=False)
        return int(getattr(arr, "size", 0))
    return int(path.stat().st_size // 4)


def load_json(path: Path):
    return _safe_json_load_path(path)


def file_size(path: Path) -> int:
    return path.stat().st_size


def infer_sequence_id(json_path: Path) -> str:
    # expected layout: <seq>/recordings/<id>.json
    return json_path.parent.parent.name if json_path.parent.name == "recordings" else json_path.parent.name


def resolve_out_paths(out_arg, default_dir: Path, default_stem: str = "records_summary"):
    """
    Returns (csv_path, xlsx_path) based on --out.
    --out can be:
      - None -> default_dir/default_stem.{csv,xlsx}
      - a path with or without suffix -> treated as base path/stem
      - a directory -> file name uses default_stem inside that directory
    """
    if out_arg:
        p = Path(out_arg)
    else:
        p = default_dir / default_stem

    # If user passed a directory, put default file name inside it
    if p.exists() and p.is_dir():
        p = p / default_stem
    elif str(p).endswith(("/", "\\")):  # in case directory doesn't exist yet
        p = p / default_stem

    # Drop suffix if they gave one (csv/xlsx/anything) -> use as base
    base = p.with_suffix("")

    csv_path = base.with_suffix(".csv")
    xlsx_path = base.with_suffix(".xlsx")
    return csv_path, xlsx_path


def main() -> int:
    ap = argparse.ArgumentParser(description="Summarize ECG records in a folder (non-zip).")
    ap.add_argument("data_dir", type=Path, help="Root folder containing sequence subfolders with recordings/")
    # ap.add_argument("--full_list_path", type=Path, required=True, help="excel with full listof annotated recordings")
    ap.add_argument("--fs", type=int, required=True, help="Sampling frequency in Hz")
    ap.add_argument("--out", type=Path, default=None, help="CSV output path (default: <data_dir>/records_summary.csv)")
    args = ap.parse_args()

    data_dir = args.data_dir
    if not data_dir.exists():
        raise SystemExit(f"data_dir not found: {data_dir}")

    # # Nuskaitome excel failą su duomenimis apie įrašus su anotacijomis ir triukšmais (anotuotų įrašų sąrašas),
    # # sukuriame failų vardų sąrašą file_names ir dataframe df_meta su įrašų duomenimis.
    # # Reikalingas, kad pagal lokalų failo vardą (tipo filename =1001_01.json) galėtume rasti Zive failo vardą
    # # (tipo basename=1626941.468)
    # file_names, df_meta = read_filenames_from_excel(args.full_list_path)
    # # print(file_names)
    # # print(df_meta.head())


    # Raskime visus JSON failus in data directory
    json_files = sorted(data_dir.rglob("*.json"))
    records: list[Dict] = []

    # Surandame kiekvieno įrašo statistiką ir kaupiame įrašus
    for jp in json_files:
        # print(f"Processing JSON: {jp}")
        match = find_data_file(jp)
        if match is None:
            print(f"WARN: no data file found for {jp}")
            continue
        data_path, data_ext = match
        seq_id = infer_sequence_id(jp)
        # rec_id = jp.stem
        filename = jp.stem
        # basename = df_meta.set_index("filename")["basename"].loc[filename]
        # print(f"Found data file: {data_path} for JSON: {jp}, sequence: {seq_id}, basename: {basename}")

        rec, rec_dict = summarize_record(
            sequenceId=seq_id,
            # recordingId=rec_id,
            # basename=basename,
            file_name=match[0].name,
            json_ref=jp,
            data_ref=data_path,
            data_ext=data_ext,
            load_json_fn=load_json,
            file_size_fn=file_size,
            sample_count_fn=sample_count,
            fs=args.fs,
        )
        records.append(rec_dict)
        # print(rec_dict)

    # Konvertuojame sukauptus įrašus į DataFrame
    df = pd.DataFrame(records)

    # 1) Add row number (nr)
    df = df.copy()
    df.insert(0, "nr", range(1, len(df) + 1))

    # Atsirenkame parametrus, kurie bus rodomi santraukoje
    # 2) Select only the columns you want (skip any missing ones safely)
    cols = [
        "nr", "file_name", "recordingId", "user_id", "samples", "duration_s",
        "rpeaks_count", "ann_n_count", "ann_s_count", "ann_v_count", "ann_u_count",
        "annotated_noises_count", "annotated_noises_fraction", "noises_count", "noises_fraction","flags"
    ]
    cols = [c for c in cols if c in df.columns]
    df_sel = df.loc[:, cols]
    # print("\ndf_sel:", df_sel.head())
    
    # Sutrumpiname stulpelių pavadinimus
    # short display names (only those present will be applied)
    col_short = {
        "recordingId": "rec_id",
        "user_id": "uid",
        "duration_s": "dur_s",
        "rpeaks_count": "rpk_cnt",
        "annotated_noises_count": "ann_nz_cnt",
        "annotated_noises_fraction": "ann_nz_frac",
        "noises_count": "nz_cnt",
        "noises_fraction": "nz_frac",
        "ann_n_count": "annN",
        "ann_s_count": "annS",
        "ann_v_count": "annV",
        "ann_u_count": "annU",
    }

    df_print = df_sel.rename(columns=col_short)  # type: ignore
    # print("\ndf_print:", df_print.head())
   
    # Inserting 2 new columns for external annotating and noises annotated
    i = list(df_print.columns).index("nr") + 1
    df_print.insert(i,   "FOR_EXTERNAL_ANNOTATING", "")
    df_print.insert(i+1, "NOISES_ANNOTATED",        "")

        # Įrašome santraukos statistiką į CSV ir Excel failą
   
    csv_path, xlsx_path = resolve_out_paths(args.out, data_dir, "records_summary")
    df_print.to_csv(csv_path, index=False)
    # df_print.to_excel(xlsx_path, index=False, engine="openpyxl")
    
        # Įrašome santraukos statistiką į Excel failą
    df_out = df_print.copy()

    # Ensure numeric so Excel can format properly
    for c in ["dur_s", "ann_nz_frac", "nz_frac"]:
        if c in df_out.columns:
            df_out[c] = pd.to_numeric(df_out[c], errors="coerce")
            
    # Įrašant pakoreguojame formatavimą kai kuriems stulpeliams
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        sheet = "summary"
        df_out.to_excel(writer, index=False, sheet_name=sheet)
        ws = writer.book[sheet]

        fmt_map = {
            "dur_s": "0.0",       # .1f
            "ann_nz_frac": '0.0"%"',  # .1f (already 0–100)
            "nz_frac": '0.0"%"',      # .1f (already 0–100)
        }

        headers = [cell.value for cell in ws[1]]
        col_idx = {name: i + 1 for i, name in enumerate(headers) if name is not None}

        for name, fmt in fmt_map.items():
            if name not in col_idx:
                continue
            c = col_idx[name]
            for r in range(2, ws.max_row + 1):  # skip header
                ws.cell(row=r, column=c).number_format = fmt


    # Išvedame santraukos statistiką į konsolę
    print(f"\nData directory: {data_dir}")
    print(f"Saved CSV : {csv_path}")
    print(f"Saved XLSX: {xlsx_path}")
    print("Summary statistics:")
    print(
        df_print.to_string(
            index=False,
            float_format=None,
            formatters={
                "dur_s": lambda x: f"{x:.2f}" if pd.notna(x) else "",
                "ann_nz_frac": lambda x: f"{x:.1f}%" if pd.notna(x) else "",
                "nz_frac": lambda x: f"{x:.1f}%" if pd.notna(x) else "",
            }
        )
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

"""
python3 analyze_records_2026.py user-65955b5f50e02b125d4998ad/659ec124870b3d1d1630be39/recordings --fs 200 --out records_summary.csv

"""