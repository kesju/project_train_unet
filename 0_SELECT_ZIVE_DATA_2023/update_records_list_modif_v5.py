#!/usr/bin/env python3
"""
Update ECG record parameters in an Excel file from per-record JSON metadata stored in a ZIP,
WITHOUT importing analyze_records_2026_v4.py (helpers are embedded), and highlight changed rows.

What it does (per JSON record):
- Find matching row in Excel by `basename`
- Update columns:
    tag      -> 1111 if JSON flags contain "AI_DATA"
    noni     -> h_nz_cnt (human annotated noise count)
    mlS/mlV/mlU
    h_nz_cnt / h_nz_len / h_nz_frac
- Mark the *basename cell* of every changed row with light blue fill.

Notes:
- Matching by basename is normalized: commas -> dots, trailing zeros after decimal are removed.
- h_nz_len is total annotated-noise length in samples (sum(end-start) over noises_annotated).
- h_nz_frac is annotated-noise fraction in percent.
"""

from __future__ import annotations

import argparse
import math
import re
import zipfile
from pathlib import Path
from typing import Any, Dict, Optional, Set

import openpyxl
from openpyxl.styles import PatternFill

# ---------------------------------------------------------------------
# Embedded analyzer helpers (copied from analyze_records_2026_v4.py)
# ---------------------------------------------------------------------
# -----------------------------
# Embedded analyzer helpers (from analyze_records_2026_v4.py)
# -----------------------------
import io
import json
from dataclasses import asdict, dataclass
from pathlib import PurePosixPath
from typing import Callable, Dict, List, Optional, Tuple, Union, Any

import numpy as np

JsonLikeRef = Union[str, Path]

@dataclass
class RecordSummary:
    sequenceId: str
    recordingId: Optional[str]
    basename: str
    channel_count: Optional[int]
    user_id: Optional[str]

    bin_bytes: int
    samples: int
    duration_s: Optional[float]

    flags_count: int
    flags: List[str]

    rpeaks_count: int
    h_n_count: int
    h_s_count: int
    h_v_count: int
    h_u_count: int

    ml_s_count: int
    ml_v_count: int
    ml_u_count: int

    json_ok: bool

    ml_noises_count: int
    ml_noises_fraction: Optional[float]

    h_noises_count: int
    h_noises_fraction: Optional[float]

    has_comment: bool
    json_keys_correct: bool


def _flatten_flags(flags: Any) -> List[str]:
    if flags is None:
        return []
    if isinstance(flags, list):
        return [str(x) for x in flags]
    return [str(flags)]


def _pick_variant_dict(obj: Any, keys_preference: Tuple[str, ...] = ("merged", "human", "ml")) -> Any:
    if not isinstance(obj, dict):
        return None
    for k in keys_preference:
        if k in obj:
            return obj.get(k)
    for k, v in obj.items():
        if k == "__info":
            continue
        return v
    return None


def _extract_rpeaks_list(rpeaks: Any) -> List[Dict[str, Any]]:
    if isinstance(rpeaks, list):
        return [x for x in rpeaks if isinstance(x, dict)]
    if isinstance(rpeaks, dict):
        v = _pick_variant_dict(rpeaks, ("merged", "human", "ml"))
        if isinstance(v, list):
            return [x for x in v if isinstance(x, dict)]
    return []


def _extract_h_counts(h_counts: Any) -> Dict[str, int]:
    if not isinstance(h_counts, dict):
        return {}
    # Variant dict?
    if any(isinstance(v, dict) for k, v in h_counts.items() if k != "__info"):
        v = _pick_variant_dict(h_counts, ("merged", "human", "ml"))
        if isinstance(v, dict):
            out: Dict[str, int] = {}
            for k, val in v.items():
                if k == "__info":
                    continue
                if isinstance(val, int):
                    out[str(k)] = int(val)
                elif isinstance(val, (float, np.floating)) and float(val).is_integer():
                    out[str(k)] = int(val)
            return out
        return {}
    # Flat dict
    out2: Dict[str, int] = {}
    for k, v in h_counts.items():
        if k == "__info":
            continue
        if isinstance(v, int):
            out2[str(k)] = int(v)
        elif isinstance(v, (float, np.floating)) and float(v).is_integer():
            out2[str(k)] = int(v)
    return out2


def _extract_rpeaks_list_for_variant(rpeaks: Any, variant: str) -> List[Dict[str, Any]]:
    if isinstance(rpeaks, dict):
        v = rpeaks.get(variant)
        if isinstance(v, list):
            return [x for x in v if isinstance(x, dict)]
    return []


def _extract_h_counts_for_variant(h_counts: Any, variant: str) -> Dict[str, int]:
    if not isinstance(h_counts, dict):
        return {}
    v = h_counts.get(variant)
    if not isinstance(v, dict):
        return {}
    out: Dict[str, int] = {}
    for k, val in v.items():
        if k == "__info":
            continue
        if isinstance(val, int):
            out[str(k)] = int(val)
        elif isinstance(val, (float, np.floating)) and float(val).is_integer():
            out[str(k)] = int(val)
    return out


def _summarize_rpeaks(rpeaks: Any) -> Tuple[int, Optional[int], Optional[int], Dict[str, int]]:
    lst = _extract_rpeaks_list(rpeaks)
    if not lst:
        return 0, None, None, {}
    idxs: List[int] = []
    ann: Dict[str, int] = {}
    for rp in lst:
        si = rp.get("sampleIndex")
        av = rp.get("annotationValue")
        if isinstance(si, int):
            idxs.append(si)
        elif isinstance(si, (float, np.floating)) and float(si).is_integer():
            idxs.append(int(si))
        if isinstance(av, str):
            ann[av] = ann.get(av, 0) + 1
        elif av is not None:
            s = str(av)
            ann[s] = ann.get(s, 0) + 1
    if not idxs:
        return len(lst), None, None, ann
    idxs.sort()
    return len(lst), idxs[0], idxs[-1], ann


def _sum_noise_samples(noises: Any, fs: Optional[int]) -> Tuple[int, int]:
    # Variant dict?
    if isinstance(noises, dict):
        v = _pick_variant_dict(noises, ("merged", "human", "ml"))
        noises = v
    if not isinstance(noises, list):
        return 0, 0
    cnt = 0
    total = 0
    for it in noises:
        a = b = None
        if isinstance(it, dict):
            a = it.get("startIndex", it.get("startSample"))
            b = it.get("endIndex", it.get("endSample"))
            if (a is None or b is None) and fs:
                t0 = it.get("startTime")
                t1 = it.get("endTime")
                if isinstance(t0, (int, float, np.floating)) and isinstance(t1, (int, float, np.floating)) and t1 > t0:
                    a = int(round(float(t0) * fs))
                    b = int(round(float(t1) * fs))
        elif isinstance(it, (list, tuple)) and len(it) >= 2:
            a, b = it[0], it[1]
        if isinstance(a, (float, np.floating)) and float(a).is_integer():
            a = int(a)
        if isinstance(b, (float, np.floating)) and float(b).is_integer():
            b = int(b)
        if isinstance(a, int) and isinstance(b, int) and b > a:
            cnt += 1
            total += (b - a)
    return cnt, total


def load_json_zip(zf: zipfile.ZipFile, member: str) -> Any:
    with zf.open(member) as f:
        return json.load(f)


def file_size_zip(zf: zipfile.ZipFile, member: str) -> int:
    return int(zf.getinfo(member).file_size)


def sample_count_zip(zf: zipfile.ZipFile, member: str, ext: str) -> int:
    ext = (ext or "").lower()
    if ext == ".npy":
        raw = zf.read(member)
        arr = np.load(io.BytesIO(raw), allow_pickle=False)
        return int(getattr(arr, "size", 0))
    return int(file_size_zip(zf, member) // 4)


def summarize_record(
    sequenceId: str,
    basename: str,
    json_ref: JsonLikeRef,
    data_ref: JsonLikeRef,
    data_ext: str,
    *,
    load_json_fn: Callable[[JsonLikeRef], Any],
    file_size_fn: Callable[[JsonLikeRef], int],
    sample_count_fn: Callable[[JsonLikeRef, str], int],
    fs: int,
) -> Tuple[RecordSummary, Dict[str, Any]]:
    try:
        meta = load_json_fn(json_ref)
        json_ok = True
    except Exception as exc:
        print(f"WARN: failed to load JSON for {sequenceId}/{basename}: {exc}")
        meta = {}
        json_ok = False

    bsz = int(file_size_fn(data_ref))
    samples = int(sample_count_fn(data_ref, data_ext))

    channel_count = meta.get("channelCount") if isinstance(meta, dict) else None
    user_id = meta.get("userId") if isinstance(meta, dict) else None
    recordingId = meta.get("recordingId") if isinstance(meta, dict) else None
    flags_list = _flatten_flags(meta.get("flags") if isinstance(meta, dict) else None)

    rpeaks_any = meta.get("rpeaks") if isinstance(meta, dict) else None
    rpk_cnt, _rpk_first, _rpk_last, rpk_ann = _summarize_rpeaks(rpeaks_any)

    h_counts_any = meta.get("rpeakAnnotationCounts") if isinstance(meta, dict) else None

    # IMPORTANT:
    # Only treat counts as "human" if they come from explicit "human" variant
    # (or from a flat dict without variants). This prevents accidentally
    # copying ML counts into hS/hV/hU when JSON lacks human labels.
    h_counts_human = _extract_h_counts_for_variant(h_counts_any, "human")

    # If JSON has no explicit human counts, try to derive them from human rpeaks variant.
    if not h_counts_human:
        human_rpeaks = _extract_rpeaks_list_for_variant(rpeaks_any, "human")
        if human_rpeaks:
            tmp: Dict[str, int] = {}
            for rp in human_rpeaks:
                av = rp.get("annotationValue")
                if isinstance(av, str):
                    tmp[av] = tmp.get(av, 0) + 1
                elif av is not None:
                    s = str(av)
                    tmp[s] = tmp.get(s, 0) + 1
            h_counts_human = tmp

    # If rpeakAnnotationCounts is a flat dict (no variants), accept it as human.
    if not h_counts_human:
        flat = _extract_h_counts(h_counts_any)
        if flat:
            h_counts_human = flat

    has_human_counts = bool(h_counts_human)

    h_n = int(h_counts_human.get("N", 0)) if has_human_counts else 0
    h_s = int(h_counts_human.get("S", 0)) if has_human_counts else 0
    h_v = int(h_counts_human.get("V", 0)) if has_human_counts else 0
    h_u = int(h_counts_human.get("U", 0)) if has_human_counts else 0

    ml_counts = _extract_h_counts_for_variant(h_counts_any, "ml")
    if not ml_counts:
        ml_rpeaks = _extract_rpeaks_list_for_variant(rpeaks_any, "ml")
        tmp: Dict[str, int] = {}
        for rp in ml_rpeaks:
            av = rp.get("annotationValue")
            if isinstance(av, str):
                tmp[av] = tmp.get(av, 0) + 1
            elif av is not None:
                s = str(av)
                tmp[s] = tmp.get(s, 0) + 1
        ml_counts = tmp

    ml_s = int(ml_counts.get("S", 0))
    ml_v = int(ml_counts.get("V", 0))
    ml_u = int(ml_counts.get("U", 0))

    h_noises = meta.get("noises_annotated") if isinstance(meta, dict) else None
    h_nz_cnt, h_nz_samples = _sum_noise_samples(h_noises, fs)
    noises = meta.get("noises") if isinstance(meta, dict) else None
    ml_nz_cnt, nz_samples = _sum_noise_samples(noises, fs)

    duration_s = (samples / fs) if (fs and fs > 0 and samples > 0) else None
    ml_noises_fraction = (nz_samples / samples) * 100.0 if samples > 0 else None
    h_noises_fraction = (h_nz_samples / samples) * 100.0 if samples > 0 else None

    allowed_keys = {
        "channelCount",
        "comment",
        "flags",
        "noises",
        "noises_annotated",
        "recordingId",
        "rpeakAnnotationCounts",
        "rpeaks",
        "userId",
    }
    meta_keys = set(meta.keys()) if isinstance(meta, dict) else set()
    json_keys_correct = meta_keys.issubset(allowed_keys) if meta_keys else False

    rec = RecordSummary(
        sequenceId=str(sequenceId),
        recordingId=str(recordingId) if isinstance(recordingId, str) else None,
        basename=str(basename),
        channel_count=int(channel_count) if isinstance(channel_count, int) else None,
        user_id=str(user_id) if isinstance(user_id, str) else None,
        bin_bytes=bsz,
        samples=samples,
        duration_s=float(duration_s) if duration_s is not None else None,
        flags_count=len(flags_list),
        flags=flags_list,
        rpeaks_count=int(rpk_cnt),
        h_n_count=h_n,
        h_s_count=h_s,
        h_v_count=h_v,
        h_u_count=h_u,
        ml_s_count=int(ml_s),
        ml_v_count=int(ml_v),
        ml_u_count=int(ml_u),
        json_ok=bool(json_ok),
        ml_noises_count=int(ml_nz_cnt),
        ml_noises_fraction=float(ml_noises_fraction) if ml_noises_fraction is not None else None,
        h_noises_count=int(h_nz_cnt),
        h_noises_fraction=float(h_noises_fraction) if h_noises_fraction is not None else None,
        has_comment=bool(meta.get("comment")) if isinstance(meta, dict) else False,
        json_keys_correct=bool(json_keys_correct),
    )

    rec_dict = asdict(rec)
    rec_dict["__has_human_counts"] = bool(has_human_counts)
    rec_dict["_noises_samples"] = nz_samples
    rec_dict["_annotated_noises_samples"] = h_nz_samples
    return rec, rec_dict


def find_data_file_zip(json_member: str, name_set: set[str]) -> Tuple[str, str] | None:
    pp = PurePosixPath(json_member)
    if pp.suffix.lower() != ".json":
        return None
    rec_dir = pp.parent
    stem = pp.stem

    exact = str(rec_dir / stem)
    ex_suffix = PurePosixPath(exact).suffix
    if exact in name_set and ex_suffix[1:].isdigit() and len(ex_suffix) == 4:
        return exact, ex_suffix

    candidates = [
        str(rec_dir / f"{stem}.npy"),
        str(rec_dir / f"{stem}.bin"),
        str(rec_dir / f"{stem}"),
    ]
    prefix = str(rec_dir / stem) + "."
    for n in sorted(name_set):
        if n.startswith(prefix):
            suf = PurePosixPath(n).suffix
            if suf[1:].isdigit() and len(suf) == 4:
                candidates.append(n)

    for c in candidates:
        if c in name_set:
            return c, PurePosixPath(c).suffix
    return None


def infer_sequence_id_zip(json_member: str, fallback: str) -> str:
    pp = PurePosixPath(json_member)
    parts = list(pp.parts)
    if "recordings" in parts:
        idx = parts.index("recordings")
        if idx >= 1:
            return parts[idx - 1]
    if len(parts) >= 2:
        return parts[-2]
    return fallback



def load_json_fs(p: Path) -> Any:
    with open(p, "r", encoding="utf-8") as f:
        return json.load(f)


def file_size_fs(p: Path) -> int:
    return int(p.stat().st_size)


def sample_count_fs(p: Path, ext: str) -> int:
    ext = (ext or "").lower()
    if ext == ".npy":
        arr = np.load(p, allow_pickle=False)
        return int(getattr(arr, "size", 0))
    # Default: assume 4 bytes per sample (int32)
    return int(file_size_fs(p) // 4)


def infer_sequence_id_fs(json_path: Path, fallback: str) -> str:
    parts = list(json_path.parts)
    if "recordings" in parts:
        idx = parts.index("recordings")
        if idx >= 1:
            return parts[idx - 1]
    if len(parts) >= 2:
        return parts[-2]
    return fallback


def find_data_file_fs(json_path: Path) -> Tuple[Path, str] | None:
    if json_path.suffix.lower() != ".json":
        return None
    rec_dir = json_path.parent
    stem = json_path.stem

    exact = rec_dir / stem
    # Prefer numeric suffix file like 1626941.468 (suffix ".468")
    if exact.exists():
        suf = exact.suffix
        if suf and suf[1:].isdigit() and len(suf) == 4:
            return exact, suf

    candidates = [
        rec_dir / f"{stem}.npy",
        rec_dir / f"{stem}.bin",
        rec_dir / f"{stem}",
    ]

    # Any file like "<stem>.<3digits>"
    for p in sorted(rec_dir.glob(f"{stem}.*")):
        suf = p.suffix
        if suf and suf[1:].isdigit() and len(suf) == 4:
            candidates.append(p)

    for c in candidates:
        if c.exists():
            return c, c.suffix
    return None

# ---------------------------------------------------------------------
# Excel update + highlighting
# ---------------------------------------------------------------------

BLUE_FILL = PatternFill(patternType="solid", fgColor="FF1E90FF")  # ARGB


# Classic pure blue
# BLUE_FILL = PatternFill(patternType="solid", fgColor="FF0000FF")  # opaque blue
# 2) Medium blue (less intense)
# BLUE_FILL = PatternFill(patternType="solid", fgColor="FF1E90FF")  # DodgerBlue
# 3) Light blue (what you currently had)
# BLUE_FILL = PatternFill(patternType="solid", fgColor="FFEAF2FF")  # very light blue
# Pure yellow
# YELLOW_FILL = PatternFill(patternType="solid", fgColor="FFFFFF00")  # opaque yellow
# 2) Light yellow (Excel-like highlight)
# YELLOW_FILL = PatternFill(patternType="solid", fgColor="FFFFF2CC")  # light yellow
# 3) Softer pastel yellow
# YELLOW_FILL = PatternFill(patternType="solid", fgColor="FFFFF9C4")


def norm_basename(x: Any) -> Optional[str]:
    """Normalize basenames for matching: ',' -> '.', strip, and drop trailing zeros in decimals."""
    if x is None:
        return None
    s = str(x).strip()
    if not s:
        return None
    s = s.replace(",", ".")
    m = re.fullmatch(r"(\d+)\.(\d+)", s)
    if m:
        a, b = m.group(1), m.group(2).rstrip("0")
        return a if b == "" else f"{a}.{b}"
    return s


def build_header_map(ws) -> Dict[str, int]:
    hdr: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        key = str(v).strip().lower()
        if key:
            hdr[key] = col
    return hdr


def pick_col(hdr: Dict[str, int], *names: str) -> Optional[int]:
    for n in names:
        k = n.strip().lower()
        if k in hdr:
            return hdr[k]
    return None


def cell_changed(old: Any, new: Any) -> bool:
    """Loose compare to avoid false mismatches (e.g. '12,3' vs 12.3)."""
    def norm(v: Any) -> Any:
        if v is None:
            return None
        if isinstance(v, bool):
            return bool(v)
        if isinstance(v, int):
            return int(v)
        if isinstance(v, float):
            if math.isnan(v):
                return None
            return float(v)
        s = str(v).strip()
        try:
            return float(s.replace(",", "."))
        except Exception:
            return s

    return norm(old) != norm(new)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", type=Path, required=True, help="Input Excel (.xlsx)")
    ap.add_argument("--zip", dest="zip_path", type=Path, required=True, help="ZIP file OR a folder containing JSON metadata and data files")
    ap.add_argument("--fs", type=int, default=200, help="Sampling frequency (Hz). Default: 200")
    ap.add_argument("--sheet", type=str, default=None, help="Sheet name (default: active sheet)")
    ap.add_argument("--out", type=Path, default=None, help="Output Excel path (default: <excel>_updated.xlsx)")
    args = ap.parse_args()

    # Load twice: one workbook with formulas and one with computed values (data_only)
    wb = openpyxl.load_workbook(args.excel)
    wb_values = openpyxl.load_workbook(args.excel, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.active
    ws_values = wb_values[args.sheet] if args.sheet else wb_values.active

    hdr = build_header_map(ws)

    c_basename = pick_col(hdr, "basename")
    if c_basename is None:
        raise RuntimeError("Column 'basename' not found in row 1 header.")

    # Required columns (case-insensitive)
    cols = {
        "tag": pick_col(hdr, "tag"),
        "noni": pick_col(hdr, "noni"),

        # Human r-peak annotation counts
        "hN": pick_col(hdr, "hN", "hn"),
        "hS": pick_col(hdr, "hS", "hs"),
        "hV": pick_col(hdr, "hV", "hv"),
        "hU": pick_col(hdr, "hU", "hu"),

        # ML r-peak annotation counts
        "mlS": pick_col(hdr, "mls", "mlS"),
        "mlV": pick_col(hdr, "mlv", "mlV"),
        "mlU": pick_col(hdr, "mlu", "mlU"),

        # Human annotated noise (noises_annotated)
        "h_nz_cnt": pick_col(hdr, "h_nz_cnt"),
        "h_nz_len": pick_col(hdr, "h_nz_len"),
        "h_nz_frac": pick_col(hdr, "h_nz_frac", "h_nz_frac%", "h_nz_frac %"),
    }
    missing = [k for k, v in cols.items() if v is None]
    if missing:
        raise RuntimeError(f"Missing required columns in Excel header row: {missing}")

    # Map basename -> row index
    basename_to_row: Dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        bn = norm_basename(ws.cell(row=r, column=c_basename).value)
        if bn:
            basename_to_row[bn] = r

    changed_rows: Set[int] = set()
    processed = 0
    not_found: Set[str] = set()


    src = args.zip_path
    print(f"Source: {src} (is_dir={src.is_dir()}, is_file={src.is_file()})")
    if src.is_dir():
        # ----------------------------
        # Source is an unzipped folder
        # ----------------------------
        json_paths = [
            p for p in src.rglob("*.json")
            if "__MACOSX" not in p.parts
            and p.is_file()
        ]
        print(f"Found JSON files: {len(json_paths)}")
        
        for jp in sorted(json_paths):
            stem = jp.stem
            bn = norm_basename(stem)
            if not bn:
                continue
    
            row = basename_to_row.get(bn)
            if row is None:
                not_found.add(bn)
                continue
    
            data_info = find_data_file_fs(jp)
            print(f"Processing JSON: {jp} (data_info={data_info})")
            
            if data_info is None:
                # JSON-only fallback: cannot compute fraction without sample count
                meta = load_json_fs(jp)
                print(f"Loaded JSON metadata for {jp}: {meta}")
                flags_list = _flatten_flags(meta.get("flags") if isinstance(meta, dict) else None)
                h_noises = meta.get("noises_annotated") if isinstance(meta, dict) else None
                h_nz_cnt, h_nz_samples = _sum_noise_samples(h_noises, args.fs)
                h_frac = None
    
                h_counts_any = meta.get("rpeakAnnotationCounts") if isinstance(meta, dict) else None

                h_counts_human = _extract_h_counts_for_variant(h_counts_any, "human")
                if not h_counts_human:
                    flat = _extract_h_counts(h_counts_any)
                    if flat:
                        h_counts_human = flat
                has_human_counts = bool(h_counts_human)

                h_n = int(h_counts_human.get("N", 0)) if has_human_counts else 0
                h_s = int(h_counts_human.get("S", 0)) if has_human_counts else 0
                h_v = int(h_counts_human.get("V", 0)) if has_human_counts else 0
                h_u = int(h_counts_human.get("U", 0)) if has_human_counts else 0
                ml_counts = _extract_h_counts_for_variant(h_counts_any, "ml")
                ml_s = int(ml_counts.get("S", 0)) if ml_counts else 0
                ml_v = int(ml_counts.get("V", 0)) if ml_counts else 0
                ml_u = int(ml_counts.get("U", 0)) if ml_counts else 0
    
                ai_data = "AI_DATA" in flags_list
            else:
                data_path, data_ext = data_info
                seq = infer_sequence_id_fs(jp, fallback=src.name or "dir")
                rec, rec_dict = summarize_record(
                    sequenceId=seq,
                    basename=bn,
                    json_ref=jp,
                    data_ref=data_path,
                    data_ext=data_ext,
                    load_json_fn=load_json_fs,
                    file_size_fn=file_size_fs,
                    sample_count_fn=sample_count_fs,
                    fs=args.fs,
                )
                has_human_counts = bool(rec_dict.get("__has_human_counts"))
                ai_data = "AI_DATA" in (rec.flags or [])
                h_nz_cnt = int(rec.h_noises_count)
                h_nz_samples = int(rec_dict.get("_annotated_noises_samples", 0))
                h_frac = rec.h_noises_fraction
                ml_s = int(rec.ml_s_count)
                ml_v = int(rec.ml_v_count)
                ml_u = int(rec.ml_u_count)
                h_n = int(rec.h_n_count)
                h_s = int(rec.h_s_count)
                h_v = int(rec.h_v_count)
                h_u = int(rec.h_u_count)
    
            def set_cell(col_key: str, new_val: Any) -> bool:
                col = cols[col_key]
                cell = ws.cell(row=row, column=col)
                old_val = cell.value
                if isinstance(new_val, float):
                    new_val = round(float(new_val), 1)
                if cell_changed(old_val, new_val):
                    cell.value = new_val
                    return True
                return False
    
            row_changed = False
    
            # tag: only set if AI_DATA present
            if ai_data:
                row_changed |= set_cell("tag", 1111)
    
            # noni = h_nz_cnt
            row_changed |= set_cell("noni", int(h_nz_cnt))

            # Human counts (N/S/V/U)
            if has_human_counts:
                row_changed |= set_cell("hN", int(h_n))
                row_changed |= set_cell("hS", int(h_s))
                row_changed |= set_cell("hV", int(h_v))
                row_changed |= set_cell("hU", int(h_u))
    
            # ML counts
            row_changed |= set_cell("mlS", int(ml_s))
            row_changed |= set_cell("mlV", int(ml_v))
            row_changed |= set_cell("mlU", int(ml_u))
    
            # Human annotated noise
            row_changed |= set_cell("h_nz_cnt", int(h_nz_cnt))
            row_changed |= set_cell("h_nz_len", int(h_nz_samples))
            if h_frac is not None:
                row_changed |= set_cell("h_nz_frac", float(h_frac))
    
            if row_changed:
                changed_rows.add(row)
    
            processed += 1
    
    else:
        # ----------------------------
        # Source is a ZIP archive
        # ----------------------------
        if not src.is_file() or src.suffix.lower() != ".zip":
            raise FileNotFoundError(f"--zip must be a .zip file or a directory. Got: {src}")
    
        with zipfile.ZipFile(src, "r") as zf:
            name_set = set(zf.namelist())
            json_members = [
                n for n in name_set
                if n.lower().endswith(".json")
                and not n.endswith("/")
                and not n.startswith("__MACOSX/")
            ]
    
            for jm in sorted(json_members):
                stem = Path(jm).name[:-5]  # drop ".json"
                bn = norm_basename(stem)
                if not bn:
                    continue
    
                row = basename_to_row.get(bn)
                if row is None:
                    not_found.add(bn)
                    continue
    
                data_info = find_data_file_zip(jm, name_set)
    
                if data_info is None:
                    # JSON-only fallback: cannot compute fraction without sample count
                    meta = load_json_zip(zf, jm)
                    flags_list = _flatten_flags(meta.get("flags") if isinstance(meta, dict) else None)
                    h_noises = meta.get("noises_annotated") if isinstance(meta, dict) else None
                    h_nz_cnt, h_nz_samples = _sum_noise_samples(h_noises, args.fs)
                    h_frac = None
    
                    h_counts_any = meta.get("rpeakAnnotationCounts") if isinstance(meta, dict) else None

                    h_counts_human = _extract_h_counts_for_variant(h_counts_any, "human")
                    if not h_counts_human:
                        flat = _extract_h_counts(h_counts_any)
                        if flat:
                            h_counts_human = flat
                    has_human_counts = bool(h_counts_human)

                    h_n = int(h_counts_human.get("N", 0)) if has_human_counts else 0
                    h_s = int(h_counts_human.get("S", 0)) if has_human_counts else 0
                    h_v = int(h_counts_human.get("V", 0)) if has_human_counts else 0
                    h_u = int(h_counts_human.get("U", 0)) if has_human_counts else 0
                    ml_counts = _extract_h_counts_for_variant(h_counts_any, "ml")
                    ml_s = int(ml_counts.get("S", 0)) if ml_counts else 0
                    ml_v = int(ml_counts.get("V", 0)) if ml_counts else 0
                    ml_u = int(ml_counts.get("U", 0)) if ml_counts else 0
    
                    ai_data = "AI_DATA" in flags_list
                else:
                    data_member, data_ext = data_info
                    seq = infer_sequence_id_zip(jm, fallback="zip")
                    rec, rec_dict = summarize_record(
                        sequenceId=seq,
                        basename=bn,
                        json_ref=jm,
                        data_ref=data_member,
                        data_ext=data_ext,
                        load_json_fn=lambda m: load_json_zip(zf, str(m)),
                        file_size_fn=lambda m: file_size_zip(zf, str(m)),
                        sample_count_fn=lambda m, ext: sample_count_zip(zf, str(m), ext),
                        fs=args.fs,
                    )
                    has_human_counts = bool(rec_dict.get("__has_human_counts"))
                    ai_data = "AI_DATA" in (rec.flags or [])
                    h_nz_cnt = int(rec.h_noises_count)
                    h_nz_samples = int(rec_dict.get("_annotated_noises_samples", 0))
                    h_frac = rec.h_noises_fraction
                    ml_s = int(rec.ml_s_count)
                    ml_v = int(rec.ml_v_count)
                    ml_u = int(rec.ml_u_count)
                    h_n = int(rec.h_n_count)
                    h_s = int(rec.h_s_count)
                    h_v = int(rec.h_v_count)
                    h_u = int(rec.h_u_count)
    
                def set_cell(col_key: str, new_val: Any) -> bool:
                    col = cols[col_key]
                    cell = ws.cell(row=row, column=col)
                    old_val = cell.value
                    if isinstance(new_val, float):
                        new_val = round(float(new_val), 1)
                    if cell_changed(old_val, new_val):
                        cell.value = new_val
                        return True
                    return False
    
                row_changed = False
    
                # tag: only set if AI_DATA present
                if ai_data:
                    row_changed |= set_cell("tag", 1111)
    
                # noni = h_nz_cnt
                row_changed |= set_cell("noni", int(h_nz_cnt))

                # Human counts (N/S/V/U)
                if has_human_counts:
                    row_changed |= set_cell("hN", int(h_n))
                    row_changed |= set_cell("hS", int(h_s))
                    row_changed |= set_cell("hV", int(h_v))
                    row_changed |= set_cell("hU", int(h_u))
    
                # ML counts
                row_changed |= set_cell("mlS", int(ml_s))
                row_changed |= set_cell("mlV", int(ml_v))
                row_changed |= set_cell("mlU", int(ml_u))
    
                # Human annotated noise
                row_changed |= set_cell("h_nz_cnt", int(h_nz_cnt))
                row_changed |= set_cell("h_nz_len", int(h_nz_samples))
                if h_frac is not None:
                    row_changed |= set_cell("h_nz_frac", float(h_frac))
    
                if row_changed:
                    changed_rows.add(row)
    
                processed += 1
    

    # -----------------------------------------------------------------
    # Preserve selected columns from the input workbook (data_only).
    # If these columns contain formulas, openpyxl won't recalculate them,
    # and cached results may be empty. To avoid blanks in the output,
    # we copy the displayed values from the data_only workbook.
    # -----------------------------------------------------------------
    preserve_cols = []
    for key in ("hN", "h_nz_frac"):
        c = cols.get(key)
        if c is not None:
            preserve_cols.append(c)

    if preserve_cols:
        for r in range(2, ws.max_row + 1):
            for c in preserve_cols:
                cell = ws.cell(row=r, column=c)
                v = cell.value
                is_formula = isinstance(v, str) and v.startswith("=")
                if v is None or is_formula:
                    v_cached = ws_values.cell(row=r, column=c).value
                    if v_cached is not None:
                        cell.value = v_cached

    # Highlight basename cell for changed rows
    for r in changed_rows:
        ws.cell(row=r, column=c_basename).fill = BLUE_FILL

    out_path = args.out if args.out else args.excel.with_name(args.excel.stem + "_updated.xlsx")
    wb.save(out_path)

    print(f"Processed JSON files: {processed}")
    print(f"Changed rows: {len(changed_rows)}")
    if not_found:
        print(f"Basenames not found in Excel: {len(not_found)}")
        ex = sorted(not_found)[:10]
        print("Examples:", ex)
    print(f"Saved: {out_path}")

if __name__ == "__main__":
    main()
