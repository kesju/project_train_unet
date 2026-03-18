#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
from decimal import Decimal, InvalidOperation
from pathlib import Path
from copy import copy

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def normalize_basename(x) -> str:
    """
    Canonicalize basename values so strings like:
      "1626933.710", "1626933.71", 1626933.71, "1626933,71"
    match each other.

    Returns a dot-decimal string without trailing zeros.
    """
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = str(x).strip().replace(",", ".")
    if not s:
        return ""
    try:
        # Decimal removes float-ish artifacts better than float->str sometimes.
        d = Decimal(s)
        # Normalize (removes trailing zeros), but keep plain string.
        s2 = format(d.normalize(), "f")
        # Edge: Decimal('1E+3') -> '1000'
        return s2
    except (InvalidOperation, ValueError):
        # Not numeric: keep as-is after comma->dot.
        return s


def read_basenames_list(txt_path: Path) -> list[str]:
    basenames: list[str] = []
    with txt_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            basenames.append(normalize_basename(line))
    # Deduplicate while preserving order
    seen = set()
    out = []
    for b in basenames:
        if b and b not in seen:
            seen.add(b)
            out.append(b)
    return out


def filter_source_rows(src_xlsx: Path, basenames: list[str]) -> pd.DataFrame:
    df = pd.read_excel(src_xlsx)

    if "basename" not in df.columns:
        raise KeyError(f"Source Excel has no 'basename' column: {src_xlsx}")

    df["_basename_key"] = df["basename"].map(normalize_basename)
    wanted = set(basenames)
    df_sel = df[df["_basename_key"].isin(wanted)].copy()
    df_sel.drop(columns=["_basename_key"], inplace=True)

    # Optional: keep same order as basenames list
    order_map = {b: i for i, b in enumerate(basenames)}
    df_sel["_ord"] = df_sel["basename"].map(normalize_basename).map(order_map).fillna(10**9).astype(int)
    df_sel.sort_values(["_ord"], inplace=True)
    df_sel.drop(columns=["_ord"], inplace=True)

    return df_sel


def autosize_columns(ws, max_rows_scan: int = 200, min_w: int = 8, max_w: int = 45) -> None:
    ncols = ws.max_column
    for c in range(1, ncols + 1):
        col_letter = get_column_letter(c)
        max_len = 0
        for r in range(1, min(ws.max_row, max_rows_scan) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(min_w, max_len + 2), max_w)


def build_output(
    template_xlsx: Path,
    out_xlsx: Path,
    df_sarasas: pd.DataFrame,
) -> None:
    # Load template to preserve formatting for "logas"
    wb_out = openpyxl.load_workbook(template_xlsx)
    ws_log = wb_out.active
    ws_log.title = "logas"

    # Load template again to capture a style row (row 2) for copying
    wb_tmp = openpyxl.load_workbook(template_xlsx)
    ws_tmp = wb_tmp.active
    style_row_idx = 2
    style_cells = [ws_tmp.cell(row=style_row_idx, column=c) for c in range(1, 14)]  # A..M

    # Clear data rows (keep header row 1)
    if ws_log.max_row > 1:
        ws_log.delete_rows(2, ws_log.max_row - 1)

    # Fill "logas" using template columns:
    # A nr
    # B filename   <- filename
    # C basename    <- basename
    # ...
    # K comment    <- comment
    # L recordingId      <- recordingId
    # M userId         <- userId
    required_cols = {"filename", "basename", "recordingId", "userId"}
    missing = [c for c in required_cols if c not in df_sarasas.columns]
    if missing:
        raise KeyError(f"Missing required columns in filtered data: {missing}")

    n = len(df_sarasas)
    for i in range(n):
        r = i + 2
        row = df_sarasas.iloc[i]

        ws_log.cell(row=r, column=1, value=i + 1)
        ws_log.cell(row=r, column=2, value=None if pd.isna(row.get("filename")) else str(row.get("filename")))
        ws_log.cell(row=r, column=3, value=normalize_basename(row.get("basename")))

        comment = row.get("comment") if "comment" in df_sarasas.columns else None
        ws_log.cell(row=r, column=11, value=None if pd.isna(comment) else str(comment))

        ws_log.cell(row=r, column=12, value=None if pd.isna(row.get("recordingId")) else str(row.get("recordingId")))
        ws_log.cell(row=r, column=13, value=None if pd.isna(row.get("userId")) else str(row.get("userId")))

        # Copy template styling for A..M
        for c in range(1, 14):
            src = style_cells[c - 1]
            tgt = ws_log.cell(row=r, column=c)
            tgt._style = copy(src._style)
            tgt.number_format = src.number_format
            tgt.protection = copy(src.protection)
            tgt.alignment = copy(src.alignment)
            tgt.font = copy(src.font)
            tgt.border = copy(src.border)
            tgt.fill = copy(src.fill)

    ws_log.freeze_panes = ws_tmp.freeze_panes
    ws_log.auto_filter.ref = "A1:M1"

    # Add "sarasas" and write the full filtered table
    if "sarasas" in wb_out.sheetnames:
        del wb_out["sarasas"]
    ws_sar = wb_out.create_sheet("sarasas")

    df_sar_out = df_sarasas.copy()
    if "basename" in df_sar_out.columns:
        df_sar_out["basename"] = df_sar_out["basename"].map(normalize_basename)

    for r_idx, excel_row in enumerate(dataframe_to_rows(df_sar_out, index=False, header=True), start=1):
        for c_idx, v in enumerate(excel_row, start=1):
            ws_sar.cell(row=r_idx, column=c_idx, value=v)


    # --- Format h_nz_frac and ml_nz_frac to 1 decimal + percent sign on 'sarasas' ---
    percent_cols = {"h_nz_frac", "ml_nz_frac"}

    # Map header name -> column index (1-based)
    header_to_col = {}
    for c in range(1, ws_sar.max_column + 1):
        v = ws_sar.cell(row=1, column=c).value
        if isinstance(v, str) and v.strip():
            header_to_col[v.strip()] = c

    for col_name in percent_cols:
        cidx = header_to_col.get(col_name)
        if not cidx:
            continue

        # Decide if values are stored as "percent units" (12.3 means 12.3%)
        # or as fractions (0.123 means 12.3%).
        vals = []
        for r in range(2, ws_sar.max_row + 1):
            v = ws_sar.cell(row=r, column=cidx).value
            if v is None:
                continue
            try:
                vals.append(float(v))
            except Exception:
                pass

        # Heuristic: if any non-null value > 1.5, treat as percent-units.
        is_percent_units = any(abs(x) > 1.5 for x in vals)

        for r in range(2, ws_sar.max_row + 1):
            cell = ws_sar.cell(row=r, column=cidx)
            if cell.value is None:
                continue
            try:
                x = float(cell.value)
            except Exception:
                continue

            if is_percent_units:
                # Keep numeric value (12.3) and show it as "12.3%"
                cell.value = x
                cell.number_format = '0.0"%"'
            else:
                # Convert fraction (0.123) to true percent (12.3%) and show as percent
                cell.value = x
                cell.number_format = "0.0%"

    # Header style borrowed from template A1 (simple, consistent)
    header_font = copy(ws_log["A1"].font)
    header_fill = copy(ws_log["A1"].fill)
    header_alignment = copy(ws_log["A1"].alignment)
    header_border = copy(ws_log["A1"].border)

    for c in range(1, df_sar_out.shape[1] + 1):
        cell = ws_sar.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    ws_sar.freeze_panes = "A2"
    ws_sar.auto_filter.ref = f"A1:{get_column_letter(df_sar_out.shape[1])}1"
    autosize_columns(ws_sar)
    
    # Make 'sarasas' the first sheet (leftmost) using public API
    ws_sar = wb_out["sarasas"]
    idx = wb_out.sheetnames.index("sarasas")   # current index (0-based)
    if idx != 0:
        wb_out.move_sheet(ws_sar, offset=-idx)  # move left by idx positions -> becomes index 0

    wb_out.save(out_xlsx)


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Create 06_Anotavimo_logas.xlsx (sarasas + logas) from basename list + source Excel."
    )
    ap.add_argument("--list-txt", type=Path, default=Path("06_Sarasas_anotavimui_testas.txt"))
    ap.add_argument("--src-xlsx", type=Path, default=Path("visi_zive_irasai_atrankai._modif_v1.xlsx"))
    ap.add_argument("--template-xlsx", type=Path, default=Path("05_Anotavimo_logas.xlsx"))
    ap.add_argument("--out-xlsx", type=Path, default=Path("06_Anotavimo_logas.xlsx"))

    args = ap.parse_args()

    basenames = read_basenames_list(args.list_txt)
    if not basenames:
        raise SystemExit(f"No basenames found in: {args.list_txt}")

    df_sel = filter_source_rows(args.src_xlsx, basenames)
    if df_sel.empty:
        raise SystemExit("No rows matched basenames list. Check basename formats/values.")

    build_output(args.template_xlsx, args.out_xlsx, df_sel)

    print(f"OK: wrote {args.out_xlsx} (rows={len(df_sel)})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
