#!/usr/bin/env python3
"""
Directory-only simplified version of update_records_list_adding_ml_noises_v1.py

Changes vs original:
- Source is ONLY a directory (no ZIP support)
- Column mapping simplified:
    h_nz_len: pick_col(hdr, "h_nz_len", "h_nz_length")
    ml_nz_len: pick_col(hdr, "ml_nz_len")
- h_nz_frac and ml_nz_frac% ALWAYS mean percentage (0..100), rounded to 1 decimal
"""

from __future__ import annotations

import argparse
import json
import math
import re
from copy import copy as _copy_style
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Set, Tuple, Union

import numpy as np
import openpyxl
from ecg_denoising_pipeline import load_ecg_npy
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension




JsonLikeRef = Union[str, Path]

# IMPORTANT: if your run_denoising_pipeline is in a different module, update the import below.
from record_noise_stats import calc_noise_stats_from_denoised_result

try:
    # <-- adjust this import to match your project structure if needed
    from ecg_denoising_pipeline import (
        ECGDenoisingPipeline,
        load_denoising_config_yaml,
        load_ecg_npy,
        DenoisingPipelineConfig,
        check_denoising_config,
        resolve_model_path
    )
except Exception as exc:
    raise ImportError(
        "Cannot import run_denoising_pipeline. Update the import to match your project.\n"
        f"Original error: {exc}"
    ) from exc

# If you have a config checker, import it; otherwise we just keep the cfg path.
try:
    from ecg_denoising_pipeline.steps import check_denoising_config
except Exception:
    check_denoising_config = None

from record_noise_stats import calc_noise_stats_from_denoised_result

try:
    from ecg_denoising_pipeline.steps import check_denoising_config
except Exception:
    check_denoising_config = None

try:
    from ecg_ectopy_pipeline import (
        ECGEctopyPipeline,
        EctopyPipelineConfig,
        EctopyPipelineResult,
        load_ectopy_config_yaml,
    )
except Exception as exc:
    raise ImportError(
        "Cannot import ECG ectopy pipeline helpers. "
        "Update imports to match your project structure.\n"
        f"Original error: {exc}"
    ) from exc



# ---------------------------------------------------------------------
# Embedded analyzer helpers
# ---------------------------------------------------------------------

@dataclass
class RecordSummary:
    sequenceId: str
    recordingId: Optional[str]
    basename: str
    channel_count: Optional[int]
    user_id: Optional[str]

    bin_bytes: int
    samples: int
    duration_s: Optional[float]

    flags_count: int
    flags: List[str]

    rpeaks_count: int
    h_n_count: int
    h_s_count: int
    h_v_count: int
    h_u_count: int

    ml_s_count: int
    ml_v_count: int
    ml_u_count: int

    json_ok: bool

    ml_noises_count: int
    ml_noises_samples: int
    ml_noises_fraction: Optional[float]  # percent 0..100

    h_noises_count: int
    h_noises_samples: int
    h_noises_fraction: Optional[float]  # percent 0..100

    cmt: Optional[str]
    json_keys_correct: bool
    notes: Optional[str] = None


def _flatten_flags(flags: Any) -> List[str]:
    if flags is None:
        return []
    if isinstance(flags, list):
        return [str(x) for x in flags]
    return [str(flags)]


def _pick_variant_dict(obj: Any, keys_preference: Tuple[str, ...] = ("merged", "human", "ml")) -> Any:
    if not isinstance(obj, dict):
        return None
    for k in keys_preference:
        if k in obj:
            return obj.get(k)
    for k, v in obj.items():
        if k == "__info":
            continue
        return v
    return None


def _extract_rpeaks_list(rpeaks: Any) -> List[Dict[str, Any]]:
    if isinstance(rpeaks, list):
        return [x for x in rpeaks if isinstance(x, dict)]
    if isinstance(rpeaks, dict):
        v = _pick_variant_dict(rpeaks, ("merged", "human", "ml"))
        if isinstance(v, list):
            return [x for x in v if isinstance(x, dict)]
    return []


def _extract_rpeaks_list_for_variant(rpeaks: Any, variant: str) -> List[Dict[str, Any]]:
    if isinstance(rpeaks, dict):
        v = rpeaks.get(variant)
        if isinstance(v, list):
            return [x for x in v if isinstance(x, dict)]
    return []


def _extract_mrk_counts_for_variant(h_counts: Any, variant: str) -> Dict[str, int]:
    if not isinstance(h_counts, dict):
        return {}
    v = h_counts.get(variant)
    if not isinstance(v, dict):
        return {}
    out: Dict[str, int] = {}
    for k, val in v.items():
        if k == "__info":
            continue
        if isinstance(val, int):
            out[str(k)] = int(val)
        elif isinstance(val, (float, np.floating)) and float(val).is_integer():
            out[str(k)] = int(val)
    return out


def _extract_mrk_counts(h_counts: Any) -> Dict[str, int]:
    if not isinstance(h_counts, dict):
        return {}
    # Variant dict?
    if any(isinstance(v, dict) for k, v in h_counts.items() if k != "__info"):
        v = _pick_variant_dict(h_counts, ("merged", "human", "ml"))
        if isinstance(v, dict):
            out: Dict[str, int] = {}
            for k, val in v.items():
                if k == "__info":
                    continue
                if isinstance(val, int):
                    out[str(k)] = int(val)
                elif isinstance(val, (float, np.floating)) and float(val).is_integer():
                    out[str(k)] = int(val)
            return out
        return {}
    # Flat dict
    out2: Dict[str, int] = {}
    for k, v in h_counts.items():
        if k == "__info":
            continue
        if isinstance(v, int):
            out2[str(k)] = int(v)
        elif isinstance(v, (float, np.floating)) and float(v).is_integer():
            out2[str(k)] = int(v)
    return out2


def _summarize_rpeaks(rpeaks: Any) -> Tuple[int, Optional[int], Optional[int], Dict[str, int]]:
    lst = _extract_rpeaks_list(rpeaks)
    if not lst:
        return 0, None, None, {}
    idxs: List[int] = []
    ann: Dict[str, int] = {}
    for rp in lst:
        si = rp.get("sampleIndex")
        av = rp.get("annotationValue")
        if isinstance(si, int):
            idxs.append(si)
        elif isinstance(si, (float, np.floating)) and float(si).is_integer():
            idxs.append(int(si))
        if isinstance(av, str):
            ann[av] = ann.get(av, 0) + 1
        elif av is not None:
            s = str(av)
            ann[s] = ann.get(s, 0) + 1
    if not idxs:
        return len(lst), None, None, ann
    idxs.sort()
    return len(lst), idxs[0], idxs[-1], ann


def _sum_noise_samples(noises: Any, fs: Optional[int]) -> Tuple[int, int]:
    # Variant dict?
    if isinstance(noises, dict):
        noises = _pick_variant_dict(noises, ("merged", "human", "ml"))
    if not isinstance(noises, list):
        return 0, 0

    cnt = 0
    total = 0
    for it in noises:
        a = b = None
        if isinstance(it, dict):
            a = it.get("startIndex", it.get("startSample"))
            b = it.get("endIndex", it.get("endSample"))
            if (a is None or b is None) and fs:
                t0 = it.get("startTime")
                t1 = it.get("endTime")
                if (
                    isinstance(t0, (int, float, np.floating))
                    and isinstance(t1, (int, float, np.floating))
                    and t1 > t0
                ):
                    a = int(round(float(t0) * fs))
                    b = int(round(float(t1) * fs))
        elif isinstance(it, (list, tuple)) and len(it) >= 2:
            a, b = it[0], it[1]

        if isinstance(a, (float, np.floating)) and float(a).is_integer():
            a = int(a)
        if isinstance(b, (float, np.floating)) and float(b).is_integer():
            b = int(b)

        if isinstance(a, int) and isinstance(b, int) and b > a:
            cnt += 1
            total += (b - a)
    return cnt, total


def load_json_fs(p: Path) -> Any:
    with open(p, "r", encoding="utf-8") as f:
        return json.load(f)


def file_size_fs(p: Path) -> int:
    return int(p.stat().st_size)


def sample_count_fs(p: Path, ext: str) -> int:
    ext = (ext or "").lower()
    if ext == ".npy":
        arr = np.load(p, allow_pickle=False)
        return int(getattr(arr, "size", 0))
    # Default: assume 4 bytes per sample (int32)
    return int(file_size_fs(p) // 4)


def summarize_record(
    sequenceId: str,
    basename: str,
    json_ref: Path,
    data_ref: Path,
    data_ext: str,
    *,
    load_json_fn: Callable[[Path], Any],
    file_size_fn: Callable[[Path], int],
    sample_count_fn: Callable[[Path, str], int],
    fs: int,
) -> Tuple[RecordSummary, Dict[str, Any]]:
    try:
        meta = load_json_fn(json_ref)
        json_ok = True
    except Exception as exc:
        print(f"WARN: failed to load JSON for {sequenceId}/{basename}: {exc}")
        meta = {}
        json_ok = False

    bsz = int(file_size_fn(data_ref))
    samples = int(sample_count_fn(data_ref, data_ext))

    channel_count = meta.get("channelCount") if isinstance(meta, dict) else None
    user_id = meta.get("userId") if isinstance(meta, dict) else None
    recordingId = meta.get("recordingId") if isinstance(meta, dict) else None
    flags_list = _flatten_flags(meta.get("flags") if isinstance(meta, dict) else None)
    comment = meta.get("comment") if isinstance(meta, dict) else None

    rpeaks_any = meta.get("rpeaks") if isinstance(meta, dict) else None
    rpeaks_count, *_ = _summarize_rpeaks(rpeaks_any)

    mrk_counts_any = meta.get("rpeakAnnotationCounts") if isinstance(meta, dict) else None

    # H counts preference: merged -> human -> derived -> flat
    def _counts_for_h_columns() -> Tuple[Dict[str, int], bool]:
        for var in ("merged", "human"):
            d = _extract_mrk_counts_for_variant(mrk_counts_any, var)
            if d:
                return d, True

        for var in ("merged", "human"):
            rp_list = _extract_rpeaks_list_for_variant(rpeaks_any, var)
            if rp_list:
                tmp: Dict[str, int] = {}
                for rp in rp_list:
                    av = rp.get("annotationValue")
                    if isinstance(av, str):
                        tmp[av] = tmp.get(av, 0) + 1
                    elif av is not None:
                        tmp[str(av)] = tmp.get(str(av), 0) + 1
                if tmp:
                    return tmp, True

        flat = _extract_mrk_counts(mrk_counts_any)
        if flat:
            return flat, True

        return {}, False

    h_counts_primary, has_h_counts = _counts_for_h_columns()
    h_n = int(h_counts_primary.get("N", 0)) if has_h_counts else 0
    h_s = int(h_counts_primary.get("S", 0)) if has_h_counts else 0
    h_v = int(h_counts_primary.get("V", 0)) if has_h_counts else 0
    h_u = int(h_counts_primary.get("U", 0)) if has_h_counts else 0

    # ML counts: explicit "ml" first, else derive from ml rpeaks
    ml_counts = _extract_mrk_counts_for_variant(mrk_counts_any, "ml")
    if not ml_counts:
        ml_rpeaks = _extract_rpeaks_list_for_variant(rpeaks_any, "ml")
        tmp2: Dict[str, int] = {}
        for rp in ml_rpeaks:
            av = rp.get("annotationValue")
            if isinstance(av, str):
                tmp2[av] = tmp2.get(av, 0) + 1
            elif av is not None:
                tmp2[str(av)] = tmp2.get(str(av), 0) + 1
        ml_counts = tmp2

    ml_s = int(ml_counts.get("S", 0))
    ml_v = int(ml_counts.get("V", 0))
    ml_u = int(ml_counts.get("U", 0))

    h_noises = meta.get("noises_annotated") if isinstance(meta, dict) else None
    h_nz_cnt, h_nz_samples = _sum_noise_samples(h_noises, fs)

    ml_noises = meta.get("noises") if isinstance(meta, dict) else None
    ml_nz_cnt, ml_nz_samples = _sum_noise_samples(ml_noises, fs)

    duration_s = (samples / fs) if (fs and fs > 0 and samples > 0) else None

    # IMPORTANT: fractions are ALWAYS percentages (0..100)
    ml_noises_fraction = (ml_nz_samples / samples) * 100.0 if samples > 0 else None
    h_noises_fraction = (h_nz_samples / samples) * 100.0 if samples > 0 else None

    allowed_keys = {
        "channelCount",
        "comment",
        "flags",
        "noises",
        "noises_annotated",
        "recordingId",
        "rpeakAnnotationCounts",
        "rpeaks",
        "userId",
    }
    meta_keys = set(meta.keys()) if isinstance(meta, dict) else set()
    json_keys_correct = meta_keys.issubset(allowed_keys) if meta_keys else False

    rec = RecordSummary(
        sequenceId=str(sequenceId),
        recordingId=str(recordingId) if isinstance(recordingId, str) else None,
        basename=str(basename),
        channel_count=int(channel_count) if isinstance(channel_count, int) else None,
        user_id=str(user_id) if isinstance(user_id, str) else None,
        bin_bytes=bsz,
        samples=samples,
        duration_s=float(duration_s) if duration_s is not None else None,
        cmt=str(comment) if isinstance(comment, str) else None,
        flags_count=len(flags_list),
        flags=flags_list,
        rpeaks_count=int(rpeaks_count),
        h_n_count=h_n,
        h_s_count=h_s,
        h_v_count=h_v,
        h_u_count=h_u,
        ml_s_count=int(ml_s),
        ml_v_count=int(ml_v),
        ml_u_count=int(ml_u),
        json_ok=bool(json_ok),
        ml_noises_count=int(ml_nz_cnt),
        ml_noises_samples=int(ml_nz_samples),
        ml_noises_fraction=float(ml_noises_fraction) if ml_noises_fraction is not None else None,
        h_noises_count=int(h_nz_cnt),
        h_noises_samples=int(h_nz_samples),
        h_noises_fraction=float(h_noises_fraction) if h_noises_fraction is not None else None,
        json_keys_correct=bool(json_keys_correct),
    )

    rec_dict = asdict(rec)
    rec_dict["_ml_noises_samples"] = int(ml_nz_samples)
    rec_dict["_h_noises_samples"] = int(h_nz_samples)
    return rec, rec_dict


def infer_sequence_id_fs(json_path: Path, fallback: str) -> str:
    parts = list(json_path.parts)
    if "recordings" in parts:
        idx = parts.index("recordings")
        if idx >= 1:
            return parts[idx - 1]
    if len(parts) >= 2:
        return parts[-2]
    return fallback


def find_data_file_fs(json_path: Path) -> Optional[Tuple[Path, str]]:
    if json_path.suffix.lower() != ".json":
        return None
    rec_dir = json_path.parent
    stem = json_path.stem

    exact = rec_dir / stem
    if exact.exists():
        suf = exact.suffix
        if suf and suf[1:].isdigit() and len(suf) == 4:
            return exact, suf

    candidates = [
        rec_dir / f"{stem}.npy",
        rec_dir / f"{stem}.bin",
        rec_dir / f"{stem}",
    ]
    for p in sorted(rec_dir.glob(f"{stem}.*")):
        suf = p.suffix
        if suf and suf[1:].isdigit() and len(suf) == 4:
            candidates.append(p)

    for c in candidates:
        if c.exists():
            return c, c.suffix
    return None


# ---------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------

LIGHT_RED_FILL = PatternFill(patternType="solid", fgColor="FFFFC7CE")  # new rows
MEDIUM_BLUE_FILL = PatternFill(patternType="solid", fgColor="FF1E90FF")  # changed existing rows


def norm_basename(x: Any) -> Optional[str]:
    """Normalize basenames for matching: ',' -> '.', strip, and drop trailing zeros in decimals."""
    if x is None:
        return None
    s = str(x).strip()
    if not s:
        return None
    s = s.replace(",", ".")
    m = re.fullmatch(r"(\d+)\.(\d+)", s)
    if m:
        a, b = m.group(1), m.group(2).rstrip("0")
        return a if b == "" else f"{a}.{b}"
    return s


def build_header_map(ws) -> Dict[str, int]:
    hdr: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        key = str(v).strip().lower()
        if key:
            hdr[key] = col
    return hdr


def pick_col(hdr: Dict[str, int], *names: str) -> Optional[int]:
    for n in names:
        k = n.strip().lower()
        if k in hdr:
            return hdr[k]
    return None


def cell_changed(old: Any, new: Any) -> bool:
    """Loose compare to avoid false mismatches (e.g. '12,3' vs 12.3)."""
    def norm(v: Any) -> Any:
        if v is None:
            return None
        if isinstance(v, bool):
            return bool(v)
        if isinstance(v, int):
            return int(v)
        if isinstance(v, float):
            if math.isnan(v):
                return None
            return float(v)
        s = str(v).strip()
        try:
            return float(s.replace(",", "."))
        except Exception:
            return s
    return norm(old) != norm(new)


def _flags_to_cell_value(flags_list: List[str]) -> str:
    return "; ".join([f for f in flags_list if f])



def _split_column_dimension_if_needed(ws, col_idx: int) -> None:
    """Ensure this column has its own ColumnDimension, not a shared grouped range.

    This prevents hiding one column (e.g. mlU) from also hiding neighboring columns
    that happen to share the same dimension range in the source workbook.
    """
    target_letter = get_column_letter(col_idx)

    # Find a dimension range that covers the target column.
    for key, dim in list(ws.column_dimensions.items()):
        min_idx = getattr(dim, "min", None)
        max_idx = getattr(dim, "max", None)
        if min_idx is None or max_idx is None:
            continue
        if not (min_idx <= col_idx <= max_idx):
            continue
        if min_idx == max_idx == col_idx:
            return

        # Snapshot current formatting/behavior of the shared range.
        width = dim.width
        hidden = dim.hidden
        best_fit = dim.bestFit
        collapsed = dim.collapsed
        outline_level = dim.outline_level
        custom_width = dim.customWidth

        # Left part stays on the original dimension key.
        if min_idx < col_idx:
            dim.min = min_idx
            dim.max = col_idx - 1
        else:
            # No left part remains: remove the original grouped dimension.
            del ws.column_dimensions[key]

        # Create a dedicated dimension for the target column.
        target_dim = ColumnDimension(ws, min=col_idx, max=col_idx, width=width)
        target_dim.hidden = hidden
        target_dim.bestFit = best_fit
        target_dim.collapsed = collapsed
        target_dim.outline_level = outline_level
        if custom_width:
            target_dim.width = width
        ws.column_dimensions[target_letter] = target_dim

        # Right part, if any, becomes its own grouped dimension.
        if col_idx < max_idx:
            right_start = col_idx + 1
            right_letter = get_column_letter(right_start)
            right_dim = ColumnDimension(ws, min=right_start, max=max_idx, width=width)
            right_dim.hidden = hidden
            right_dim.bestFit = best_fit
            right_dim.collapsed = collapsed
            right_dim.outline_level = outline_level
            if custom_width:
                right_dim.width = width
            ws.column_dimensions[right_letter] = right_dim
        return


def hide_columns_by_keys(ws, cols: dict, *keys: str) -> None:
    for key in keys:
        col_idx = cols.get(key) or cols.get(str(key).strip().lower())
        if col_idx:
            _split_column_dimension_if_needed(ws, col_idx)
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].hidden = True


def prepare_denoising_pipeline(
    denoising_config_path: Path,
    denoising_model_dir: Path,
    disable_motions: bool = False,
) -> DenoisingPipelineConfig:
    if not denoising_config_path.exists():
        raise FileNotFoundError(f"Config file not found: {denoising_config_path}")
    if not denoising_model_dir.exists():
        raise FileNotFoundError(f"Model directory not found: {denoising_model_dir}")

    cfg = load_denoising_config_yaml(str(denoising_config_path))
    if check_denoising_config is not None:
        check_denoising_config(cfg)

    cfg.motions.model_name = resolve_model_path(denoising_model_dir, cfg.motions.model_name)
    cfg.motions.enabled = not disable_motions
    return cfg


def create_unet_marker(cfg_denoising: DenoisingPipelineConfig) -> str:
    """Create a marker string based on the UNet model name and threshold for output column naming.""" 
    unet_model_name = Path(cfg_denoising.motions.model_name).name
    # print("\nunet_model_name:", unet_model_name, type(unet_model_name))
    MARKER = unet_model_name.removeprefix("resunet_ecg").removesuffix(".keras")  # -> "_1024_0_5_3_7"
    threshold = cfg_denoising.motions.threshold
    threshold_str = str( threshold).replace('.', '_')
    MARKER += f"_{threshold_str}"
    return MARKER


def prepare_ectopy_pipeline(
    ectopy_config_path: Path,
    ectopy_model_dir: Path,
) -> EctopyPipelineConfig:
    
    """Entry point when the ectopy pipeline is used ad-hoc (e.g. in notebooks)."""
    
    cfg: EctopyPipelineConfig = load_ectopy_config_yaml(ectopy_config_path)

    cfg.ectopy.model_name = resolve_model_path(ectopy_model_dir, cfg.ectopy.model_name)
    cfg.ectopy.scaler_name = resolve_model_path(ectopy_model_dir, cfg.ectopy.scaler_name)

    cfg.ectopy.ectopy_removing = False

    # print("[ECTOPY] Model:", cfg.ectopy.model_name)
    # print("[ECTOPY] Scaler:", cfg.ectopy.scaler_name)
    # print("[ECTOPY] Ectopy removing enabled:", bool(cfg.ectopy.ectopy_removing))

    return cfg

def compute_ectopy_stats(res_ectopy: EctopyPipelineResult | None) -> Dict[str, int]:
    """
    Count ectopic beat classes from rpeaks_on_denoised_df.

    Expected mapping in column 'pred':
        0 -> ectN
        1 -> ectS
        2 -> ectV
        3 -> ectU
        other values are ignored
    """
    stats = {"ectN": 0, "ectS": 0, "ectV": 0, "ectU": 0}

    if res_ectopy is None:
        return stats

    rpeaks_on_denoised_df = res_ectopy.rpeaks_on_denoised_df
    if rpeaks_on_denoised_df is None:
        return stats

    if "pred" not in rpeaks_on_denoised_df.columns:
        return stats

    counts = rpeaks_on_denoised_df["pred"].value_counts()

    return {
        "ectN": int(counts.get(0, 0)),
        "ectS": int(counts.get(1, 0)),
        "ectV": int(counts.get(2, 0)),
        "ectU": int(counts.get(3, 0)),
    }


def auto_adjust_widths(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        col_idx = None
        for cell in column_cells:
            col_idx = cell.column
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        if col_idx is not None:
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 80)

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", type=Path, required=True, help="Input Excel (.xlsx)")
    ap.add_argument("--dir", type=Path, required=True, help="Directory containing JSON metadata and data files")
    ap.add_argument("--fs", type=int, default=200, help="Sampling frequency (Hz). Default: 200")
    ap.add_argument("--sheet", type=str, default=None, help="Sheet name (default: active sheet)")
    ap.add_argument("--out", type=Path, default=None, help="Output Excel path (default: <excel>_updated.xlsx)")
   # NEW: denoising args (prepared once)
    ap.add_argument("--cfg-denoising", type=Path, required=True, help="Denoising config path")
    ap.add_argument("--unet-model-dir", type=Path, required=True, help="Model directory for denoising/motions")
    ap.add_argument("--disable-motions", action="store_true", help="Disable motions stage in denoising pipeline")
    ap.add_argument("--cfg-ectopy", type=Path, required=True, help="Ectopy config path")
    ap.add_argument("--ectopy-model-dir", type=Path, required=True, help="Model directory for ectopy detection")
    ap.add_argument("--quiet", action="store_true", help="Silence stdout during denoising/stats")
    ap.add_argument("--denoising", action="store_true", help="Silence stdout during denoising/stats")
   
   
    args = ap.parse_args()

    # Open Excel workbook and prepare for updates  ++++++++++++++++++++++++++++++
    
    src = args.dir
    if not src.exists() or not src.is_dir():
        raise FileNotFoundError(f"--dir must be an existing directory. Got: {src}")

    wb = openpyxl.load_workbook(args.excel)
    ws = wb[args.sheet] if args.sheet else wb.active

    # create header mapping
    hdr = build_header_map(ws)

    cols: Dict[str, Optional[int]] = {}
    cols["filename"] = pick_col(hdr, "filename", "file", "path")
    cols["basename"] = pick_col(hdr, "basename")
    cols["rec_id"] = pick_col(hdr, "rec_id", "recordingid", "recording_id", "recording id")
    cols["uid"] = pick_col(hdr, "uid", "user_id", "userid", "user id")
    cols["samples"] = pick_col(hdr, "samples", "sample_cnt", "sample_count")
    cols["cmt"] = pick_col(hdr, "cmt", "comment")
    cols["dur_s"] = pick_col(hdr, "dur_s", "duration_s", "duration", "dur")
    cols["rpk_cnt"] = pick_col(hdr, "rpk_cnt", "rpeaks_cnt", "rpeaks_count", "rpeaks")
    cols["hN"] = pick_col(hdr, "hn", "hN")
    cols["hS"] = pick_col(hdr, "hs", "hS")
    cols["hV"] = pick_col(hdr, "hv", "hV")
    cols["hU"] = pick_col(hdr, "hu", "hU")
    cols["mlS"] = pick_col(hdr, "mls", "mlS")
    cols["mlV"] = pick_col(hdr, "mlv", "mlV")
    cols["mlU"] = pick_col(hdr, "mlu", "mlU")
    cols["ectN"] = pick_col(hdr, "ectn", "ectN")
    cols["ectS"] = pick_col(hdr, "ects", "ectS")
    cols["ectV"] = pick_col(hdr, "ectv", "ectV")
    cols["ectU"] = pick_col(hdr, "ectu", "ectU")
    cols["h_nz_cnt"] = pick_col(hdr, "h_nz_cnt", "h_nz_count")

    # requested simplifications
    cols["h_nz_len"] = pick_col(hdr, "h_nz_len", "h_nz_length")
    cols["h_nz_frac"] = pick_col(hdr, "h_nz_frac", "h_nz_frac%", "h_nz_frac %")
    cols["out"] = pick_col(hdr, "out")
    cols["rdr"] = pick_col(hdr, "rdr")
    cols["noi"] = pick_col(hdr, "noi")
    cols["tp_pct"] = pick_col(hdr, "tp%", "tp %", "tp_pct", "tp")
    cols["ml_nz_cnt"] = pick_col(hdr, "ml_nz_cnt", "ml_nz_count")
    cols["ml_nz_len"] = pick_col(hdr, "ml_nz_len")
    cols["ml_nz_frac%"] = pick_col(hdr, "ml_nz_frac%", "ml_nz_frac", "ml_nz_frac %")
    cols["flags"] = pick_col(hdr, "flags")

    if cols["basename"] is None:
        raise RuntimeError("Missing required column in Excel header row: ['basename']")

    c_basename = int(cols["basename"])
    c_filename_raw = cols.get("filename")
    if c_filename_raw is None:
        raise ValueError("Required column 'filename' not found in Excel header")
    c_filename = int(c_filename_raw)
    
    # Map basename -> row index
    basename_to_row: Dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        bn = norm_basename(ws.cell(row=r, column=c_basename).value)
        if bn:
            basename_to_row[bn] = r

    # Copy formatting template from row 2
    style_row_idx = 2 if ws.max_row >= 2 else None

    def copy_row_style(dst_row: int) -> None:
        if style_row_idx is None:
            return
        for c in range(1, ws.max_column + 1):
            src_cell = ws.cell(row=style_row_idx, column=c)
            dst_cell = ws.cell(row=dst_row, column=c)
            dst_cell._style = _copy_style(src_cell._style)
            dst_cell.font = _copy_style(src_cell.font)
            dst_cell.border = _copy_style(src_cell.border)
            dst_cell.fill = _copy_style(src_cell.fill)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = _copy_style(src_cell.protection)
            dst_cell.alignment = _copy_style(src_cell.alignment)

    new_rows: Set[int] = set()
    changed_rows: Set[int] = set()

    def get_or_create_row(bn: str) -> Tuple[int, bool]:
        row = basename_to_row.get(bn)
        if row is not None:
            return row, False

        row = ws.max_row + 1
        copy_row_style(row)
        ws.cell(row=row, column=c_basename).value = bn
        basename_to_row[bn] = row
        new_rows.add(row)
        ws.cell(row=row, column=c_basename).fill = LIGHT_RED_FILL
        return row, True

    def set_cell(row: int, col_key: str, new_val: Any) -> bool:
        col = cols.get(col_key)
        if col is None:
            return False
        col_i = int(col)
        cell = ws.cell(row=row, column=col_i)

        # percent columns always mean 0..100
        if col_key in ("h_nz_frac", "ml_nz_frac%", "tp_pct") and isinstance(new_val, (int, float, np.floating)):
            new_val = round(float(new_val), 1)

        if cell_changed(cell.value, new_val):
            cell.value = new_val
            return True
        return False

    json_paths = [p for p in src.rglob("*.json") if "__MACOSX" not in p.parts and p.is_file()]
    print(f"Source: {src}")
    print(f"Found JSON files: {len(json_paths)}")

    # print(json_paths)
    
        # PREPARE DENOISING PART 
    
    if (not args.denoising):
        print("\nDenoising is disabled. Noise stats columns will be left empty.")
        MARKER = ""
        noise_stats = {}
        denoising_model_dir = "Not used"
        cfg_denoising = "Not used"
        cfg_ectopy = "Not used"
        ectopy_pipe = None
        denoising_pipe = None
        denoising_pipe_disabled_motions = None
    else:
        print("\nDenoising will be performed for each record using the specified config and model.")
    
        # prepare the denoising pipeline (if needed) ++++++++++++++++++++++++++++++++++++++
        
        #  Preparing the Denoising pipeline with configuration and model paths
        print("\n*****Denoising pipeline config:")

        cfg_denoising_path = args.cfg_denoising
        denoising_model_dir = args.unet_model_dir
        print(f"\nDenoising config path: {cfg_denoising_path}")
        print(f"Denoising model directory: {denoising_model_dir}")
        
        cfg_denoising = prepare_denoising_pipeline(cfg_denoising_path, denoising_model_dir)
        MARKER = create_unet_marker(cfg_denoising)
        print(f"\nUNet marker: {MARKER}\n")
        denoising_pipe = ECGDenoisingPipeline(cfg_denoising)
        
        cfg_denoising_disabled_motions = prepare_denoising_pipeline(
        denoising_config_path=cfg_denoising_path,
        denoising_model_dir=denoising_model_dir,
        disable_motions=True,
        # disable_motions=args.disable_motions,
        )
        denoising_pipe_disabled_motions = ECGDenoisingPipeline(cfg_denoising_disabled_motions)
        
        # PREPARE ECTOPY DETECTION PART 
        
        cfg_ectopy = prepare_ectopy_pipeline(
            ectopy_config_path=args.cfg_ectopy,
            ectopy_model_dir=args.ectopy_model_dir
        )
        ectopy_pipe = ECGEctopyPipeline(cfg_ectopy)
        
        # CYCLE through JSON and data files and update Excel rows 
    
    processed = 0
    for jp in sorted(json_paths):
        bn = norm_basename(jp.stem)
        if not bn:
            continue
        if bn.lower() == "manifest":
            continue

        row, _is_new = get_or_create_row(bn)
        data_filename = ws.cell(row=row, column=c_filename).value
        print(f"\nProcessing: {bn} ({data_filename}) -> row {row} (new: {_is_new})")

        data_info = find_data_file_fs(jp)
        if data_info is None:
            # JSON exists but data missing -> still fill what we can from JSON only
            meta = load_json_fs(jp)
            if not isinstance(meta, dict):
                meta = {}
            # Minimal extraction in this simplified version when data file missing
            # (If you still want the full JSON-only extraction logic, tell me and I’ll add it back.)
            extracted = {
                "rec_id": meta.get("recordingId") if isinstance(meta.get("recordingId"), str) else None,
                "uid": meta.get("userId") if isinstance(meta.get("userId"), str) else None,
                "flags": _flags_to_cell_value(_flatten_flags(meta.get("flags"))),
            }
        else:
            data_path, data_ext = data_info
            seq = infer_sequence_id_fs(jp, fallback=src.name or "dir")

            noise_stats: Dict[str, Any] = {}
            ectopy_stats: Dict[str, Any] = {"ectN": 0, "ectS": 0, "ectV": 0, "ectU": 0}
            
            if (args.denoising and denoising_pipe is not None):
                try:
                     # With activated motions stage:
                    x = load_ecg_npy(data_path)
                    
                    # denoising and getting noise stats
                    res_denoising = denoising_pipe.run(x, gaps_indices=[])
                    noise_stats = calc_noise_stats_from_denoised_result(res_denoising) or {}
                    print(f"Noise stats for {data_path.name}: out={noise_stats['out']}, rdr={noise_stats['rdr']}, noi={noise_stats['noi']}, tp_pct={noise_stats['tp_pct']:.1f} with enabled detecting motions")
                    print()
                    
                    # Without detecting motions stage (for comparison/debugging):
                    x = load_ecg_npy(data_path)
                    
                    # denoising and getting noise stats
                    res_denoising_disabled_motions = denoising_pipe_disabled_motions.run(x, gaps_indices=[])
                    noise_stats_disabled_motions = calc_noise_stats_from_denoised_result(res_denoising_disabled_motions) or {}
                    print(f"Noise stats for {data_path.name}: out={noise_stats_disabled_motions['out']}, rdr={noise_stats_disabled_motions['rdr']}, noi={noise_stats_disabled_motions['noi']}, tp_pct={noise_stats_disabled_motions['tp_pct']:.1f} with disabled detecting motions")
                    
                    # detecting ectopies and getting ectopy stats (without  detecting motions):
                    res_ectopy = ectopy_pipe.run(res_denoising_disabled_motions, fs=args.fs)
                    # print(f"DEBUG: Ectopy stats for {data_path.name}: {res_ectopy}")
                    ectopy_stats = compute_ectopy_stats(res_ectopy)
                    print(f"ectopy_stats={ectopy_stats} with disabled detecting motions")
                
                except Exception as exc:
                    print(f"WARN: failed denoising/stat extraction for {data_path.name}: {exc}")
                    noise_stats = {}
                    ectopy_stats = {"ectN": 0, "ectS": 0, "ectV": 0, "ectU": 0}

            rec, _rec_dict = summarize_record(
                sequenceId=seq,
                basename=bn,
                json_ref=jp,
                data_ref=data_path,
                data_ext=data_ext,
                load_json_fn=load_json_fs,
                file_size_fn=file_size_fs,
                sample_count_fn=sample_count_fs,
                fs=args.fs,
            )
            extracted = {
                "rec_id": rec.recordingId,
                "uid": rec.user_id,
                "samples": int(rec.samples),
                "cmt": rec.cmt,
                "dur_s": float(rec.duration_s) if rec.duration_s is not None else None,
                "rpk_cnt": int(rec.rpeaks_count),
                "hN": int(rec.h_n_count),
                "hS": int(rec.h_s_count),
                "hV": int(rec.h_v_count),
                "hU": int(rec.h_u_count),
                "mlS": int(rec.ml_s_count),
                "mlV": int(rec.ml_v_count),
                "mlU": int(rec.ml_u_count),
                "ectN": int(ectopy_stats.get("ectN", 0)),
                "ectS": int(ectopy_stats.get("ectS", 0)),
                "ectV": int(ectopy_stats.get("ectV", 0)),
                "ectU": int(ectopy_stats.get("ectU", 0)),
                "h_nz_cnt": int(rec.h_noises_count),
                "h_nz_len": int(_rec_dict.get("_h_noises_samples", 0)),
                "h_nz_frac": float(rec.h_noises_fraction) if rec.h_noises_fraction is not None else None,  # percent
                "ml_nz_cnt": int(rec.ml_noises_count),
                "ml_nz_len": int(_rec_dict.get("_ml_noises_samples", 0)),
                "ml_nz_frac%": float(rec.ml_noises_fraction) if rec.ml_noises_fraction is not None else None,  # percent
                "flags": _flags_to_cell_value(rec.flags or []),
                "out": noise_stats.get("out"),
                "rdr": noise_stats.get("rdr"),
                "noi": noise_stats.get("noi"),
                "tp_pct": noise_stats.get("tp_pct"),
            }

        row_changed = False
        for k, v in extracted.items():
            if v is None:
                continue
            row_changed |= set_cell(row, k, v)

        if row_changed:
            changed_rows.add(row)

        processed += 1

    for r in changed_rows:
        if r in new_rows:
            continue
        ws.cell(row=r, column=c_basename).fill = MEDIUM_BLUE_FILL

    # Optional: hide temporarely the columns if they exist
    hide_columns_by_keys(ws, cols, "ml_nz_cnt", "ml_nz_len", "ml_nz_frac%", "mlS", "mlV", "mlU")

    # Ensure ectopy summary columns stay clearly visible
    for key in ("ectN", "ectS", "ectV", "ectU"):
        col_idx = cols.get(key) or cols.get(key.lower())
        if col_idx:
            _split_column_dimension_if_needed(ws, col_idx)
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].hidden = False
            ws.column_dimensions[col_letter].width = 8

    # Save the updated workbook with a new name (original name + "_updated" + marker)
    out_path = args.out if args.out else args.excel.with_name(args.excel.stem + "_updated" + MARKER + ".xlsx")


# Save the parameters used for denoising and ectopy detection in a new sheet called "Parameters"
    ws_params = wb["Parameters"] if "Parameters" in wb.sheetnames else wb.create_sheet(title="Parameters")
    # denoising model info:
    ws_params["A1"] = "Unet model dir:"
    ws_params["B1"] = str(denoising_model_dir)
    ws_params["A2"] = "Unet model:"
    ws_params["B2"] = "Not used" if cfg_denoising == "Not used" else Path(cfg_denoising.motions.model_name).name
    ws_params["A3"] = "threshold:"
    ws_params["B3"] = "Not used" if cfg_denoising == "Not used" else cfg_denoising.motions.threshold

    # ectopy model info:
    ws_params["A5"] = "Ectopy model dir:"
    ws_params["B5"] = "Not used" if cfg_ectopy == "Not used" else str(args.ectopy_model_dir)
    ws_params["A6"] = "Ectopy model:"
    ws_params["B6"] = "Not used" if cfg_ectopy == "Not used" else Path(cfg_ectopy.ectopy.model_name).name

    ws_params["A7"] = "Ectopy scaler:"
    ws_params["B7"] = "Not used" if cfg_ectopy == "Not used" else Path(cfg_ectopy.ectopy.scaler_name).name

    # Bolding the parameter labels in column A
    for cell_ref in ("A1", "A2", "A3", "A5", "A6", "A7"):
        ws_params[cell_ref].font = Font(bold=True)

    auto_adjust_widths(ws_params)

    wb.save(out_path)
   
    print(f"Processed JSON files: {processed}")
    print(f"Changed rows: {len(changed_rows)}")
    print(f"Added new rows: {len(new_rows)}")
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()