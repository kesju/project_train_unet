#!/usr/bin/env python3
"""
Directory-only simplified version of update_records_list_adding_ml_noises_v1.py

python update_records_list_adding_ml_noises_v3_denoising.py \
  --excel /path/to/visi_zive_irasai_atrankai.xlsx \
  --dir /path/to/AtsisiuntimasZiveDuomenu/DuomenysTestui \
  --cfg-denoising /path/to/denoising_config.yaml \
  --model-dir /path/to/MODEL_UNET \
  --fs 200 \
  --quiet
  
MODIFICATION (per request):
- Instead of writing ml_nz_cnt and ml_nz_len, write:
    out, rdr, noi  (from stats = calc_noise_stats_from_result(res_denoising))
- Set ml_nz_frac% = tp% (tp_pct from stats)
- Preparation for denoising (paths/config/model loading checks) is done ONCE before main loop
- In main loop: read ECG -> run denoising -> calc stats -> write to Excel

Notes:
- This script assumes ECG data is available as .npy (preferred).
- For non-.npy files it will attempt to read int32 big-endian as a fallback.
  If your binary scaling is different (e.g., convert to mV), adjust READ_SCALE below.
"""

from __future__ import annotations

import argparse
import contextlib
import io
import json
import math
import re
from copy import copy as _copy_style
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Set, Tuple, Union

from ecg_denoising_pipeline import DenoisingPipelineConfig, ECGDenoisingPipeline, run_denoising_pipeline
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill

JsonLikeRef = Union[str, Path]

# ---------------------------------------------------------------------
# NEW: denoising + noise stats imports
# ---------------------------------------------------------------------
# These are based on the filenames you mentioned:
# - record_noise_stats.py contains calc_noise_stats_from_result
# - denoising_pipeline_demo_2.ipynb shows how run_denoising_pipeline is called
#
# IMPORTANT: if your run_denoising_pipeline is in a different module, update the import below.
from record_noise_stats import calc_noise_stats_from_result

try:
    # <-- adjust this import to match your project structure if needed
    from ecg_denoising_pipeline import (
        run_denoising_pipeline,
        load_denoising_config_yaml,
        load_ecg_npy,
        DenoisingPipelineConfig,
        check_denoising_config,
        resolve_model_path
    )
except Exception as exc:
    raise ImportError(
        "Cannot import run_denoising_pipeline. Update the import to match your project.\n"
        f"Original error: {exc}"
    ) from exc

# If you have a config checker, import it; otherwise we just keep the cfg path.
try:
    from ecg_denoising_pipeline.steps import check_denoising_config
except Exception:
    check_denoising_config = None


# ---------------------------------------------------------------------
# Embedded analyzer helpers (unchanged)
# ---------------------------------------------------------------------

@dataclass
class RecordSummary:
    sequenceId: str
    recordingId: Optional[str]
    basename: str
    channel_count: Optional[int]
    user_id: Optional[str]

    bin_bytes: int
    samples: int
    duration_s: Optional[float]

    flags_count: int
    flags: List[str]

    rpeaks_count: int
    h_n_count: int
    h_s_count: int
    h_v_count: int
    h_u_count: int

    ml_s_count: int
    ml_v_count: int
    ml_u_count: int

    json_ok: bool

    ml_noises_count: int
    ml_noises_samples: int
    ml_noises_fraction: Optional[float]  # percent 0..100

    h_noises_count: int
    h_noises_samples: int
    h_noises_fraction: Optional[float]  # percent 0..100

    cmt: Optional[str]
    json_keys_correct: bool
    notes: Optional[str] = None


def _flatten_flags(flags: Any) -> List[str]:
    if flags is None:
        return []
    if isinstance(flags, list):
        return [str(x) for x in flags]
    return [str(flags)]


def _pick_variant_dict(obj: Any, keys_preference: Tuple[str, ...] = ("merged", "human", "ml")) -> Any:
    if not isinstance(obj, dict):
        return None
    for k in keys_preference:
        if k in obj:
            return obj.get(k)
    for k, v in obj.items():
        if k == "__info":
            continue
        return v
    return None


def _extract_rpeaks_list(rpeaks: Any) -> List[Dict[str, Any]]:
    if isinstance(rpeaks, list):
        return [x for x in rpeaks if isinstance(x, dict)]
    if isinstance(rpeaks, dict):
        v = _pick_variant_dict(rpeaks, ("merged", "human", "ml"))
        if isinstance(v, list):
            return [x for x in v if isinstance(x, dict)]
    return []


def _extract_rpeaks_list_for_variant(rpeaks: Any, variant: str) -> List[Dict[str, Any]]:
    if isinstance(rpeaks, dict):
        v = rpeaks.get(variant)
        if isinstance(v, list):
            return [x for x in v if isinstance(x, dict)]
    return []


def _extract_mrk_counts_for_variant(h_counts: Any, variant: str) -> Dict[str, int]:
    if not isinstance(h_counts, dict):
        return {}
    v = h_counts.get(variant)
    if not isinstance(v, dict):
        return {}
    out: Dict[str, int] = {}
    for k, val in v.items():
        if k == "__info":
            continue
        if isinstance(val, int):
            out[str(k)] = int(val)
        elif isinstance(val, (float, np.floating)) and float(val).is_integer():
            out[str(k)] = int(val)
    return out


def _extract_mrk_counts(h_counts: Any) -> Dict[str, int]:
    if not isinstance(h_counts, dict):
        return {}
    # Variant dict?
    if any(isinstance(v, dict) for k, v in h_counts.items() if k != "__info"):
        v = _pick_variant_dict(h_counts, ("merged", "human", "ml"))
        if isinstance(v, dict):
            out: Dict[str, int] = {}
            for k, val in v.items():
                if k == "__info":
                    continue
                if isinstance(val, int):
                    out[str(k)] = int(val)
                elif isinstance(val, (float, np.floating)) and float(val).is_integer():
                    out[str(k)] = int(val)
            return out
        return {}
    # Flat dict
    out2: Dict[str, int] = {}
    for k, v in h_counts.items():
        if k == "__info":
            continue
        if isinstance(v, int):
            out2[str(k)] = int(v)
        elif isinstance(v, (float, np.floating)) and float(v).is_integer():
            out2[str(k)] = int(v)
    return out2


def _summarize_rpeaks(rpeaks: Any) -> Tuple[int, Optional[int], Optional[int], Dict[str, int]]:
    lst = _extract_rpeaks_list(rpeaks)
    if not lst:
        return 0, None, None, {}
    idxs: List[int] = []
    ann: Dict[str, int] = {}
    for rp in lst:
        si = rp.get("sampleIndex")
        av = rp.get("annotationValue")
        if isinstance(si, int):
            idxs.append(si)
        elif isinstance(si, (float, np.floating)) and float(si).is_integer():
            idxs.append(int(si))
        if isinstance(av, str):
            ann[av] = ann.get(av, 0) + 1
        elif av is not None:
            s = str(av)
            ann[s] = ann.get(s, 0) + 1
    if not idxs:
        return len(lst), None, None, ann
    idxs.sort()
    return len(lst), idxs[0], idxs[-1], ann


def _sum_noise_samples(noises: Any, fs: Optional[int]) -> Tuple[int, int]:
    # Variant dict?
    if isinstance(noises, dict):
        noises = _pick_variant_dict(noises, ("merged", "human", "ml"))
    if not isinstance(noises, list):
        return 0, 0

    cnt = 0
    total = 0
    for it in noises:
        a = b = None
        if isinstance(it, dict):
            a = it.get("startIndex", it.get("startSample"))
            b = it.get("endIndex", it.get("endSample"))
            if (a is None or b is None) and fs:
                t0 = it.get("startTime")
                t1 = it.get("endTime")
                if (
                    isinstance(t0, (int, float, np.floating))
                    and isinstance(t1, (int, float, np.floating))
                    and t1 > t0
                ):
                    a = int(round(float(t0) * fs))
                    b = int(round(float(t1) * fs))
        elif isinstance(it, (list, tuple)) and len(it) >= 2:
            a, b = it[0], it[1]

        if isinstance(a, (float, np.floating)) and float(a).is_integer():
            a = int(a)
        if isinstance(b, (float, np.floating)) and float(b).is_integer():
            b = int(b)

        if isinstance(a, int) and isinstance(b, int) and b > a:
            cnt += 1
            total += (b - a)
    return cnt, total


def load_json_fs(p: Path) -> Any:
    with open(p, "r", encoding="utf-8") as f:
        return json.load(f)


def file_size_fs(p: Path) -> int:
    return int(p.stat().st_size)


def sample_count_fs(p: Path, ext: str) -> int:
    ext = (ext or "").lower()
    if ext == ".npy":
        arr = np.load(p, allow_pickle=False)
        return int(getattr(arr, "size", 0))
    # Default: assume 4 bytes per sample (int32)
    return int(file_size_fs(p) // 4)


def summarize_record(
    sequenceId: str,
    basename: str,
    json_ref: Path,
    data_ref: Path,
    data_ext: str,
    *,
    load_json_fn: Callable[[Path], Any],
    file_size_fn: Callable[[Path], int],
    sample_count_fn: Callable[[Path, str], int],
    fs: int,
) -> Tuple[RecordSummary, Dict[str, Any]]:
    try:
        meta = load_json_fn(json_ref)
        json_ok = True
    except Exception as exc:
        print(f"WARN: failed to load JSON for {sequenceId}/{basename}: {exc}")
        meta = {}
        json_ok = False

    bsz = int(file_size_fn(data_ref))
    samples = int(sample_count_fn(data_ref, data_ext))

    channel_count = meta.get("channelCount") if isinstance(meta, dict) else None
    user_id = meta.get("userId") if isinstance(meta, dict) else None
    recordingId = meta.get("recordingId") if isinstance(meta, dict) else None
    flags_list = _flatten_flags(meta.get("flags") if isinstance(meta, dict) else None)
    comment = meta.get("comment") if isinstance(meta, dict) else None

    rpeaks_any = meta.get("rpeaks") if isinstance(meta, dict) else None
    rpeaks_count, *_ = _summarize_rpeaks(rpeaks_any)

    mrk_counts_any = meta.get("rpeakAnnotationCounts") if isinstance(meta, dict) else None

    # H counts preference: merged -> human -> derived -> flat
    def _counts_for_h_columns() -> Tuple[Dict[str, int], bool]:
        for var in ("merged", "human"):
            d = _extract_mrk_counts_for_variant(mrk_counts_any, var)
            if d:
                return d, True

        for var in ("merged", "human"):
            rp_list = _extract_rpeaks_list_for_variant(rpeaks_any, var)
            if rp_list:
                tmp: Dict[str, int] = {}
                for rp in rp_list:
                    av = rp.get("annotationValue")
                    if isinstance(av, str):
                        tmp[av] = tmp.get(av, 0) + 1
                    elif av is not None:
                        tmp[str(av)] = tmp.get(str(av), 0) + 1
                if tmp:
                    return tmp, True

        flat = _extract_mrk_counts(mrk_counts_any)
        if flat:
            return flat, True

        return {}, False

    h_counts_primary, has_h_counts = _counts_for_h_columns()
    h_n = int(h_counts_primary.get("N", 0)) if has_h_counts else 0
    h_s = int(h_counts_primary.get("S", 0)) if has_h_counts else 0
    h_v = int(h_counts_primary.get("V", 0)) if has_h_counts else 0
    h_u = int(h_counts_primary.get("U", 0)) if has_h_counts else 0

    # ML counts: explicit "ml" first, else derive from ml rpeaks
    ml_counts = _extract_mrk_counts_for_variant(mrk_counts_any, "ml")
    if not ml_counts:
        ml_rpeaks = _extract_rpeaks_list_for_variant(rpeaks_any, "ml")
        tmp2: Dict[str, int] = {}
        for rp in ml_rpeaks:
            av = rp.get("annotationValue")
            if isinstance(av, str):
                tmp2[av] = tmp2.get(av, 0) + 1
            elif av is not None:
                tmp2[str(av)] = tmp2.get(str(av), 0) + 1
        ml_counts = tmp2

    ml_s = int(ml_counts.get("S", 0))
    ml_v = int(ml_counts.get("V", 0))
    ml_u = int(ml_counts.get("U", 0))

    h_noises = meta.get("noises_annotated") if isinstance(meta, dict) else None
    h_nz_cnt, h_nz_samples = _sum_noise_samples(h_noises, fs)

    ml_noises = meta.get("noises") if isinstance(meta, dict) else None
    ml_nz_cnt, ml_nz_samples = _sum_noise_samples(ml_noises, fs)

    duration_s = (samples / fs) if (fs and fs > 0 and samples > 0) else None

    # IMPORTANT: fractions are ALWAYS percentages (0..100)
    ml_noises_fraction = (ml_nz_samples / samples) * 100.0 if samples > 0 else None
    h_noises_fraction = (h_nz_samples / samples) * 100.0 if samples > 0 else None

    allowed_keys = {
        "channelCount",
        "comment",
        "flags",
        "noises",
        "noises_annotated",
        "recordingId",
        "rpeakAnnotationCounts",
        "rpeaks",
        "userId",
    }
    meta_keys = set(meta.keys()) if isinstance(meta, dict) else set()
    json_keys_correct = meta_keys.issubset(allowed_keys) if meta_keys else False

    rec = RecordSummary(
        sequenceId=str(sequenceId),
        recordingId=str(recordingId) if isinstance(recordingId, str) else None,
        basename=str(basename),
        channel_count=int(channel_count) if isinstance(channel_count, int) else None,
        user_id=str(user_id) if isinstance(user_id, str) else None,
        bin_bytes=bsz,
        samples=samples,
        duration_s=float(duration_s) if duration_s is not None else None,
        cmt=str(comment) if isinstance(comment, str) else None,
        flags_count=len(flags_list),
        flags=flags_list,
        rpeaks_count=int(rpeaks_count),
        h_n_count=h_n,
        h_s_count=h_s,
        h_v_count=h_v,
        h_u_count=h_u,
        ml_s_count=int(ml_s),
        ml_v_count=int(ml_v),
        ml_u_count=int(ml_u),
        json_ok=bool(json_ok),
        ml_noises_count=int(ml_nz_cnt),
        ml_noises_samples=int(ml_nz_samples),
        ml_noises_fraction=float(ml_noises_fraction) if ml_noises_fraction is not None else None,
        h_noises_count=int(h_nz_cnt),
        h_noises_samples=int(h_nz_samples),
        h_noises_fraction=float(h_noises_fraction) if h_noises_fraction is not None else None,
        json_keys_correct=bool(json_keys_correct),
    )

    rec_dict = asdict(rec)
    rec_dict["_ml_noises_samples"] = int(ml_nz_samples)
    rec_dict["_h_noises_samples"] = int(h_nz_samples)
    return rec, rec_dict


def infer_sequence_id_fs(json_path: Path, fallback: str) -> str:
    parts = list(json_path.parts)
    if "recordings" in parts:
        idx = parts.index("recordings")
        if idx >= 1:
            return parts[idx - 1]
    if len(parts) >= 2:
        return parts[-2]
    return fallback


def find_data_file_fs(json_path: Path) -> Optional[Tuple[Path, str]]:
    if json_path.suffix.lower() != ".json":
        return None
    rec_dir = json_path.parent
    stem = json_path.stem

    exact = rec_dir / stem
    if exact.exists():
        suf = exact.suffix
        if suf and suf[1:].isdigit() and len(suf) == 4:
            return exact, suf

    candidates = [
        rec_dir / f"{stem}.npy",
        rec_dir / f"{stem}.bin",
        rec_dir / f"{stem}",
    ]
    for p in sorted(rec_dir.glob(f"{stem}.*")):
        suf = p.suffix
        if suf and suf[1:].isdigit() and len(suf) == 4:
            candidates.append(p)

    for c in candidates:
        if c.exists():
            return c, c.suffix
    return None


# ---------------------------------------------------------------------
# NEW: ECG loader for denoising
# ---------------------------------------------------------------------

READ_SCALE = 1.0  # <-- if your binary int32 needs scaling to mV, change this (e.g., 1/1000.0)

# def load_ecg_1d_for_denoising(data_path: Path, data_ext: str) -> np.ndarray:
#     ext = (data_ext or data_path.suffix or "").lower()
#     if ext == ".npy":
#         x = np.load(data_path, allow_pickle=False)
#         x = np.asarray(x).squeeze()
#         if x.ndim != 1:
#             raise RuntimeError(f"{data_path} is not 1D after squeeze(), got shape={x.shape}")
#         return x.astype(np.float32, copy=False)

#     # Fallback: read int32 big-endian, cast to float
#     raw = np.fromfile(data_path, dtype=">i4")  # big-endian int32
#     x = raw.astype(np.float32) * float(READ_SCALE)
#     return x


def load_ecg_1d_for_denoising(data_path: Path, data_ext: str) -> np.ndarray:

    x = load_ecg_npy(data_path)

    return x
# ---------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------

LIGHT_RED_FILL = PatternFill(patternType="solid", fgColor="FFFFC7CE")  # new rows
MEDIUM_BLUE_FILL = PatternFill(patternType="solid", fgColor="FF1E90FF")  # changed existing rows


def norm_basename(x: Any) -> Optional[str]:
    """Normalize basenames for matching: ',' -> '.', strip, and drop trailing zeros in decimals."""
    if x is None:
        return None
    s = str(x).strip()
    if not s:
        return None
    s = s.replace(",", ".")
    m = re.fullmatch(r"(\d+)\.(\d+)", s)
    if m:
        a, b = m.group(1), m.group(2).rstrip("0")
        return a if b == "" else f"{a}.{b}"
    return s


def build_header_map(ws) -> Dict[str, int]:
    hdr: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        key = str(v).strip().lower()
        if key:
            hdr[key] = col
    return hdr


def pick_col(hdr: Dict[str, int], *names: str) -> Optional[int]:
    for n in names:
        k = n.strip().lower()
        if k in hdr:
            return hdr[k]
    return None


def cell_changed(old: Any, new: Any) -> bool:
    """Loose compare to avoid false mismatches (e.g. '12,3' vs 12.3)."""
    def norm(v: Any) -> Any:
        if v is None:
            return None
        if isinstance(v, bool):
            return bool(v)
        if isinstance(v, int):
            return int(v)
        if isinstance(v, float):
            if math.isnan(v):
                return None
            return float(v)
        s = str(v).strip()
        try:
            return float(s.replace(",", "."))
        except Exception:
            return s
    return norm(old) != norm(new)


def _flags_to_cell_value(flags_list: List[str]) -> str:
    return "; ".join([f for f in flags_list if f])

def prepare_denoising_pipeline(
    config_path: Path,
    model_dir: Path,
) -> DenoisingPipelineConfig:
    """Pipeline-aware runner that wires configuration, data loading and execution."""

    if not config_path.exists():
        raise FileNotFoundError(f"Config file not found: {config_path}")
    if not model_dir.exists():
        raise FileNotFoundError(f"Model directory not found: {model_dir}")

    # print_heading("Project paths")
    # print("DATA_DIR:", data_dir)
    # print("CONFIG  :", config_path)
    # print("MODEL_DIR:", model_dir)

    cfg = load_denoising_config_yaml(str(config_path))
    fs = float(cfg.fs)

    # print_heading("Denoising pipeline config")
    # friendly_print_denoising_cfg(cfg)
    check_denoising_config(cfg)
    
    # print_heading("Input data")
    # print(f"File: {file_name}")
    # len_secs = math.ceil(len(x) / fs)
    # h, m, s = convert_seconds_to_hms(len_secs)
    # print(f"len(ecg): {len(x)} samples (~{len_secs:.1f} s) | {h:02d}:{m:02d}:{s:02d}")

    # print("path:", path)
    # print(f"Loaded {len(gaps_indices)} gap intervals.")

    cfg.motions.model_name = resolve_model_path(model_dir, cfg.motions.model_name)
    cfg.motions.enabled = True

    # print_heading("UNet model")
    # print("UNet model path:", cfg.motions.model_name)
    # print("Motions enabled:", bool(getattr(cfg.motions, "enabled", True)))

    return cfg


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", type=Path, required=True, help="Input Excel (.xlsx)")
    ap.add_argument("--dir", type=Path, required=True, help="Directory containing JSON metadata and data files")
    ap.add_argument("--fs", type=int, default=200, help="Sampling frequency (Hz). Default: 200")
    ap.add_argument("--sheet", type=str, default=None, help="Sheet name (default: active sheet)")
    ap.add_argument("--out", type=Path, default=None, help="Output Excel path (default: <excel>_updated.xlsx)")

    # NEW: denoising args (prepared once)
    ap.add_argument("--cfg-denoising", type=Path, required=True, help="Denoising config path")
    ap.add_argument("--model-dir", type=Path, required=True, help="Model directory for denoising/motions")
    ap.add_argument("--disable-motions", action="store_true", help="Disable motions stage in denoising pipeline")
    ap.add_argument("--quiet", action="store_true", help="Silence stdout during denoising/stats")

    args = ap.parse_args()
    
    def _fmt(v):
        return str(v) if isinstance(v, Path) else v

    d = vars(args)
    width = max(len(k) for k in d)
    print("Arguments:")
    for k in sorted(d):
        print(f"  {k:<{width}} : {_fmt(d[k])}")
    
    src = args.dir
    if not src.exists() or not src.is_dir():
        raise FileNotFoundError(f"--dir must be an existing directory. Got: {src}")

    # -------------------------------------------------------------
    # NEW: Prepare denoising ONCE (before main cycle)
    # -------------------------------------------------------------
    cfg_denoising_path = args.cfg_denoising
    model_dir = args.model_dir
    
    #  Preparing the Denoising pipeline with configuration and model paths
    print("\n*****Denoising pipeline config:")
    cfg_denoising = prepare_denoising_pipeline(cfg_denoising_path, model_dir)
    pipe = ECGDenoisingPipeline(cfg_denoising)

    wb = openpyxl.load_workbook(args.excel)
    ws = wb[args.sheet] if args.sheet else wb.active

    hdr = build_header_map(ws)

    cols: Dict[str, Optional[int]] = {}
    cols["basename"] = pick_col(hdr, "basename")
    cols["rec_id"] = pick_col(hdr, "rec_id", "recordingid", "recording_id", "recording id")
    cols["uid"] = pick_col(hdr, "uid", "user_id", "userid", "user id")
    cols["samples"] = pick_col(hdr, "samples", "sample_cnt", "sample_count")
    cols["cmt"] = pick_col(hdr, "cmt", "comment")
    cols["dur_s"] = pick_col(hdr, "dur_s", "duration_s", "duration", "dur")
    cols["rpk_cnt"] = pick_col(hdr, "rpk_cnt", "rpeaks_cnt", "rpeaks_count", "rpeaks")
    cols["hN"] = pick_col(hdr, "hn", "hN")
    cols["hS"] = pick_col(hdr, "hs", "hS")
    cols["hV"] = pick_col(hdr, "hv", "hV")
    cols["hU"] = pick_col(hdr, "hu", "hU")
    cols["mlS"] = pick_col(hdr, "mls", "mlS")
    cols["mlV"] = pick_col(hdr, "mlv", "mlV")
    cols["mlU"] = pick_col(hdr, "mlu", "mlU")
    cols["h_nz_cnt"] = pick_col(hdr, "h_nz_cnt", "h_nz_count")

    cols["h_nz_len"] = pick_col(hdr, "h_nz_len", "h_nz_length")
    cols["h_nz_frac"] = pick_col(hdr, "h_nz_frac", "h_nz_frac%", "h_nz_frac %")
    cols["noni"] = pick_col(hdr, "noni")

    # KEEP ml_nz_frac column (we will set it = tp%)
    cols["ml_nz_frac"] = pick_col(hdr, "ml_nz_frac", "ml_nz_frac%", "ml_nz_frac %")

    cols["flags"] = pick_col(hdr, "flags")

    # NEW: columns to write instead of ml_nz_cnt/ml_nz_len
    cols["out"] = pick_col(hdr, "out", "out_cnt", "outliers")
    cols["rdr"] = pick_col(hdr, "rdr", "rdr_cnt", "rdropouts")
    cols["noi"] = pick_col(hdr, "noi", "noi_cnt", "motions")
    cols["tp"] = pick_col(hdr, "tp", "tp%", "tp_pct", "tp %")

    if cols["basename"] is None:
        raise RuntimeError("Missing required column in Excel header row: ['basename']")

    c_basename = int(cols["basename"])

    # Map basename -> row index
    basename_to_row: Dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        bn = norm_basename(ws.cell(row=r, column=c_basename).value)
        if bn:
            basename_to_row[bn] = r

    # Copy formatting template from row 2
    style_row_idx = 2 if ws.max_row >= 2 else None

    def copy_row_style(dst_row: int) -> None:
        if style_row_idx is None:
            return
        for c in range(1, ws.max_column + 1):
            src_cell = ws.cell(row=style_row_idx, column=c)
            dst_cell = ws.cell(row=dst_row, column=c)
            dst_cell._style = _copy_style(src_cell._style)
            dst_cell.font = _copy_style(src_cell.font)
            dst_cell.border = _copy_style(src_cell.border)
            dst_cell.fill = _copy_style(src_cell.fill)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = _copy_style(src_cell.protection)
            dst_cell.alignment = _copy_style(src_cell.alignment)

    new_rows: Set[int] = set()
    changed_rows: Set[int] = set()

    def get_or_create_row(bn: str) -> Tuple[int, bool]:
        row = basename_to_row.get(bn)
        if row is not None:
            return row, False

        row = ws.max_row + 1
        copy_row_style(row)
        ws.cell(row=row, column=c_basename).value = bn
        basename_to_row[bn] = row
        new_rows.add(row)
        ws.cell(row=row, column=c_basename).fill = LIGHT_RED_FILL
        return row, True

    def set_cell(row: int, col_key: str, new_val: Any) -> bool:
        col = cols.get(col_key)
        if col is None:
            return False
        col_i = int(col)
        cell = ws.cell(row=row, column=col_i)

        # percent columns always mean 0..100
        if col_key in ("h_nz_frac", "ml_nz_frac", "tp") and isinstance(new_val, (int, float, np.floating)):
            new_val = round(float(new_val), 1)

        if cell_changed(cell.value, new_val):
            cell.value = new_val
            return True
        return False

    json_paths = [p for p in src.rglob("*.json") if "__MACOSX" not in p.parts and p.is_file()]
    print(f"Source: {src}")
    print(f"Found JSON files: {len(json_paths)}")

    processed = 0
    for jp in sorted(json_paths):
        bn = norm_basename(jp.stem)
        if not bn:
            continue
        if bn.lower() == "manifest":
            continue

        row, _is_new = get_or_create_row(bn)
        print(f"Processing: {jp} {bn} -> row {row} (new: {_is_new})")

        data_info = find_data_file_fs(jp)
        if data_info is None:
            # JSON exists but data missing -> fill what we can from JSON only (no denoising)
            meta = load_json_fs(jp)
            if not isinstance(meta, dict):
                meta = {}
            extracted = {
                "rec_id": meta.get("recordingId") if isinstance(meta.get("recordingId"), str) else None,
                "uid": meta.get("userId") if isinstance(meta.get("userId"), str) else None,
                "flags": _flags_to_cell_value(_flatten_flags(meta.get("flags"))),
            }
        else:
            data_path, data_ext = data_info
            seq = infer_sequence_id_fs(jp, fallback=src.name or "dir")
            rec, _rec_dict = summarize_record(
                sequenceId=seq,
                basename=bn,
                json_ref=jp,
                data_ref=data_path,
                data_ext=data_ext,
                load_json_fn=load_json_fs,
                file_size_fn=file_size_fs,
                sample_count_fn=sample_count_fs,
                fs=args.fs,
            )

            # -------------------------------------------------------------
            # NEW: denoise + stats per record
            # -------------------------------------------------------------
            try:
                x = load_ecg_1d_for_denoising(data_path, data_ext)

                # Execute the pipeline in the notebook with explicit arguments
                print("\nRunning Denoising pipeline...")
                if args.quiet:
                    with contextlib.redirect_stdout(io.StringIO()):
                        res_denoising = pipe.run(x, gaps_indices=[])
                        stats = calc_noise_stats_from_result(res_denoising)
                else:
                    res_denoising = pipe.run(x, gaps_indices=[])
                    stats = calc_noise_stats_from_result(res_denoising)

                out_cnt = int(stats.get("out", 0))
                rdr_cnt = int(stats.get("rdr", 0))
                noi_cnt = int(stats.get("noi", 0))
                tp_pct = float(stats.get("tp_pct", 0.0))

            except Exception as exc:
                print(f"WARN: denoising/stats failed for {bn}: {exc}")
                out_cnt = rdr_cnt = noi_cnt = 0
                tp_pct = 0.0

            # -------------------------------------------------------------
            # Extracted fields written to Excel
            # -------------------------------------------------------------
            extracted = {
                "rec_id": rec.recordingId,
                "uid": rec.user_id,
                "samples": int(rec.samples),
                "cmt": rec.cmt,
                "dur_s": float(rec.duration_s) if rec.duration_s is not None else None,
                "rpk_cnt": int(rec.rpeaks_count),
                "hN": int(rec.h_n_count),
                "hS": int(rec.h_s_count),
                "hV": int(rec.h_v_count),
                "hU": int(rec.h_u_count),
                "mlS": int(rec.ml_s_count),
                "mlV": int(rec.ml_v_count),
                "mlU": int(rec.ml_u_count),
                "h_nz_cnt": int(rec.h_noises_count),
                "h_nz_len": int(_rec_dict.get("_h_noises_samples", 0)),
                "h_nz_frac": float(rec.h_noises_fraction) if rec.h_noises_fraction is not None else None,  # percent

                # NEW: instead of ml_nz_cnt/ml_nz_len
                "out": out_cnt,
                "rdr": rdr_cnt,
                "noi": noi_cnt,
                "tp": tp_pct,

                # NEW: ml_nz_frac% = tp%
                "ml_nz_frac": tp_pct,

                "flags": _flags_to_cell_value(rec.flags or []),
                "noni": int(rec.h_noises_count),
            }

        row_changed = False
        for k, v in extracted.items():
            if v is None:
                continue
            row_changed |= set_cell(row, k, v)
            if k == "h_nz_cnt":
                row_changed |= set_cell(row, "noni", v)

        if row_changed:
            changed_rows.add(row)

        processed += 1

    for r in changed_rows:
        if r in new_rows:
            continue
        ws.cell(row=r, column=c_basename).fill = MEDIUM_BLUE_FILL

    out_path = args.out if args.out else args.excel.with_name(args.excel.stem + "_updated_3.xlsx")
    wb.save(out_path)

    print(f"Processed JSON files: {processed}")
    print(f"Changed rows: {len(changed_rows)}")
    print(f"Added new rows: {len(new_rows)}")
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()