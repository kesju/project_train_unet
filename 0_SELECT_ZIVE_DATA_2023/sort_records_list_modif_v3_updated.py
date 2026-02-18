import os, re
from collections import OrderedDict
from copy import copy

import openpyxl
from openpyxl.styles import PatternFill


IN_PATH = "visi_zive_irasai_atrankai._modif_v1 - Darb_updated.xlsx"


def norm_header(x):
    if x is None:
        return ""
    s = str(x).strip().lower()
    s = re.sub(r"\s+", "", s)
    return s


def to_float(x, default=0.0):
    """Accepts numbers or strings like '12,5%' / '12.5%' / '12,5'."""
    if x is None:
        return default
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if not s:
        return default
    s = s.replace("%", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return default


def copy_cell_style(src_cell, dst_cell):
    """Copy everything EXCEPT value."""
    if src_cell.has_style:
        dst_cell._style = copy(src_cell._style)
    dst_cell.number_format = src_cell.number_format
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.protection = copy(src_cell.protection)
    dst_cell.comment = src_cell.comment


def copy_cell(src_cell, dst_cell):
    """Copy value + style."""
    dst_cell.value = src_cell.value
    copy_cell_style(src_cell, dst_cell)


def main(in_path: str):
    if not os.path.exists(in_path):
        raise FileNotFoundError(f"Input not found: {in_path}")

    wb = openpyxl.load_workbook(in_path)
    ws = wb.active

    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    hmap = {norm_header(v): i + 1 for i, v in enumerate(header)}

    def col_idx(*cands):
        for c in cands:
            k = norm_header(c)
            if k in hmap:
                return hmap[k]
        return None

    c_user  = col_idx("userId", "userid", "user_id")
    c_hS    = col_idx("hS", "hs")
    c_hV    = col_idx("hV", "hv")
    c_mlS   = col_idx("mlS", "mls")
    c_mlV   = col_idx("mlV", "mlv")
    c_hnzf  = col_idx("h_nz_frac%", "h_nz_frac", "hnzfrac%", "hnzfrac")
    c_mlnzf = col_idx("ml_nz_frac%", "ml_nz_frac", "mlnzfrac%", "mlnzfrac")

    missing = [
        name for name, c in [
            ("userId", c_user),
            ("hS", c_hS),
            ("hV", c_hV),
            ("mlS", c_mlS),
            ("mlV", c_mlV),
            ("h_nz_frac%", c_hnzf),
            ("ml_nz_frac%", c_mlnzf),
        ] if c is None
    ]
    if missing:
        raise RuntimeError(
            f"Missing required columns in header row: {missing}\n"
            f"Available headers (normalized): {sorted(hmap.keys())}"
        )

    # Read rows
    rows = []
    for r in range(2, ws.max_row + 1):
        uid = ws.cell(row=r, column=c_user).value
        if uid is None and all(
            ws.cell(row=r, column=c).value in (None, "")
            for c in range(1, ws.max_column + 1)
        ):
            continue

        row_cells = [ws.cell(row=r, column=c) for c in range(1, ws.max_column + 1)]
        hs_sum = to_float(ws.cell(row=r, column=c_hS).value) + to_float(ws.cell(row=r, column=c_hV).value)
        ml_sum = to_float(ws.cell(row=r, column=c_mlS).value) + to_float(ws.cell(row=r, column=c_mlV).value)
        hnzf = to_float(ws.cell(row=r, column=c_hnzf).value, default=0.0)
        mlnzf = to_float(ws.cell(row=r, column=c_mlnzf).value, default=0.0)

        rows.append({
            "orig_row": r,
            "userId": "" if uid is None else str(uid).strip(),
            "hs_sum": hs_sum,
            "ml_sum": ml_sum,
            "hnzf": hnzf,
            "mlnzf": mlnzf,
            "cells": row_cells,
        })

    # Group by userId in first-appearance order
    groups = OrderedDict()
    for item in rows:
        groups.setdefault(item["userId"], []).append(item)

    def sort_group(items):
        all_hs_zero = all(abs(x["hs_sum"]) < 1e-12 for x in items)
        if all_hs_zero:
            return sorted(items, key=lambda x: (-x["ml_sum"], x["mlnzf"], x["orig_row"]))
        return sorted(items, key=lambda x: (-x["hs_sum"], x["hnzf"], x["orig_row"]))

    sorted_groups = [(uid, sort_group(items)) for uid, items in groups.items()]

    base, ext = os.path.splitext(in_path)
    out_path = base + "_sorted" + ext

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = ws.title
    out_ws.sheet_view.freeze_panes = ws.freeze_panes

    # Column widths
    out_ws.column_dimensions["A"].width = 8
    for col_letter, dim in ws.column_dimensions.items():
        if len(col_letter) == 1 and "A" <= col_letter <= "Y":
            shifted = chr(ord(col_letter) + 1)
            out_ws.column_dimensions[shifted].width = dim.width

    # Header row: NEW A1 value, style copied without overwriting value
    h1 = out_ws.cell(row=1, column=1, value="grp_no")
    copy_cell_style(ws.cell(row=1, column=1), h1)

    # Copy original headers shifted by +1
    for c in range(1, ws.max_column + 1):
        copy_cell(ws.cell(row=1, column=c), out_ws.cell(row=1, column=c + 1))

    # Highlight for first row of each group
    # group_fill = PatternFill("solid", fgColor="E8F4FF")
    # group_fill = PatternFill("solid", fgColor="FFF2CC")  # light yellow
    # group_fill = PatternFill("solid", fgColor="FFFFCC")  # very pale yellow
    # group_fill = PatternFill("solid", fgColor="FFF9C4")  # soft light yellow
    group_fill = PatternFill("solid", fgColor="FFFF00")  # pure yellow


    out_r = 2
    group_no = 1
    for uid, items in sorted_groups:
        first = True
        for item in items:
            # grp_no value must not be overwritten by style copy
            gcell = out_ws.cell(row=out_r, column=1, value=group_no)
            copy_cell_style(item["cells"][0], gcell)

            # Copy row cells shifted by +1
            for j, src_cell in enumerate(item["cells"], start=2):
                copy_cell(src_cell, out_ws.cell(row=out_r, column=j))

            if first:
                for j in range(1, ws.max_column + 2):
                    out_ws.cell(row=out_r, column=j).fill = copy(group_fill)
                first = False

            out_r += 1
        group_no += 1

    if ws.row_dimensions[1].height is not None:
        out_ws.row_dimensions[1].height = ws.row_dimensions[1].height

    out_wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main(IN_PATH)
