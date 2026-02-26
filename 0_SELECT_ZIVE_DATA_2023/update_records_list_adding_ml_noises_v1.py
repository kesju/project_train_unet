#!/usr/bin/env python3
"""
Update ECG record parameters in an Excel file from per-record JSON metadata stored in a ZIP
(or an unzipped folder containing the same structure).

Required behavior (added/changed in this version):
- If a JSON record basename is NOT present in the input Excel, append a NEW row.
- For newly appended rows, fill all columns that can be derived from JSON (+ data file if present):
  basename, rec_id, uid, samples, dur_s, rpk_cnt, hN, hS, hV, hU, mlS, mlV, mlU,
  h_nz_cnt, h_nz_frac, ml_nz_cnt, ml_nz_frac, flags
- Mark the NEW row's "basename" cell with LIGHT RED fill.

Other:
- Matching basenames is normalized: ',' -> '.', trailing zeros after decimal are removed.
- Fractions (h_nz_frac / ml_nz_frac) are stored as percent values (0..100), rounded to 1 decimal.
"""

from __future__ import annotations

import argparse
import io
import json
import math
import re
import zipfile
from copy import copy as _copy_style
from dataclasses import asdict, dataclass
from pathlib import Path, PurePosixPath
from typing import Any, Callable, Dict, List, Optional, Set, Tuple, Union

import numpy as np
import openpyxl
from openpyxl.styles import PatternFill

JsonLikeRef = Union[str, Path]


# ---------------------------------------------------------------------
# Embedded analyzer helpers (from analyze_records_2026_v4.py)
# ---------------------------------------------------------------------

@dataclass
class RecordSummary:
    sequenceId: str
    recordingId: Optional[str]
    basename: str
    channel_count: Optional[int]
    user_id: Optional[str]

    bin_bytes: int
    samples: int
    duration_s: Optional[float]

    flags_count: int
    flags: List[str]

    rpeaks_count: int
    h_n_count: int
    h_s_count: int
    h_v_count: int
    h_u_count: int

    ml_s_count: int
    ml_v_count: int
    ml_u_count: int

    json_ok: bool

    ml_noises_count: int
    ml_noises_samples: int
    ml_noises_fraction: Optional[float]

    h_noises_count: int
    h_noises_samples: int
    h_noises_fraction: Optional[float]

    cmt: Optional[str]
    json_keys_correct: bool
    notes: Optional[str] = None


def _flatten_flags(flags: Any) -> List[str]:
    if flags is None:
        return []
    if isinstance(flags, list):
        return [str(x) for x in flags]
    return [str(flags)]


def _pick_variant_dict(obj: Any, keys_preference: Tuple[str, ...] = ("merged", "human", "ml")) -> Any:
    if not isinstance(obj, dict):
        return None
    for k in keys_preference:
        if k in obj:
            return obj.get(k)
    for k, v in obj.items():
        if k == "__info":
            continue
        return v
    return None


def _extract_rpeaks_list(rpeaks: Any) -> List[Dict[str, Any]]:
    if isinstance(rpeaks, list):
        return [x for x in rpeaks if isinstance(x, dict)]
    if isinstance(rpeaks, dict):
        v = _pick_variant_dict(rpeaks, ("merged", "human", "ml"))
        if isinstance(v, list):
            return [x for x in v if isinstance(x, dict)]
    return []


def _extract_mrk_counts(h_counts: Any) -> Dict[str, int]:
    if not isinstance(h_counts, dict):
        return {}
    # Variant dict?
    if any(isinstance(v, dict) for k, v in h_counts.items() if k != "__info"):
        v = _pick_variant_dict(h_counts, ("merged", "human", "ml"))
        if isinstance(v, dict):
            out: Dict[str, int] = {}
            for k, val in v.items():
                if k == "__info":
                    continue
                if isinstance(val, int):
                    out[str(k)] = int(val)
                elif isinstance(val, (float, np.floating)) and float(val).is_integer():
                    out[str(k)] = int(val)
            return out
        return {}
    # Flat dict
    out2: Dict[str, int] = {}
    for k, v in h_counts.items():
        if k == "__info":
            continue
        if isinstance(v, int):
            out2[str(k)] = int(v)
        elif isinstance(v, (float, np.floating)) and float(v).is_integer():
            out2[str(k)] = int(v)
    return out2


def _extract_rpeaks_list_for_variant(rpeaks: Any, variant: str) -> List[Dict[str, Any]]:
    if isinstance(rpeaks, dict):
        v = rpeaks.get(variant)
        if isinstance(v, list):
            return [x for x in v if isinstance(x, dict)]
    return []


def _extract_mrk_counts_for_variant(h_counts: Any, variant: str) -> Dict[str, int]:
    if not isinstance(h_counts, dict):
        return {}
    v = h_counts.get(variant)
    if not isinstance(v, dict):
        return {}
    out: Dict[str, int] = {}
    for k, val in v.items():
        if k == "__info":
            continue
        if isinstance(val, int):
            out[str(k)] = int(val)
        elif isinstance(val, (float, np.floating)) and float(val).is_integer():
            out[str(k)] = int(val)
    return out


def _summarize_rpeaks(rpeaks: Any) -> Tuple[int, Optional[int], Optional[int], Dict[str, int]]:
    lst = _extract_rpeaks_list(rpeaks)
    if not lst:
        return 0, None, None, {}
    idxs: List[int] = []
    ann: Dict[str, int] = {}
    for rp in lst:
        si = rp.get("sampleIndex")
        av = rp.get("annotationValue")
        if isinstance(si, int):
            idxs.append(si)
        elif isinstance(si, (float, np.floating)) and float(si).is_integer():
            idxs.append(int(si))
        if isinstance(av, str):
            ann[av] = ann.get(av, 0) + 1
        elif av is not None:
            s = str(av)
            ann[s] = ann.get(s, 0) + 1
    if not idxs:
        return len(lst), None, None, ann
    idxs.sort()
    return len(lst), idxs[0], idxs[-1], ann


def _sum_noise_samples(noises: Any, fs: Optional[int]) -> Tuple[int, int]:
    # Variant dict?
    if isinstance(noises, dict):
        v = _pick_variant_dict(noises, ("merged", "human", "ml"))
        noises = v
    if not isinstance(noises, list):
        return 0, 0
    cnt = 0
    total = 0
    for it in noises:
        a = b = None
        if isinstance(it, dict):
            a = it.get("startIndex", it.get("startSample"))
            b = it.get("endIndex", it.get("endSample"))
            if (a is None or b is None) and fs:
                t0 = it.get("startTime")
                t1 = it.get("endTime")
                if isinstance(t0, (int, float, np.floating)) and isinstance(t1, (int, float, np.floating)) and t1 > t0:
                    a = int(round(float(t0) * fs))
                    b = int(round(float(t1) * fs))
        elif isinstance(it, (list, tuple)) and len(it) >= 2:
            a, b = it[0], it[1]
        if isinstance(a, (float, np.floating)) and float(a).is_integer():
            a = int(a)
        if isinstance(b, (float, np.floating)) and float(b).is_integer():
            b = int(b)
        if isinstance(a, int) and isinstance(b, int) and b > a:
            cnt += 1
            total += (b - a)
    return cnt, total


def load_json_zip(zf: zipfile.ZipFile, member: str) -> Any:
    with zf.open(member) as f:
        return json.load(f)


def file_size_zip(zf: zipfile.ZipFile, member: str) -> int:
    return int(zf.getinfo(member).file_size)


def sample_count_zip(zf: zipfile.ZipFile, member: str, ext: str) -> int:
    ext = (ext or "").lower()
    if ext == ".npy":
        raw = zf.read(member)
        arr = np.load(io.BytesIO(raw), allow_pickle=False)
        return int(getattr(arr, "size", 0))
    # Default: assume 4 bytes per sample (int32)
    return int(file_size_zip(zf, member) // 4)


def summarize_record(
    sequenceId: str,
    basename: str,
    json_ref: JsonLikeRef,
    data_ref: JsonLikeRef,
    data_ext: str,
    *,
    load_json_fn: Callable[[JsonLikeRef], Any],
    file_size_fn: Callable[[JsonLikeRef], int],
    sample_count_fn: Callable[[JsonLikeRef, str], int],
    fs: int,
) -> Tuple[RecordSummary, Dict[str, Any]]:
    try:
        meta = load_json_fn(json_ref)
        json_ok = True
    except Exception as exc:
        print(f"WARN: failed to load JSON for {sequenceId}/{basename}: {exc}")
        meta = {}
        json_ok = False

    bsz = int(file_size_fn(data_ref))
    samples = int(sample_count_fn(data_ref, data_ext))

    channel_count = meta.get("channelCount") if isinstance(meta, dict) else None
    user_id = meta.get("userId") if isinstance(meta, dict) else None
    recordingId = meta.get("recordingId") if isinstance(meta, dict) else None
    flags_list = _flatten_flags(meta.get("flags") if isinstance(meta, dict) else None)
    comment = meta.get("comment") if isinstance(meta, dict) else None
    rpeaks_any = meta.get("rpeaks") if isinstance(meta, dict) else None
    rpeaks_count, _rpk_first, _rpk_last, _rpk_ann = _summarize_rpeaks(rpeaks_any)

    mrk_counts_any = meta.get("rpeakAnnotationCounts") if isinstance(meta, dict) else None
    # print(f"DEBUG: {sequenceId}/{basename} - raw mrk_counts_any: {mrk_counts_any}")

    # Pick counts for the "H" columns (hN/hS/hV/hU).
    # IMPORTANT: Prefer explicit "merged" counts if present, then "human".
    # Only if neither exists do we fall back to a flat dict / derived counts.
    def _counts_for_h_columns() -> Tuple[Dict[str, int], str]:
        # 1) Explicit counts by variant
        for var in ("merged", "human"):
            d = _extract_mrk_counts_for_variant(mrk_counts_any, var)
            if d:
                return d, var

        # 2) Derive from rpeaks by variant (merged first, then human)
        for var in ("merged", "human"):
            rp_list = _extract_rpeaks_list_for_variant(rpeaks_any, var)
            if rp_list:
                tmp: Dict[str, int] = {}
                for rp in rp_list:
                    av = rp.get("annotationValue")
                    if isinstance(av, str):
                        tmp[av] = tmp.get(av, 0) + 1
                    elif av is not None:
                        s = str(av)
                        tmp[s] = tmp.get(s, 0) + 1
                if tmp:
                    return tmp, f"derived:{var}"

        # 3) Flat dict (no variants) -> treat as primary counts
        flat = _extract_mrk_counts(mrk_counts_any)
        if flat:
            return flat, "flat"

        return {}, "none"

    h_counts_primary, h_counts_src = _counts_for_h_columns()
    # print(f"DEBUG: {sequenceId}/{basename} - h_counts_primary source={h_counts_src}: {h_counts_primary}")

    has_h_counts = bool(h_counts_primary)

    h_n = int(h_counts_primary.get("N", 0)) if has_h_counts else 0
    h_s = int(h_counts_primary.get("S", 0)) if has_h_counts else 0
    h_v = int(h_counts_primary.get("V", 0)) if has_h_counts else 0
    h_u = int(h_counts_primary.get("U", 0)) if has_h_counts else 0

    ml_counts = _extract_mrk_counts_for_variant(mrk_counts_any, "ml")
    if not ml_counts:
        ml_rpeaks = _extract_rpeaks_list_for_variant(rpeaks_any, "ml")
        tmp2: Dict[str, int] = {}
        for rp in ml_rpeaks:
            av = rp.get("annotationValue")
            if isinstance(av, str):
                tmp2[av] = tmp2.get(av, 0) + 1
            elif av is not None:
                s = str(av)
                tmp2[s] = tmp2.get(s, 0) + 1
        ml_counts = tmp2

    ml_s = int(ml_counts.get("S", 0))
    ml_v = int(ml_counts.get("V", 0))
    ml_u = int(ml_counts.get("U", 0))

    h_noises = meta.get("noises_annotated") if isinstance(meta, dict) else None
    h_nz_cnt, h_nz_samples = _sum_noise_samples(h_noises, fs)
    ml_noises = meta.get("noises") if isinstance(meta, dict) else None
    ml_nz_cnt, ml_nz_samples = _sum_noise_samples(ml_noises, fs)

    duration_s = (samples / fs) if (fs and fs > 0 and samples > 0) else None
    ml_noises_fraction = (ml_nz_samples / samples) * 100.0 if samples > 0 else None
    h_noises_fraction = (h_nz_samples / samples) * 100.0 if samples > 0 else None

    allowed_keys = {
        "channelCount",
        "comment",
        "flags",
        "noises",
        "noises_annotated",
        "recordingId",
        "rpeakAnnotationCounts",
        "rpeaks",
        "userId",
    }
    meta_keys = set(meta.keys()) if isinstance(meta, dict) else set()
    json_keys_correct = meta_keys.issubset(allowed_keys) if meta_keys else False

    rec = RecordSummary(
        sequenceId=str(sequenceId),
        recordingId=str(recordingId) if isinstance(recordingId, str) else None,
        basename=str(basename),
        channel_count=int(channel_count) if isinstance(channel_count, int) else None,
        user_id=str(user_id) if isinstance(user_id, str) else None,
        bin_bytes=bsz,
        samples=samples,
        duration_s=float(duration_s) if duration_s is not None else None,
        cmt=str(comment) if isinstance(comment, str) else None,
        flags_count=len(flags_list),
        flags=flags_list,
        rpeaks_count=int(rpeaks_count),
        h_n_count=h_n,
        h_s_count=h_s,
        h_v_count=h_v,
        h_u_count=h_u,
        ml_s_count=int(ml_s),
        ml_v_count=int(ml_v),
        ml_u_count=int(ml_u),
        json_ok=bool(json_ok),
        ml_noises_count=int(ml_nz_cnt),
        ml_noises_samples=int(ml_nz_samples),
        ml_noises_fraction=float(ml_noises_fraction) if ml_noises_fraction is not None else None,
        h_noises_count=int(h_nz_cnt),
        h_noises_samples=int(h_nz_samples),
        h_noises_fraction=float(h_noises_fraction) if h_noises_fraction is not None else None,
        json_keys_correct=bool(json_keys_correct),
    )

    rec_dict = asdict(rec)
    rec_dict["__has_human_counts"] = bool(has_h_counts)  # kept key name for backward compatibility
    rec_dict["_ml_noises_samples"] = int(ml_nz_samples)
    rec_dict["_h_noises_samples"] = int(h_nz_samples)
    return rec, rec_dict


def find_data_file_zip(json_member: str, name_set: Set[str]) -> Optional[Tuple[str, str]]:
    pp = PurePosixPath(json_member)
    if pp.suffix.lower() != ".json":
        return None
    rec_dir = pp.parent
    stem = pp.stem

    exact = str(rec_dir / stem)
    ex_suffix = PurePosixPath(exact).suffix
    if exact in name_set and ex_suffix[1:].isdigit() and len(ex_suffix) == 4:
        return exact, ex_suffix

    candidates = [
        str(rec_dir / f"{stem}.npy"),
        str(rec_dir / f"{stem}.bin"),
        str(rec_dir / f"{stem}"),
    ]
    prefix = str(rec_dir / stem) + "."
    for n in sorted(name_set):
        if n.startswith(prefix):
            suf = PurePosixPath(n).suffix
            if suf[1:].isdigit() and len(suf) == 4:
                candidates.append(n)

    for c in candidates:
        if c in name_set:
            return c, PurePosixPath(c).suffix
    return None


def infer_sequence_id_zip(json_member: str, fallback: str) -> str:
    pp = PurePosixPath(json_member)
    parts = list(pp.parts)
    if "recordings" in parts:
        idx = parts.index("recordings")
        if idx >= 1:
            return parts[idx - 1]
    if len(parts) >= 2:
        return parts[-2]
    return fallback


def load_json_fs(p: Path) -> Any:
    with open(p, "r", encoding="utf-8") as f:
        return json.load(f)


def file_size_fs(p: Path) -> int:
    return int(p.stat().st_size)


def sample_count_fs(p: Path, ext: str) -> int:
    ext = (ext or "").lower()
    if ext == ".npy":
        arr = np.load(p, allow_pickle=False)
        return int(getattr(arr, "size", 0))
    return int(file_size_fs(p) // 4)


def infer_sequence_id_fs(json_path: Path, fallback: str) -> str:
    parts = list(json_path.parts)
    if "recordings" in parts:
        idx = parts.index("recordings")
        if idx >= 1:
            return parts[idx - 1]
    if len(parts) >= 2:
        return parts[-2]
    return fallback


def find_data_file_fs(json_path: Path) -> Optional[Tuple[Path, str]]:
    if json_path.suffix.lower() != ".json":
        return None
    rec_dir = json_path.parent
    stem = json_path.stem

    exact = rec_dir / stem
    if exact.exists():
        suf = exact.suffix
        if suf and suf[1:].isdigit() and len(suf) == 4:
            return exact, suf

    candidates = [
        rec_dir / f"{stem}.npy",
        rec_dir / f"{stem}.bin",
        rec_dir / f"{stem}",
    ]
    for p in sorted(rec_dir.glob(f"{stem}.*")):
        suf = p.suffix
        if suf and suf[1:].isdigit() and len(suf) == 4:
            candidates.append(p)

    for c in candidates:
        if c.exists():
            return c, c.suffix
    return None


# ---------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------

LIGHT_RED_FILL = PatternFill(patternType="solid", fgColor="FFFFC7CE") # Excel-like light red (new rows)
MEDIUM_BLUE_FILL = PatternFill(patternType="solid", fgColor="FF1E90FF")  # ARGB


# Classic pure blue
# PURE_BLUE_FILL = PatternFill(patternType="solid", fgColor="FF0000FF")  # opaque blue
# 2) Medium blue (less intense)
# MEDIUM_BLUE_FILL = PatternFill(patternType="solid", fgColor="FF1E90FF")  # DodgerBlue
# 3) Light blue (what you currently had)
# LIGHT_BLUE_FILL = PatternFill(patternType="solid", fgColor="FFEAF2FF")  # very light blue
# Pure yellow
# PURE_YELLOW_FILL = PatternFill(patternType="solid", fgColor="FFFFFF00")  # opaque yellow
# 2) Light yellow (Excel-like highlight)
# LIGHT_YELLOW_FILL = PatternFill(patternType="solid", fgColor="FFFFF2CC")  # light yellow
# 3) Softer pastel yellow
# SOFTER_YELLOW_FILL = PatternFill(patternType="solid", fgColor="FFFFF9C4")



def norm_basename(x: Any) -> Optional[str]:
    """Normalize basenames for matching: ',' -> '.', strip, and drop trailing zeros in decimals."""
    if x is None:
        return None
    s = str(x).strip()
    if not s:
        return None
    s = s.replace(",", ".")
    m = re.fullmatch(r"(\d+)\.(\d+)", s)
    if m:
        a, b = m.group(1), m.group(2).rstrip("0")
        return a if b == "" else f"{a}.{b}"
    return s


def build_header_map(ws) -> Dict[str, int]:
    hdr: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        key = str(v).strip().lower()
        if key:
            hdr[key] = col
    return hdr


def pick_col(hdr: Dict[str, int], *names: str) -> Optional[int]:
    for n in names:
        k = n.strip().lower()
        if k in hdr:
            return hdr[k]
    return None


def cell_changed(old: Any, new: Any) -> bool:
    """Loose compare to avoid false mismatches (e.g. '12,3' vs 12.3)."""
    def norm(v: Any) -> Any:
        if v is None:
            return None
        if isinstance(v, bool):
            return bool(v)
        if isinstance(v, int):
            return int(v)
        if isinstance(v, float):
            if math.isnan(v):
                return None
            return float(v)
        s = str(v).strip()
        try:
            return float(s.replace(",", "."))
        except Exception:
            return s
    return norm(old) != norm(new)


def _flags_to_cell_value(flags_list: List[str]) -> str:
    # Keep Excel cell as a simple string.
    # Use '; ' to keep it readable and stable.
    return "; ".join([f for f in flags_list if f])



def _infer_total_samples_from_meta(meta: Dict[str, Any], fs: int) -> Optional[int]:
    """Best-effort total sample count from JSON metadata (when data file is missing)."""
    if not isinstance(meta, dict):
        return None
    for k in ("samples", "sampleCount", "sample_count", "n_samples", "signalSamples", "signalLength"):
        v = meta.get(k)
        if isinstance(v, int) and v > 0:
            return int(v)
        if isinstance(v, (float, np.floating)) and float(v).is_integer() and float(v) > 0:
            return int(v)
    for k in ("duration_s", "durationSec", "duration", "dur_s"):
        v = meta.get(k)
        if isinstance(v, (int, float, np.floating)) and float(v) > 0 and fs > 0:
            return int(round(float(v) * fs))
    return None

def _extract_from_json_only(meta: Dict[str, Any], fs: int) -> Dict[str, Any]:
    """
    Extract everything we can from JSON metadata alone (no samples/duration/fractions without data).
    Returns a dict with keys aligned to requested columns.
    """
    recording_id = meta.get("recordingId") if isinstance(meta.get("recordingId"), str) else None
    user_id = meta.get("userId") if isinstance(meta.get("userId"), str) else None
    flags_list = _flatten_flags(meta.get("flags"))

    comment = meta.get("comment") if isinstance(meta.get("comment"), str) else None
    rpeaks_any = meta.get("rpeaks")
    rpk_cnt, *_ = _summarize_rpeaks(rpeaks_any)

    mrk_counts_any = meta.get("rpeakAnnotationCounts")
    # Pick counts for the "H" columns (hN/hS/hV/hU).
    # Prefer explicit "merged" counts, then "human", then derive from rpeaks, then flat dict.
    def _counts_for_h_columns() -> Dict[str, int]:
        for var in ("merged", "human"):
            d = _extract_mrk_counts_for_variant(mrk_counts_any, var)
            if d:
                return d
        for var in ("merged", "human"):
            rp_list = _extract_rpeaks_list_for_variant(rpeaks_any, var)
            if rp_list:
                tmp: Dict[str, int] = {}
                for rp in rp_list:
                    av = rp.get("annotationValue")
                    if isinstance(av, str):
                        tmp[av] = tmp.get(av, 0) + 1
                    elif av is not None:
                        s = str(av)
                        tmp[s] = tmp.get(s, 0) + 1
                if tmp:
                    return tmp
        flat = _extract_mrk_counts(mrk_counts_any)
        if flat:
            return flat
        return {}

    h_counts_primary = _counts_for_h_columns()

    ml_counts = _extract_mrk_counts_for_variant(mrk_counts_any, "ml")
    if not ml_counts:
        ml_rpeaks = _extract_rpeaks_list_for_variant(rpeaks_any, "ml")
        tmp2: Dict[str, int] = {}
        for rp in ml_rpeaks:
            av = rp.get("annotationValue")
            if isinstance(av, str):
                tmp2[av] = tmp2.get(av, 0) + 1
            elif av is not None:
                tmp2[str(av)] = tmp2.get(str(av), 0) + 1
        ml_counts = tmp2

    h_n = int(h_counts_primary.get("N", 0)) if h_counts_primary else 0
    h_s = int(h_counts_primary.get("S", 0)) if h_counts_primary else 0
    h_v = int(h_counts_primary.get("V", 0)) if h_counts_primary else 0
    h_u = int(h_counts_primary.get("U", 0)) if h_counts_primary else 0

    ml_s = int(ml_counts.get("S", 0)) if ml_counts else 0
    ml_v = int(ml_counts.get("V", 0)) if ml_counts else 0
    ml_u = int(ml_counts.get("U", 0)) if ml_counts else 0

    h_noises = meta.get("noises_annotated")
    h_nz_cnt, _h_nz_samples = _sum_noise_samples(h_noises, fs)
    
    ml_noises = meta.get("noises")
    ml_nz_cnt, _ml_nz_samples = _sum_noise_samples(ml_noises, fs)

    total_samples = _infer_total_samples_from_meta(meta, fs)
    h_nz_frac = (float(_h_nz_samples) / total_samples * 100.0) if (total_samples and total_samples > 0) else None
    ml_nz_frac = (float(_ml_nz_samples) / total_samples * 100.0) if (total_samples and total_samples > 0) else None


    return {
        "rec_id": recording_id,
        "uid": user_id,
        "samples": int(total_samples) if total_samples is not None else None,
        "cmt": comment,
        "dur_s": (float(total_samples) / fs) if (total_samples is not None and fs and fs > 0) else None,
        "rpk_cnt": int(rpk_cnt),
        "hN": int(h_n),
        "hS": int(h_s),
        "hV": int(h_v),
        "hU": int(h_u),
        "mlS": int(ml_s),
        "mlV": int(ml_v),
        "mlU": int(ml_u),
        "h_nz_cnt": int(h_nz_cnt),
        "h_nz_len": int(_h_nz_samples) if _h_nz_samples is not None else None,
        "h_nz_frac": float(h_nz_frac) if h_nz_frac is not None else None,
        "ml_nz_cnt": int(ml_nz_cnt),
        "ml_nz_len": int(_ml_nz_samples) if _ml_nz_samples is not None else None,
        "ml_nz_frac": float(ml_nz_frac) if ml_nz_frac is not None else None,
        "flags": _flags_to_cell_value(flags_list),
    }


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", type=Path, required=True, help="Input Excel (.xlsx)")
    ap.add_argument("--zip", dest="zip_path", type=Path, required=True, help="ZIP file OR a folder containing JSON metadata and data files")
    ap.add_argument("--fs", type=int, default=200, help="Sampling frequency (Hz). Default: 200")
    ap.add_argument("--sheet", type=str, default=None, help="Sheet name (default: active sheet)")
    ap.add_argument("--out", type=Path, default=None, help="Output Excel path (default: <excel>_updated.xlsx)")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.excel)
    ws = wb[args.sheet] if args.sheet else wb.active
    
    # Print header row and first row for debugging
    # values = ["" if cell.value is None else str(cell.value) for cell in ws[1]]
    # print("\t".join(values))
    # values = ["" if cell.value is None else str(cell.value) for cell in ws[10]]
    # print("\t".join(values))

    hdr = build_header_map(ws)

    # Required columns for this script (with some aliases)
    cols: Dict[str, int] = {}
    cols["basename"] = pick_col(hdr, "basename")
    cols["rec_id"] = pick_col(hdr, "rec_id", "recordingid", "recording_id", "recording id")
    cols["uid"] = pick_col(hdr, "uid", "user_id", "userid", "user id")
    cols["samples"] = pick_col(hdr, "samples", "sample_cnt", "sample_count")
    cols["cmt"] = pick_col(hdr, "cmt", "comment")
    cols["dur_s"] = pick_col(hdr, "dur_s", "duration_s", "duration", "dur")
    cols["rpk_cnt"] = pick_col(hdr, "rpk_cnt", "rpeaks_cnt", "rpeaks_count", "rpeaks")
    cols["hN"] = pick_col(hdr, "hn", "hN")
    cols["hS"] = pick_col(hdr, "hs", "hS")
    cols["hV"] = pick_col(hdr, "hv", "hV")
    cols["hU"] = pick_col(hdr, "hu", "hU")
    cols["mlS"] = pick_col(hdr, "mls", "mlS")
    cols["mlV"] = pick_col(hdr, "mlv", "mlV")
    cols["mlU"] = pick_col(hdr, "mlu", "mlU")
    cols["h_nz_cnt"] = pick_col(hdr, "h_nz_cnt", "h_nz_count")
    cols["h_nz_len"] = pick_col(hdr, "h_nz_len", "h_nz_length", "h_nz_len_s", "h_nz_seconds")
    cols["h_nz_frac"] = pick_col(hdr, "h_nz_frac", "h_nz_frac%", "h_nz_frac %")
    cols["noni"] = pick_col(hdr, "noni")
    cols["ml_nz_cnt"] = pick_col(hdr, "ml_nz_cnt", "ml_nz_count")
    cols["ml_nz_len"] = pick_col(hdr, "ml_nz_len", "ml_nz_length", "ml_nz_len_s", "ml_nz_seconds")
    cols["ml_nz_frac"] = pick_col(hdr, "ml_nz_frac", "ml_nz_frac%", "ml_nz_frac %")
    cols["flags"] = pick_col(hdr, "flags")

    # print(f"Column mapping: {cols}")
    
    # Only 'basename' is strictly required. Other columns are optional; if missing in the Excel,
    # the script will simply not populate them.
    if cols["basename"] is None:
        raise RuntimeError("Missing required column in Excel header row: ['basename']")

    c_basename = cols["basename"]

    # Map basename -> row index
    basename_to_row: Dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        bn = norm_basename(ws.cell(row=r, column=c_basename).value)
        if bn:
            basename_to_row[bn] = r

    # Style template row for newly inserted rows (copy formatting from row 2 if present)
    style_row_idx = 2 if ws.max_row >= 2 else None

    def copy_row_style(dst_row: int) -> None:
        if style_row_idx is None:
            return
        for c in range(1, ws.max_column + 1):
            src_cell = ws.cell(row=style_row_idx, column=c)
            dst_cell = ws.cell(row=dst_row, column=c)
            dst_cell._style = _copy_style(src_cell._style)
            dst_cell.font = _copy_style(src_cell.font)
            dst_cell.border = _copy_style(src_cell.border)
            dst_cell.fill = _copy_style(src_cell.fill)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = _copy_style(src_cell.protection)
            dst_cell.alignment = _copy_style(src_cell.alignment)

    new_rows: Set[int] = set()
    changed_rows: Set[int] = set()
    processed = 0

    def get_or_create_row(bn: str) -> Tuple[int, bool]:
        """Return (row_index, is_new). If basename not found, append a new styled row and set basename."""
        row = basename_to_row.get(bn)
        if row is not None:
            return row, False

        row = ws.max_row + 1
        copy_row_style(row)
        ws.cell(row=row, column=c_basename).value = bn
        basename_to_row[bn] = row
        new_rows.add(row)
        # Mark NEW rows: basename cell in light red
        ws.cell(row=row, column=c_basename).fill = LIGHT_RED_FILL
        return row, True

    def set_cell(row: int, col_key: str, new_val: Any) -> bool:

        col = cols.get(col_key)
        if col is None:
            return False
        cell = ws.cell(row=row, column=col)
        old_val = cell.value

        # Normalize percent columns to 1 decimal (when numeric)
        if col_key in ("h_nz_frac", "ml_nz_frac") and isinstance(new_val, (int, float, np.floating)):
            new_val = round(float(new_val), 1)

        if col_key in ("h_nz_len","ml_nz_len") and isinstance(new_val, (int, float, np.floating)):
            new_val = round(float(new_val), 1)

        if cell_changed(old_val, new_val):
            cell.value = new_val
            return True
        return False

    src = args.zip_path
    print(f"Source: {src} (is_dir={src.is_dir()}, is_file={src.is_file()})")

    if src.is_dir():
        json_paths = [p for p in src.rglob("*.json") if "__MACOSX" not in p.parts and p.is_file()]
        print(f"Found JSON files: {len(json_paths)}")

        for jp in sorted(json_paths):
            bn = norm_basename(jp.stem)
            if not bn:
                continue
            if bn.lower() == "manifest":
                continue

            row, _is_new = get_or_create_row(bn)
            # print(f"Processing: {jp} -> row {row} (new: {_is_new})")

            data_info = find_data_file_fs(jp)
            # print(f"  Data file info: {data_info}")
            if data_info is None:
                meta = load_json_fs(jp)
                extracted = _extract_from_json_only(meta if isinstance(meta, dict) else {}, args.fs)
                # print(f"  No data file found for {jp}. Extracted from JSON only: {extracted}")
            else:
                data_path, data_ext = data_info
                seq = infer_sequence_id_fs(jp, fallback=src.name or "dir")
                rec, _rec_dict = summarize_record(
                    sequenceId=seq,
                    basename=bn,
                    json_ref=jp,
                    data_ref=data_path,
                    data_ext=data_ext,
                    load_json_fn=load_json_fs,
                    file_size_fn=file_size_fs,
                    sample_count_fn=sample_count_fs,
                    fs=args.fs,
                )
                # print(f"  Found data file for {jp}: {data_path}. Extracted: {rec}")
                extracted = {
                    "rec_id": rec.recordingId,
                    "uid": rec.user_id,
                    "samples": int(rec.samples),
                    "cmt": rec.cmt,
                    "dur_s": float(rec.duration_s) if rec.duration_s is not None else None,
                    "rpk_cnt": int(rec.rpeaks_count),
                    "hN": int(rec.h_n_count),
                    "hS": int(rec.h_s_count),
                    "hV": int(rec.h_v_count),
                    "hU": int(rec.h_u_count),
                    "mlS": int(rec.ml_s_count),
                    "mlV": int(rec.ml_v_count),
                    "mlU": int(rec.ml_u_count),
                    "h_nz_cnt": int(rec.h_noises_count),
                    "h_nz_len": int(_rec_dict.get("_h_noises_samples", 0)) if (args.fs and args.fs > 0 and _rec_dict.get("_h_noises_samples") is not None) else None,
                    "noni": int(rec.h_noises_count),
                    "h_nz_frac": float(rec.h_noises_fraction) if rec.h_noises_fraction is not None else None,
                    "ml_nz_cnt": int(rec.ml_noises_count),
                    "ml_nz_len": int(_rec_dict.get("_ml_noises_samples", 0)) if (args.fs and args.fs > 0 and _rec_dict.get("_ml_noises_samples") is not None) else None,
                    "ml_nz_frac": float(rec.ml_noises_fraction) if rec.ml_noises_fraction is not None else None,
                    "flags": _flags_to_cell_value(rec.flags or []),
                }

            row_changed = False
            # basename cell already set in get_or_create_row (and styled if new)

            for k, v in extracted.items():
                if v is None:
                    continue
                row_changed |= set_cell(row, k, v)
                if k == "h_nz_cnt":
                    # Keep legacy column in sync if present
                    row_changed |= set_cell(row, "noni", v)

            if row_changed:
                changed_rows.add(row)

            processed += 1

    else:
        if not src.is_file() or src.suffix.lower() != ".zip":
            raise FileNotFoundError(f"--zip must be a .zip file or a directory. Got: {src}")

        with zipfile.ZipFile(src, "r") as zf:
            name_set = set(zf.namelist())
            json_members = [
                n for n in name_set
                if n.lower().endswith(".json")
                and not n.endswith("/")
                and not n.startswith("__MACOSX/")
            ]
            print(f"Found JSON files: {len(json_members)}")

            for jm in sorted(json_members):
                stem = Path(jm).name[:-5]  # drop ".json"
                bn = norm_basename(stem)
                if not bn:
                    continue
                if bn.lower() == "manifest":
                    continue

                row, _is_new = get_or_create_row(bn)

                data_info = find_data_file_zip(jm, name_set)
                if data_info is None:
                    meta = load_json_zip(zf, jm)
                    extracted = _extract_from_json_only(meta if isinstance(meta, dict) else {}, args.fs)
                else:
                    data_member, data_ext = data_info
                    seq = infer_sequence_id_zip(jm, fallback="zip")
                    rec, _rec_dict = summarize_record(
                        sequenceId=seq,
                        basename=bn,
                        json_ref=jm,
                        data_ref=data_member,
                        data_ext=data_ext,
                        load_json_fn=lambda m: load_json_zip(zf, str(m)),
                        file_size_fn=lambda m: file_size_zip(zf, str(m)),
                        sample_count_fn=lambda m, ext: sample_count_zip(zf, str(m), ext),
                        fs=args.fs,
                    )
                    extracted = {
                        "rec_id": rec.recordingId,
                        "uid": rec.user_id,
                        "samples": int(rec.samples),
                        "cmt": rec.cmt,
                        "dur_s": float(rec.duration_s) if rec.duration_s is not None else None,
                        "rpk_cnt": int(rec.rpeaks_count),
                        "hN": int(rec.h_n_count),
                        "hS": int(rec.h_s_count),
                        "hV": int(rec.h_v_count),
                        "hU": int(rec.h_u_count),
                        "mlS": int(rec.ml_s_count),
                        "mlV": int(rec.ml_v_count),
                        "mlU": int(rec.ml_u_count),
                        "h_nz_cnt": int(rec.h_noises_count),
                    "h_nz_len": float(_rec_dict.get("_h_noises_samples", 0)) / args.fs if (args.fs and args.fs > 0 and _rec_dict.get("_h_noises_samples") is not None) else None,
                    "noni": int(rec.h_noises_count),
                        "h_nz_frac": float(rec.h_noises_fraction) if rec.h_noises_fraction is not None else None,
                        "ml_nz_cnt": int(rec.ml_noises_count),
                        "ml_nz_frac": float(rec.ml_noises_fraction) if rec.ml_noises_fraction is not None else None,
                        "flags": _flags_to_cell_value(rec.flags or []),
                    }

                row_changed = False
                for k, v in extracted.items():
                    if v is None:
                        continue
                    row_changed |= set_cell(row, k, v)
                    if k == "h_nz_cnt":
                        # Keep legacy column in sync if present
                        row_changed |= set_cell(row, "noni", v)

                if row_changed:
                    changed_rows.add(row)

                processed += 1

    # Highlight basename cell for changed EXISTING rows (keep red for new rows)
    for r in changed_rows:
        if r in new_rows:
            continue
        ws.cell(row=r, column=c_basename).fill = MEDIUM_BLUE_FILL

    out_path = args.out if args.out else args.excel.with_name(args.excel.stem + "_updated.xlsx")
    wb.save(out_path)

    print(f"Processed JSON files: {processed}")
    print(f"Changed rows: {len(changed_rows)}")
    print(f"Added new rows: {len(new_rows)}")
    if new_rows:
        ex_added = [ws.cell(row=r, column=c_basename).value for r in sorted(new_rows)[:10]]
        print("Examples added:", ex_added)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
