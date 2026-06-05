#!/usr/bin/env python3
"""
Build an Excel summary table from JSON metadata files only.

This is the metadata-only counterpart of analyze_records_2026_v7.py. It does
not import or call ECG loading, denoising, or ectopy pipeline helpers. Values
are extracted from JSON metadata, and the output workbook is written to the
data directory unless --out is supplied.
"""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass, fields
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


@dataclass
class RecordSummary:
    sequenceId: str
    recordingId: Optional[str]
    basename: str
    user_id: Optional[str]
    samples: Optional[int]
    duration_s: Optional[float]
    cmt: Optional[str]
    flags_count: int
    flags: List[str]
    rpeaks_count: int
    h_n_count: int
    h_s_count: int
    h_v_count: int
    h_u_count: int
    ml_s_count: int
    ml_v_count: int
    ml_u_count: int
    json_ok: bool
    ml_noises_count: int
    ml_noises_samples: int
    ml_noises_fraction: Optional[float]
    h_noises_count: int
    h_noises_samples: int
    h_noises_fraction: Optional[float]
    json_keys_correct: bool


@dataclass
class OutputRow:
    basename: str
    samples: Optional[int]
    tag: Optional[str]
    cmt: Optional[str]
    mark: Optional[str]
    rpk_cnt: int
    hN: int
    hS: int
    hV: int
    hU: int
    h_nz_cnt: int
    h_nz_len: int
    h_nz_frac_pct: Optional[float]
    mlS: int
    mlV: int
    mlU: int
    ml_nz_cnt: int
    ml_nz_len: int
    ml_nz_frac_pct: Optional[float]
    flags: Optional[str]
    recordingId: Optional[str]
    userId: Optional[str]
    notes: Optional[str]


# Single source of truth for Excel column order:
# (header shown in Excel, OutputRow attribute name)
RECORD_COLUMNS: List[Tuple[str, str]] = [
    ("basename", "basename"),
    ("samples", "samples"),
    ("tag", "tag"),
    ("cmt", "cmt"),
    ("mark", "mark"),
    ("rpk_cnt", "rpk_cnt"),
    ("hN", "hN"),
    ("hS", "hS"),
    ("hV", "hV"),
    ("hU", "hU"),
    ("h_nz_cnt", "h_nz_cnt"),
    ("h_nz_len", "h_nz_len"),
    ("h_nz_frac%", "h_nz_frac_pct"),
    ("mlS", "mlS"),
    ("mlV", "mlV"),
    ("mlU", "mlU"),
    ("ml_nz_cnt", "ml_nz_cnt"),
    ("ml_nz_len", "ml_nz_len"),
    ("ml_nz_frac%", "ml_nz_frac_pct"),
    ("flags", "flags"),
    ("recordingId", "recordingId"),
    ("userId", "userId"),
    ("notes", "notes"),
]

SAMPLE_KEYS: Tuple[str, ...] = (
    "samples",
    "sampleCount",
    "sample_count",
    "signalLength",
    "signal_length",
    "lengthSamples",
    "length_samples",
    "totalSamples",
    "total_samples",
    "numberOfSamples",
    "number_of_samples",
    "nSamples",
    "n_samples",
)

DURATION_KEYS: Tuple[str, ...] = (
    "duration_s",
    "durationSec",
    "durationSeconds",
    "duration",
)


def _flatten_flags(flags: Any) -> List[str]:
    if flags is None:
        return []
    if isinstance(flags, list):
        return [str(x) for x in flags]
    return [str(flags)]


def _pick_variant_dict(obj: Any, keys_preference: Tuple[str, ...] = ("merged", "human", "ml")) -> Any:
    if not isinstance(obj, dict):
        return None
    for key in keys_preference:
        if key in obj:
            return obj.get(key)
    for key, value in obj.items():
        if key != "__info":
            return value
    return None


def _as_int(value: Any) -> Optional[int]:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return int(value)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return None


def _as_float(value: Any) -> Optional[float]:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _extract_rpeaks_list(rpeaks: Any) -> List[Dict[str, Any]]:
    if isinstance(rpeaks, list):
        return [item for item in rpeaks if isinstance(item, dict)]
    if isinstance(rpeaks, dict):
        value = _pick_variant_dict(rpeaks, ("merged", "human", "ml"))
        if isinstance(value, list):
            return [item for item in value if isinstance(item, dict)]
    return []


def _extract_rpeaks_list_for_variant(rpeaks: Any, variant: str) -> List[Dict[str, Any]]:
    if isinstance(rpeaks, dict):
        value = rpeaks.get(variant)
        if isinstance(value, list):
            return [item for item in value if isinstance(item, dict)]
    return []


def _extract_mrk_counts_for_variant(counts: Any, variant: str) -> Dict[str, int]:
    if not isinstance(counts, dict):
        return {}
    value = counts.get(variant)
    if not isinstance(value, dict):
        return {}
    out: Dict[str, int] = {}
    for key, item in value.items():
        if key == "__info":
            continue
        int_value = _as_int(item)
        if int_value is not None:
            out[str(key)] = int_value
    return out


def _extract_flat_mrk_counts(counts: Any) -> Dict[str, int]:
    if not isinstance(counts, dict):
        return {}
    out: Dict[str, int] = {}
    for key, item in counts.items():
        if key == "__info" or isinstance(item, dict):
            continue
        int_value = _as_int(item)
        if int_value is not None:
            out[str(key)] = int_value
    return out


def _counts_from_rpeaks(rpeaks: List[Dict[str, Any]]) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for rpeak in rpeaks:
        annotation = rpeak.get("annotationValue")
        if annotation is not None:
            key = str(annotation)
            out[key] = out.get(key, 0) + 1
    return out


def _extract_h_counts(counts: Any, rpeaks: Any) -> Dict[str, int]:
    for variant in ("merged", "human"):
        variant_counts = _extract_mrk_counts_for_variant(counts, variant)
        if variant_counts:
            return variant_counts

    for variant in ("merged", "human"):
        variant_rpeaks = _extract_rpeaks_list_for_variant(rpeaks, variant)
        if variant_rpeaks:
            return _counts_from_rpeaks(variant_rpeaks)

    return _extract_flat_mrk_counts(counts)


def _summarize_rpeaks(rpeaks: Any) -> Tuple[int, Optional[int]]:
    items = _extract_rpeaks_list(rpeaks)
    max_idx: Optional[int] = None
    for item in items:
        sample_index = _as_int(item.get("sampleIndex"))
        if sample_index is not None:
            max_idx = sample_index if max_idx is None else max(max_idx, sample_index)
    return len(items), max_idx


def _iter_noise_items(noises: Any) -> List[Any]:
    if isinstance(noises, dict):
        noises = _pick_variant_dict(noises, ("merged", "human", "ml"))
    if isinstance(noises, list):
        return noises
    return []


def _sum_noise_samples(noises: Any, fs: Optional[int]) -> Tuple[int, int, Optional[int]]:
    count = 0
    total = 0
    max_idx: Optional[int] = None

    for item in _iter_noise_items(noises):
        start = end = None
        if isinstance(item, dict):
            start = item.get("startIndex", item.get("startSample"))
            end = item.get("endIndex", item.get("endSample"))
            if (start is None or end is None) and fs:
                start_time = _as_float(item.get("startTime"))
                end_time = _as_float(item.get("endTime"))
                if start_time is not None and end_time is not None and end_time > start_time:
                    start = int(round(start_time * fs))
                    end = int(round(end_time * fs))
        elif isinstance(item, (list, tuple)) and len(item) >= 2:
            start, end = item[0], item[1]

        start_i = _as_int(start)
        end_i = _as_int(end)
        if start_i is not None and end_i is not None and end_i > start_i:
            count += 1
            total += end_i - start_i
            max_idx = end_i if max_idx is None else max(max_idx, end_i)

    return count, total, max_idx


def _extract_explicit_samples(meta: Dict[str, Any]) -> Optional[int]:
    for key in SAMPLE_KEYS:
        sample_count = _as_int(meta.get(key))
        if sample_count is not None and sample_count >= 0:
            return sample_count
    return None


def _extract_duration(meta: Dict[str, Any]) -> Optional[float]:
    for key in DURATION_KEYS:
        duration = _as_float(meta.get(key))
        if duration is not None and duration >= 0:
            return duration
    return None


def _infer_samples_from_metadata(
    meta: Dict[str, Any],
    *,
    fs: Optional[int],
    rpeaks_max_idx: Optional[int],
    ml_noise_max_idx: Optional[int],
    h_noise_max_idx: Optional[int],
) -> Optional[int]:
    explicit_samples = _extract_explicit_samples(meta)
    if explicit_samples is not None:
        return explicit_samples

    duration = _extract_duration(meta)
    if duration is not None and fs and fs > 0:
        return int(round(duration * fs))

    max_idx = None
    for value in (rpeaks_max_idx, ml_noise_max_idx, h_noise_max_idx):
        if value is not None:
            max_idx = value if max_idx is None else max(max_idx, value)
    return max_idx + 1 if max_idx is not None else None


def _flags_to_cell_value(flags_list: List[str]) -> str:
    return "; ".join([flag for flag in flags_list if flag])


def _validate_record_columns() -> None:
    outputrow_fields = {field.name for field in fields(OutputRow)}
    missing = [attr_name for _header, attr_name in RECORD_COLUMNS if attr_name not in outputrow_fields]
    if missing:
        raise ValueError(f"RECORD_COLUMNS references missing OutputRow fields: {missing}")


def _record_headers() -> List[str]:
    return [header for header, _attr_name in RECORD_COLUMNS]


def _record_values(row: OutputRow) -> List[Any]:
    return [getattr(row, attr_name) for _header, attr_name in RECORD_COLUMNS]


def load_json_fs(path: Path) -> Any:
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def infer_sequence_id_fs(json_path: Path, fallback: str) -> str:
    parts = list(json_path.parts)
    if "recordings" in parts:
        idx = parts.index("recordings")
        if idx >= 1:
            return parts[idx - 1]
    if len(parts) >= 2:
        return parts[-2]
    return fallback


def summarize_metadata(sequenceId: str, basename: str, json_ref: Path, *, fs: Optional[int]) -> RecordSummary:
    try:
        meta_any = load_json_fs(json_ref)
        json_ok = True
    except Exception as exc:
        print(f"WARN: failed to load JSON for {sequenceId}/{basename}: {exc}")
        meta_any = {}
        json_ok = False

    meta = meta_any if isinstance(meta_any, dict) else {}
    user_id = meta.get("userId")
    recordingId = meta.get("recordingId")
    flags_list = _flatten_flags(meta.get("flags"))
    comment = meta.get("comment")

    rpeaks_any = meta.get("rpeaks")
    rpeaks_count, rpeaks_max_idx = _summarize_rpeaks(rpeaks_any)

    counts_any = meta.get("rpeakAnnotationCounts")
    h_counts = _extract_h_counts(counts_any, rpeaks_any)
    h_n = int(h_counts.get("N", 0))
    h_s = int(h_counts.get("S", 0))
    h_v = int(h_counts.get("V", 0))
    h_u = int(h_counts.get("U", 0))

    ml_counts = _extract_mrk_counts_for_variant(counts_any, "ml")
    if not ml_counts:
        ml_counts = _counts_from_rpeaks(_extract_rpeaks_list_for_variant(rpeaks_any, "ml"))

    ml_s = int(ml_counts.get("S", 0))
    ml_v = int(ml_counts.get("V", 0))
    ml_u = int(ml_counts.get("U", 0))

    h_noises = meta.get("noises_annotated")
    h_nz_cnt, h_nz_samples, h_noise_max_idx = _sum_noise_samples(h_noises, fs)

    ml_noises = meta.get("noises")
    ml_nz_cnt, ml_nz_samples, ml_noise_max_idx = _sum_noise_samples(ml_noises, fs)

    samples = _infer_samples_from_metadata(
        meta,
        fs=fs,
        rpeaks_max_idx=rpeaks_max_idx,
        ml_noise_max_idx=ml_noise_max_idx,
        h_noise_max_idx=h_noise_max_idx,
    )
    duration_s = _extract_duration(meta)
    if duration_s is None and samples is not None and fs and fs > 0:
        duration_s = samples / fs

    ml_noises_fraction = (ml_nz_samples / samples) * 100.0 if samples and samples > 0 else None
    h_noises_fraction = (h_nz_samples / samples) * 100.0 if samples and samples > 0 else None

    allowed_keys = {
        "channelCount",
        "comment",
        "flags",
        "noises",
        "noises_annotated",
        "recordingId",
        "rpeakAnnotationCounts",
        "rpeaks",
        "userId",
        *SAMPLE_KEYS,
        *DURATION_KEYS,
    }
    meta_keys = set(meta.keys())
    json_keys_correct = meta_keys.issubset(allowed_keys) if meta_keys else False

    return RecordSummary(
        sequenceId=str(sequenceId),
        recordingId=str(recordingId) if isinstance(recordingId, str) else None,
        basename=str(basename),
        user_id=str(user_id) if isinstance(user_id, str) else None,
        samples=samples,
        duration_s=float(duration_s) if duration_s is not None else None,
        cmt=str(comment) if isinstance(comment, str) else None,
        flags_count=len(flags_list),
        flags=flags_list,
        rpeaks_count=int(rpeaks_count),
        h_n_count=h_n,
        h_s_count=h_s,
        h_v_count=h_v,
        h_u_count=h_u,
        ml_s_count=int(ml_s),
        ml_v_count=int(ml_v),
        ml_u_count=int(ml_u),
        json_ok=bool(json_ok),
        h_noises_count=int(h_nz_cnt),
        h_noises_samples=int(h_nz_samples),
        h_noises_fraction=float(h_noises_fraction) if h_noises_fraction is not None else None,
        ml_noises_count=int(ml_nz_cnt),
        ml_noises_samples=int(ml_nz_samples),
        ml_noises_fraction=float(ml_noises_fraction) if ml_noises_fraction is not None else None,
        json_keys_correct=bool(json_keys_correct),
    )


def build_output_row(rec: RecordSummary) -> OutputRow:
    return OutputRow(
        basename=rec.basename,
        samples=int(rec.samples) if rec.samples is not None else None,
        tag=None,
        cmt=rec.cmt,
        mark=None,
        rpk_cnt=int(rec.rpeaks_count),
        hN=int(rec.h_n_count),
        hS=int(rec.h_s_count),
        hV=int(rec.h_v_count),
        hU=int(rec.h_u_count),
        h_nz_cnt=int(rec.h_noises_count),
        h_nz_len=int(rec.h_noises_samples),
        h_nz_frac_pct=round(float(rec.h_noises_fraction), 1) if rec.h_noises_fraction is not None else None,
        mlS=int(rec.ml_s_count),
        mlV=int(rec.ml_v_count),
        mlU=int(rec.ml_u_count),
        ml_nz_cnt=int(rec.ml_noises_count),
        ml_nz_len=int(rec.ml_noises_samples),
        ml_nz_frac_pct=round(float(rec.ml_noises_fraction), 1) if rec.ml_noises_fraction is not None else None,
        flags=_flags_to_cell_value(rec.flags or []),
        recordingId=rec.recordingId,
        userId=rec.user_id,
        notes=None,
    )


def auto_adjust_widths(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)


def write_output_workbook(rows: List[OutputRow], out_path: Path) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "Records"

    _validate_record_columns()
    ws.append(_record_headers())

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(_record_values(row))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    auto_adjust_widths(ws)

    ws_marks = wb.create_sheet(title="Marks")
    marks_text = {
        "B2": "Stulpelis tag:",
        "B3": "tag = 1111",
        "C3": "Kokybe gera, tinka mokymui",
        "B4": "tag = 11110",
        "C4": "Kokybe gera, pretendentas mokymui, bus patvirtinta po anotavimo",
        "B6": "tag = 2222",
        "C6": "Kokybe nebloga, taciau ant ribos, gal labiau tinka triuksmu testavimui",
        "B7": "tag = 22220",
        "C7": "Pretendentas testavimui, bus patvirtinta po anotavimo",
        "B9": "tag = 3333",
        "C9": "Gana daug triuksmu, gali tikti triuksmu testavimui",
        "B10": "tag = 33330",
        "C10": "Gana daug triuksmu, gali pretenduoti triuksmu testavimui, bus patvirtinta po anotavimo",
        "B12": "tag = 5555",
        "C12": "Ypatingi, 'keisti' atvejai",
        "B13": "tag = 9999",
        "C13": "Kurie niekam netinka",
        "B16": "Stulpelis cmt:",
        "B17": "N, J, Z",
        "C17": "Nika anotavo triuksmu intervalus, J - Jonas anotavo ekstrasistoles, Z - Zygimantas anotavo ekstrasistoles",
        "B19": "Stulpelis mark:",
        "B20": "excl",
        "C20": "eliminuojamas irasas is mokymo ir testines imciu",
        "B21": "ect8",
        "C21": "sarasas ektopiniu duziu anotavimui",
        "B22": "nz8",
        "C22": "sarasas triuksmu zymejimui",
        "B23": "ect8.nz8",
        "C23": "abu sarasai",
    }

    for cell_ref, value in marks_text.items():
        ws_marks[cell_ref] = value
    for cell_ref in ("B3", "B16", "B19"):
        ws_marks[cell_ref].font = Font(bold=True)
    for row in ws_marks.iter_rows(min_row=1, max_row=30, min_col=2, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws_marks.column_dimensions["A"].width = 4
    ws_marks.column_dimensions["B"].width = 18
    ws_marks.column_dimensions["C"].width = 110
    for row_idx in (16, 20, 21, 22, 23):
        ws_marks.row_dimensions[row_idx].height = 30

    wb.create_sheet(title="Notes")
    wb.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dir", type=Path, required=True, help="Directory containing JSON metadata files")
    parser.add_argument("--fs", type=int, default=200, help="Sampling frequency in Hz for time-based metadata. Default: 200")
    parser.add_argument("--out", type=Path, default=None, help="Output Excel path")
    args = parser.parse_args()

    src = args.dir
    if not src.exists() or not src.is_dir():
        raise FileNotFoundError(f"--dir must be an existing directory. Got: {src}")

    json_paths = [path for path in src.rglob("*.json") if "__MACOSX" not in path.parts and path.is_file()]
    print(f"Source: {src}")
    print(f"Found JSON files: {len(json_paths)}")

    output_rows: List[OutputRow] = []
    processed = 0

    for json_path in sorted(json_paths):
        basename = json_path.stem
        if not basename or basename.lower() == "manifest":
            continue

        sequence_id = infer_sequence_id_fs(json_path, fallback=src.name or "dir")
        rec = summarize_metadata(sequenceId=sequence_id, basename=basename, json_ref=json_path, fs=args.fs)
        output_rows.append(build_output_row(rec))
        processed += 1

    out_path = args.out if args.out else src / f"{src.name}_metadata_summary.xlsx"
    write_output_workbook(output_rows, out_path)

    print(f"Processed metadata files: {processed}")
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
