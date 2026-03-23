#!/usr/bin/env python3
"""

https://chatgpt.com/c/69c0fe7b-cf84-838a-bdee-7a01e99c1d54

Build an Excel summary table directly from a directory of ECG data files and JSON metadata.

This script is tailored from update_records_list_v4.py.
It keeps the RecordSummary dataclass and the existing JSON/record summarization helpers,
but changes the overall workflow from "update an existing Excel workbook" to
"scan a directory, compute/export one output workbook from scratch".

Input
-----
A directory containing JSON metadata files and accompanying ECG data files
(e.g. .npy, .bin, or files with 4-digit numeric suffixes).

Output
------
A new .xlsx file with columns in this exact order:

filename, basename, samples, tag, cmt, rpk_cnt, hN, hS, hV, hU,
ectS, ectV, ectU, h_nz_cnt, h_nz_len, h_nz_frac%, out, rdr, noi, tp%,
ml_nz_cnt, ml_nz_len, ml_nz_frac%, flags, recordingId, userId, notes

Notes
-----
- Columns tag and notes are intentionally left empty for later manual editing.
- ECG-derived denoising statistics are calculated from:
      x = load_ecg_npy(data_path)
      res_denoising = pipe.run(x, gaps_indices=[])
      noise_stats = calc_noise_stats_from_denoised_result(res_denoising) or {}
- Ectopy columns ectS/ectV/ectU currently come from a placeholder function
  returning zeros and can be replaced later with a real implementation.
"""

from __future__ import annotations

import argparse
import json
from dataclasses import asdict, dataclass, fields
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple, Union

import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

from openpyxl.worksheet.dimensions import ColumnDimension
from record_noise_stats import calc_noise_stats_from_denoised_result

try:
    from ecg_denoising_pipeline import (
        ECGDenoisingPipeline,
        DenoisingPipelineConfig,
        DenoisingPipelineResult,
        load_denoising_config_yaml,
        load_ecg_npy,
        resolve_model_path,
    )
except Exception as exc:
    raise ImportError(
        "Cannot import ECG denoising pipeline helpers. "
        "Update imports to match your project structure.\n"
        f"Original error: {exc}"
    ) from exc

try:
    from ecg_denoising_pipeline.steps import check_denoising_config
except Exception:
    check_denoising_config = None

try:
    from ecg_ectopy_pipeline import (
        ECGEctopyPipeline,
        EctopyPipelineConfig,
        EctopyPipelineResult,
        load_ectopy_config_yaml,
    )
except Exception as exc:
    raise ImportError(
        "Cannot import ECG ectopy pipeline helpers. "
        "Update imports to match your project structure.\n"
        f"Original error: {exc}"
    ) from exc

JsonLikeRef = Union[str, Path]


@dataclass
class RecordSummary:
    sequenceId: str
    recordingId: Optional[str]
    basename: str
    channel_count: Optional[int]
    user_id: Optional[str]

    bin_bytes: int
    samples: int
    duration_s: Optional[float]

    flags_count: int
    flags: List[str]

    rpeaks_count: int
    h_n_count: int
    h_s_count: int
    h_v_count: int
    h_u_count: int

    ml_s_count: int
    ml_v_count: int
    ml_u_count: int

    json_ok: bool

    ml_noises_count: int
    ml_noises_samples: int
    ml_noises_fraction: Optional[float]  # percent 0..100

    h_noises_count: int
    h_noises_samples: int
    h_noises_fraction: Optional[float]  # percent 0..100

    cmt: Optional[str]
    json_keys_correct: bool
    notes: Optional[str] = None


@dataclass
class OutputRow:
    filename: Optional[str]
    basename: str
    samples: Optional[int]
    quality: Optional[int]
    tag: Optional[str]
    cmt: Optional[str]
    mark: Optional[str]
    rpk_cnt: Optional[int]
    hN: Optional[int]
    hS: Optional[int]
    hV: Optional[int]
    hU: Optional[int]
    mlS: Optional[int]
    mlV: Optional[int]
    mlU: Optional[int]
    ectN: int
    ectS: int
    ectV: int
    ectU: int
    h_nz_cnt: Optional[int]
    h_nz_len: Optional[int]
    h_nz_frac_pct: Optional[float]
    out: Optional[int]
    rdr: Optional[int]
    noi: Optional[int]
    tp_pct: Optional[float]
    ml_nz_cnt: Optional[int]
    ml_nz_len: Optional[int]
    ml_nz_frac_pct: Optional[float]
    flags: Optional[str]
    recordingId: Optional[str]
    userId: Optional[str]
    notes: Optional[str]


def _flatten_flags(flags: Any) -> List[str]:
    if flags is None:
        return []
    if isinstance(flags, list):
        return [str(x) for x in flags]
    return [str(flags)]


def _pick_variant_dict(obj: Any, keys_preference: Tuple[str, ...] = ("merged", "human", "ml")) -> Any:
    if not isinstance(obj, dict):
        return None
    for k in keys_preference:
        if k in obj:
            return obj.get(k)
    for k, v in obj.items():
        if k == "__info":
            continue
        return v
    return None


def _extract_rpeaks_list(rpeaks: Any) -> List[Dict[str, Any]]:
    if isinstance(rpeaks, list):
        return [x for x in rpeaks if isinstance(x, dict)]
    if isinstance(rpeaks, dict):
        v = _pick_variant_dict(rpeaks, ("merged", "human", "ml"))
        if isinstance(v, list):
            return [x for x in v if isinstance(x, dict)]
    return []


def _extract_rpeaks_list_for_variant(rpeaks: Any, variant: str) -> List[Dict[str, Any]]:
    if isinstance(rpeaks, dict):
        v = rpeaks.get(variant)
        if isinstance(v, list):
            return [x for x in v if isinstance(x, dict)]
    return []


def _extract_mrk_counts_for_variant(h_counts: Any, variant: str) -> Dict[str, int]:
    if not isinstance(h_counts, dict):
        return {}
    v = h_counts.get(variant)
    if not isinstance(v, dict):
        return {}
    out: Dict[str, int] = {}
    for k, val in v.items():
        if k == "__info":
            continue
        if isinstance(val, int):
            out[str(k)] = int(val)
        elif isinstance(val, (float, np.floating)) and float(val).is_integer():
            out[str(k)] = int(val)
    return out


def _extract_mrk_counts(h_counts: Any) -> Dict[str, int]:
    if not isinstance(h_counts, dict):
        return {}
    if any(isinstance(v, dict) for k, v in h_counts.items() if k != "__info"):
        v = _pick_variant_dict(h_counts, ("merged", "human", "ml"))
        if isinstance(v, dict):
            out: Dict[str, int] = {}
            for k, val in v.items():
                if k == "__info":
                    continue
                if isinstance(val, int):
                    out[str(k)] = int(val)
                elif isinstance(val, (float, np.floating)) and float(val).is_integer():
                    out[str(k)] = int(val)
            return out
        return {}

    out2: Dict[str, int] = {}
    for k, v in h_counts.items():
        if k == "__info":
            continue
        if isinstance(v, int):
            out2[str(k)] = int(v)
        elif isinstance(v, (float, np.floating)) and float(v).is_integer():
            out2[str(k)] = int(v)
    return out2


def _summarize_rpeaks(rpeaks: Any) -> Tuple[int, Optional[int], Optional[int], Dict[str, int]]:
    lst = _extract_rpeaks_list(rpeaks)
    if not lst:
        return 0, None, None, {}
    idxs: List[int] = []
    ann: Dict[str, int] = {}
    for rp in lst:
        si = rp.get("sampleIndex")
        av = rp.get("annotationValue")
        if isinstance(si, int):
            idxs.append(si)
        elif isinstance(si, (float, np.floating)) and float(si).is_integer():
            idxs.append(int(si))
        if isinstance(av, str):
            ann[av] = ann.get(av, 0) + 1
        elif av is not None:
            s = str(av)
            ann[s] = ann.get(s, 0) + 1
    if not idxs:
        return len(lst), None, None, ann
    idxs.sort()
    return len(lst), idxs[0], idxs[-1], ann


def _sum_noise_samples(noises: Any, fs: Optional[int]) -> Tuple[int, int]:
    if isinstance(noises, dict):
        noises = _pick_variant_dict(noises, ("merged", "human", "ml"))
    if not isinstance(noises, list):
        return 0, 0

    cnt = 0
    total = 0
    for it in noises:
        a = b = None
        if isinstance(it, dict):
            a = it.get("startIndex", it.get("startSample"))
            b = it.get("endIndex", it.get("endSample"))
            if (a is None or b is None) and fs:
                t0 = it.get("startTime")
                t1 = it.get("endTime")
                if (
                    isinstance(t0, (int, float, np.floating))
                    and isinstance(t1, (int, float, np.floating))
                    and t1 > t0
                ):
                    a = int(round(float(t0) * fs))
                    b = int(round(float(t1) * fs))
        elif isinstance(it, (list, tuple)) and len(it) >= 2:
            a, b = it[0], it[1]

        if isinstance(a, (float, np.floating)) and float(a).is_integer():
            a = int(a)
        if isinstance(b, (float, np.floating)) and float(b).is_integer():
            b = int(b)

        if isinstance(a, int) and isinstance(b, int) and b > a:
            cnt += 1
            total += (b - a)
    return cnt, total


def load_json_fs(p: Path) -> Any:
    with open(p, "r", encoding="utf-8") as f:
        return json.load(f)


def file_size_fs(p: Path) -> int:
    return int(p.stat().st_size)


def sample_count_fs(p: Path, ext: str) -> int:
    ext = (ext or "").lower()
    if ext == ".npy":
        arr = np.load(p, allow_pickle=False)
        return int(getattr(arr, "size", 0))
    return int(file_size_fs(p) // 4)


def summarize_record(
    sequenceId: str,
    basename: str,
    json_ref: Path,
    data_ref: Path,
    data_ext: str,
    *,
    load_json_fn: Callable[[Path], Any],
    file_size_fn: Callable[[Path], int],
    sample_count_fn: Callable[[Path, str], int],
    fs: int,
) -> Tuple[RecordSummary, Dict[str, Any]]:
    try:
        meta = load_json_fn(json_ref)
        json_ok = True
    except Exception as exc:
        print(f"WARN: failed to load JSON for {sequenceId}/{basename}: {exc}")
        meta = {}
        json_ok = False

    bsz = int(file_size_fn(data_ref))
    samples = int(sample_count_fn(data_ref, data_ext))

    channel_count = meta.get("channelCount") if isinstance(meta, dict) else None
    user_id = meta.get("userId") if isinstance(meta, dict) else None
    recordingId = meta.get("recordingId") if isinstance(meta, dict) else None
    flags_list = _flatten_flags(meta.get("flags") if isinstance(meta, dict) else None)
    comment = meta.get("comment") if isinstance(meta, dict) else None

    rpeaks_any = meta.get("rpeaks") if isinstance(meta, dict) else None
    rpeaks_count, *_ = _summarize_rpeaks(rpeaks_any)

    mrk_counts_any = meta.get("rpeakAnnotationCounts") if isinstance(meta, dict) else None

    def _counts_for_h_columns() -> Tuple[Dict[str, int], bool]:
        for var in ("merged", "human"):
            d = _extract_mrk_counts_for_variant(mrk_counts_any, var)
            if d:
                return d, True

        for var in ("merged", "human"):
            rp_list = _extract_rpeaks_list_for_variant(rpeaks_any, var)
            if rp_list:
                tmp: Dict[str, int] = {}
                for rp in rp_list:
                    av = rp.get("annotationValue")
                    if isinstance(av, str):
                        tmp[av] = tmp.get(av, 0) + 1
                    elif av is not None:
                        tmp[str(av)] = tmp.get(str(av), 0) + 1
                if tmp:
                    return tmp, True

        flat = _extract_mrk_counts(mrk_counts_any)
        if flat:
            return flat, True

        return {}, False

    h_counts_primary, has_h_counts = _counts_for_h_columns()
    h_n = int(h_counts_primary.get("N", 0)) if has_h_counts else 0
    h_s = int(h_counts_primary.get("S", 0)) if has_h_counts else 0
    h_v = int(h_counts_primary.get("V", 0)) if has_h_counts else 0
    h_u = int(h_counts_primary.get("U", 0)) if has_h_counts else 0

    ml_counts = _extract_mrk_counts_for_variant(mrk_counts_any, "ml")
    if not ml_counts:
        ml_rpeaks = _extract_rpeaks_list_for_variant(rpeaks_any, "ml")
        tmp2: Dict[str, int] = {}
        for rp in ml_rpeaks:
            av = rp.get("annotationValue")
            if isinstance(av, str):
                tmp2[av] = tmp2.get(av, 0) + 1
            elif av is not None:
                tmp2[str(av)] = tmp2.get(str(av), 0) + 1
        ml_counts = tmp2

    ml_s = int(ml_counts.get("S", 0))
    ml_v = int(ml_counts.get("V", 0))
    ml_u = int(ml_counts.get("U", 0))

    h_noises = meta.get("noises_annotated") if isinstance(meta, dict) else None
    h_nz_cnt, h_nz_samples = _sum_noise_samples(h_noises, fs)

    ml_noises = meta.get("noises") if isinstance(meta, dict) else None
    ml_nz_cnt, ml_nz_samples = _sum_noise_samples(ml_noises, fs)

    duration_s = (samples / fs) if (fs and fs > 0 and samples > 0) else None
    ml_noises_fraction = (ml_nz_samples / samples) * 100.0 if samples > 0 else None
    h_noises_fraction = (h_nz_samples / samples) * 100.0 if samples > 0 else None

    allowed_keys = {
        "channelCount",
        "comment",
        "flags",
        "noises",
        "noises_annotated",
        "recordingId",
        "rpeakAnnotationCounts",
        "rpeaks",
        "userId",
    }
    meta_keys = set(meta.keys()) if isinstance(meta, dict) else set()
    json_keys_correct = meta_keys.issubset(allowed_keys) if meta_keys else False

    rec = RecordSummary(
        sequenceId=str(sequenceId),
        recordingId=str(recordingId) if isinstance(recordingId, str) else None,
        basename=str(basename),
        channel_count=int(channel_count) if isinstance(channel_count, int) else None,
        user_id=str(user_id) if isinstance(user_id, str) else None,
        bin_bytes=bsz,
        samples=samples,
        duration_s=float(duration_s) if duration_s is not None else None,
        cmt=str(comment) if isinstance(comment, str) else None,
        flags_count=len(flags_list),
        flags=flags_list,
        rpeaks_count=int(rpeaks_count),
        h_n_count=h_n,
        h_s_count=h_s,
        h_v_count=h_v,
        h_u_count=h_u,
        ml_s_count=int(ml_s),
        ml_v_count=int(ml_v),
        ml_u_count=int(ml_u),
        json_ok=bool(json_ok),
        ml_noises_count=int(ml_nz_cnt),
        ml_noises_samples=int(ml_nz_samples),
        ml_noises_fraction=float(ml_noises_fraction) if ml_noises_fraction is not None else None,
        h_noises_count=int(h_nz_cnt),
        h_noises_samples=int(h_nz_samples),
        h_noises_fraction=float(h_noises_fraction) if h_noises_fraction is not None else None,
        json_keys_correct=bool(json_keys_correct),
    )

    rec_dict = asdict(rec)
    rec_dict["_ml_noises_samples"] = int(ml_nz_samples)
    rec_dict["_h_noises_samples"] = int(h_nz_samples)
    return rec, rec_dict


def infer_sequence_id_fs(json_path: Path, fallback: str) -> str:
    parts = list(json_path.parts)
    if "recordings" in parts:
        idx = parts.index("recordings")
        if idx >= 1:
            return parts[idx - 1]
    if len(parts) >= 2:
        return parts[-2]
    return fallback


def find_data_file_fs(json_path: Path) -> Optional[Tuple[Path, str]]:
    if json_path.suffix.lower() != ".json":
        return None
    rec_dir = json_path.parent
    stem = json_path.stem

    exact = rec_dir / stem
    if exact.exists():
        suf = exact.suffix
        if suf and suf[1:].isdigit() and len(suf) == 4:
            return exact, suf

    candidates = [
        rec_dir / f"{stem}.npy",
        rec_dir / f"{stem}.bin",
        rec_dir / f"{stem}",
    ]
    for p in sorted(rec_dir.glob(f"{stem}.*")):
        suf = p.suffix
        if suf and suf[1:].isdigit() and len(suf) == 4:
            candidates.append(p)

    for c in candidates:
        if c.exists():
            return c, c.suffix
    return None


def _flags_to_cell_value(flags_list: List[str]) -> str:
    return "; ".join([f for f in flags_list if f])


# Single source of truth for Excel column order:
# (header shown in Excel, OutputRow attribute name)
RECORD_COLUMNS: List[Tuple[str, str]] = [
    ("filename", "filename"),
    ("basename", "basename"),
    ("samples", "samples"),
    ("quality", "quality"),
    ("tag", "tag"),
    ("cmt", "cmt"),
    ("mark", "mark"),
    ("rpk_cnt", "rpk_cnt"),
    ("hN", "hN"),
    ("hS", "hS"),
    ("hV", "hV"),
    ("hU", "hU"),
    ("mlS", "mlS"),
    ("mlV", "mlV"),
    ("mlU", "mlU"),
    ("ectN", "ectN"),
    ("ectS", "ectS"),
    ("ectV", "ectV"),
    ("ectU", "ectU"),
    ("h_nz_cnt", "h_nz_cnt"),
    ("h_nz_len", "h_nz_len"),
    ("h_nz_frac%", "h_nz_frac_pct"),
    ("out", "out"),
    ("rdr", "rdr"),
    ("noi", "noi"),
    ("tp%", "tp_pct"),
    ("ml_nz_cnt", "ml_nz_cnt"),
    ("ml_nz_len", "ml_nz_len"),
    ("ml_nz_frac%", "ml_nz_frac_pct"),
    ("flags", "flags"),
    ("recordingId", "recordingId"),
    ("userId", "userId"),
    ("notes", "notes"),
]


def _validate_record_columns() -> None:
    outputrow_fields = {f.name for f in fields(OutputRow)}
    missing = [attr_name for _header, attr_name in RECORD_COLUMNS if attr_name not in outputrow_fields]
    if missing:
        raise ValueError(f"RECORD_COLUMNS references missing OutputRow fields: {missing}")


def _record_headers() -> List[str]:
    return [header for header, _attr_name in RECORD_COLUMNS]


def _record_values(row: OutputRow) -> List[Any]:
    return [getattr(row, attr_name) for _header, attr_name in RECORD_COLUMNS]


def _record_cols_map() -> Dict[str, int]:
    return {header: i + 1 for i, (header, _attr_name) in enumerate(RECORD_COLUMNS)}


def hide_columns_by_keys(ws, cols: dict, *keys: str) -> None:
    for key in keys:
        col_idx = cols.get(key) or cols.get(str(key).strip().lower())
        if col_idx:
            _split_column_dimension_if_needed(ws, col_idx)
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].hidden = True


def _split_column_dimension_if_needed(ws, col_idx: int) -> None:
    """Ensure this column has its own ColumnDimension, not a shared grouped range.

    This prevents hiding one column (e.g. mlU) from also hiding neighboring columns
    that happen to share the same dimension range in the source workbook.
    """
    target_letter = get_column_letter(col_idx)

    # Find a dimension range that covers the target column.
    for key, dim in list(ws.column_dimensions.items()):
        min_idx = getattr(dim, "min", None)
        max_idx = getattr(dim, "max", None)
        if min_idx is None or max_idx is None:
            continue
        if not (min_idx <= col_idx <= max_idx):
            continue
        if min_idx == max_idx == col_idx:
            return

        # Snapshot current formatting/behavior of the shared range.
        width = dim.width
        hidden = dim.hidden
        best_fit = dim.bestFit
        collapsed = dim.collapsed
        outline_level = dim.outline_level
        custom_width = dim.customWidth

        # Left part stays on the original dimension key.
        if min_idx < col_idx:
            dim.min = min_idx
            dim.max = col_idx - 1
        else:
            # No left part remains: remove the original grouped dimension.
            del ws.column_dimensions[key]

        # Create a dedicated dimension for the target column.
        target_dim = ColumnDimension(ws, min=col_idx, max=col_idx, width=width)
        target_dim.hidden = hidden
        target_dim.bestFit = best_fit
        target_dim.collapsed = collapsed
        target_dim.outline_level = outline_level
        if custom_width:
            target_dim.width = width
        ws.column_dimensions[target_letter] = target_dim

        # Right part, if any, becomes its own grouped dimension.
        if col_idx < max_idx:
            right_start = col_idx + 1
            right_letter = get_column_letter(right_start)
            right_dim = ColumnDimension(ws, min=right_start, max=max_idx, width=width)
            right_dim.hidden = hidden
            right_dim.bestFit = best_fit
            right_dim.collapsed = collapsed
            right_dim.outline_level = outline_level
            if custom_width:
                right_dim.width = width
            ws.column_dimensions[right_letter] = right_dim
        return



def prepare_denoising_pipeline(
    denoising_config_path: Path,
    denoising_model_dir: Path,
    disable_motions: bool = False,
) -> DenoisingPipelineConfig:
    if not denoising_config_path.exists():
        raise FileNotFoundError(f"Config file not found: {denoising_config_path}")
    if not denoising_model_dir.exists():
        raise FileNotFoundError(f"Model directory not found: {denoising_model_dir}")

    cfg = load_denoising_config_yaml(str(denoising_config_path))
    if check_denoising_config is not None:
        check_denoising_config(cfg)

    cfg.motions.model_name = resolve_model_path(denoising_model_dir, cfg.motions.model_name)
    cfg.motions.enabled = not disable_motions
    return cfg

def prepare_ectopy_pipeline(
    ectopy_config_path: Path,
    ectopy_model_dir: Path,
) -> EctopyPipelineConfig:
    
    """Entry point when the ectopy pipeline is used ad-hoc (e.g. in notebooks)."""
    
    cfg: EctopyPipelineConfig = load_ectopy_config_yaml(ectopy_config_path)

    cfg.ectopy.model_name = resolve_model_path(ectopy_model_dir, cfg.ectopy.model_name)
    cfg.ectopy.scaler_name = resolve_model_path(ectopy_model_dir, cfg.ectopy.scaler_name)

    cfg.ectopy.ectopy_removing = False

    # print("[ECTOPY] Model:", cfg.ectopy.model_name)
    # print("[ECTOPY] Scaler:", cfg.ectopy.scaler_name)
    # print("[ECTOPY] Ectopy removing enabled:", bool(cfg.ectopy.ectopy_removing))

    return cfg


def compute_ectopy_stats_placeholder(_res_denoising: Any) -> Dict[str, int]:
    """Temporary placeholder. Replace later with the real ectopy function."""
    return {"ectS": 0, "ectV": 0, "ectU": 0}


def compute_ectopy_stats(res_ectopy: EctopyPipelineResult | None) -> Dict[str, int]:
    """
    Count ectopic beat classes from rpeaks_on_denoised_df.

    Expected mapping in column 'pred':
        0 -> ectN
        1 -> ectS
        2 -> ectV
        3 -> ectU
        other values are ignored
    """
    stats = {"ectN": 0, "ectS": 0, "ectV": 0, "ectU": 0}

    if res_ectopy is None:
        return stats

    rpeaks_on_denoised_df = res_ectopy.rpeaks_on_denoised_df
    if rpeaks_on_denoised_df is None:
        return stats

    if "pred" not in rpeaks_on_denoised_df.columns:
        return stats

    counts = rpeaks_on_denoised_df["pred"].value_counts()

    return {
        "ectN": int(counts.get(0, 0)),
        "ectS": int(counts.get(1, 0)),
        "ectV": int(counts.get(2, 0)),
        "ectU": int(counts.get(3, 0)),
    }

def build_output_row(
    *,
    rec: RecordSummary,
    rec_dict: Dict[str, Any],
    filename: Optional[str],
    noise_stats: Dict[str, Any],
    ectopy_stats: Dict[str, Any],
) -> OutputRow:
    return OutputRow(
        filename=filename,
        basename=rec.basename,
        samples=int(rec.samples) if rec.samples is not None else None,
        quality=None,
        tag=None,
        cmt=rec.cmt,
        mark=None,
        rpk_cnt=int(rec.rpeaks_count),
        hN=int(rec.h_n_count),
        hS=int(rec.h_s_count),
        hV=int(rec.h_v_count),
        hU=int(rec.h_u_count),
        mlS=int(rec.ml_s_count),
        mlV=int(rec.ml_v_count),
        mlU=int(rec.ml_u_count),
        ectN=int(ectopy_stats.get("ectN", 0)),
        ectS=int(ectopy_stats.get("ectS", 0)),
        ectV=int(ectopy_stats.get("ectV", 0)),
        ectU=int(ectopy_stats.get("ectU", 0)),
        h_nz_cnt=int(rec.h_noises_count),
        h_nz_len=int(rec_dict.get("_h_noises_samples", 0)),
        h_nz_frac_pct=round(float(rec.h_noises_fraction), 1) if rec.h_noises_fraction is not None else None,
        out=noise_stats.get("out"),
        rdr=noise_stats.get("rdr"),
        noi=noise_stats.get("noi"),
        tp_pct=round(float(noise_stats.get("tp_pct")), 1) if noise_stats.get("tp_pct") is not None else None,
        ml_nz_cnt=int(rec.ml_noises_count),
        ml_nz_len=int(rec_dict.get("_ml_noises_samples", 0)),
        ml_nz_frac_pct=round(float(rec.ml_noises_fraction), 1) if rec.ml_noises_fraction is not None else None,
        flags=_flags_to_cell_value(rec.flags or []),
        recordingId=rec.recordingId,
        userId=rec.user_id,
        notes=None,
    )


def auto_adjust_widths(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)



def write_output_workbook(
    rows: List[OutputRow],
    out_path: Path,
    *,
    cfg_denoising: DenoisingPipelineConfig,
    denoising_model_dir: Path,
    cfg_ectopy: EctopyPipelineConfig,
    ectopy_model_dir: Path,
) -> None:
    # creates a new Excel workbook in memory
    wb = Workbook()

    # First sheet = records
    ws = wb.active
    ws.title = "Records"

    _validate_record_columns()
    headers = _record_headers()
    cols = _record_cols_map()

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(_record_values(row))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    auto_adjust_widths(ws)

    # Hide internal / intermediate ML columns safely.
    # _split_column_dimension_if_needed() prevents one hidden column from
    # accidentally hiding neighboring columns that share a grouped dimension.
    hide_columns_by_keys(
        ws,
        cols,
        "ml_nz_cnt",
        "ml_nz_len",
        "ml_nz_frac%",
        "mlS",
        "mlV",
        "mlU",
    )

    # Second sheet = Parameters
    ws_params = wb.create_sheet(title="Parameters")
    # denoising model info:
    ws_params["A1"] = "Unet model dir:"
    ws_params["B1"] = str(denoising_model_dir)
    ws_params["A2"] = "Unet model:"
    ws_params["B2"] = Path(cfg_denoising.motions.model_name).name
    ws_params["A3"] = "threshold:"
    ws_params["B3"] = cfg_denoising.motions.threshold

    # ectopy model info:
    ws_params["A5"] = "Ectopy model dir:"
    ws_params["B5"] = str(ectopy_model_dir)
    ws_params["A6"] = "Ectopy model:"
    ws_params["B6"] = Path(cfg_ectopy.ectopy.model_name).name
    
    ws_params["A7"] = "Ectopy scaler:"
    ws_params["B7"] = Path(cfg_ectopy.ectopy.scaler_name).name

    # Bolding the parameter labels in column A   
    for cell_ref in ("A1", "A2", "A3",  "A5", "A6", "A7"):
        ws_params[cell_ref].font = Font(bold=True)

    auto_adjust_widths(ws_params)

    # Sheet "Marks" with instructions for annotators
    ws_marks = wb.create_sheet(title="Marks")

    marks_text = {
        "B2":  "Stulpelis tag:",
        "B3":  "tag = 1111",
        "C3":  "Kokybė gera, tinka mokymui",
        "B4":  "tag = 11110",
        "C4":  "Kokybė gera, pretendentas mokymui, bus patvirtinta po anotavimo",
        "B6":  "tag = 2222",
        "C6":  "Kokybė nebloga, tačiau ant ribos, gal labiau tinka triukšmų testavimui",
        "B7":  "tag = 22220",
        "C7":  "Pretendentas testavimui, bus patvirtinta po anotavimo",
        "B9":  "tag = 3333",
        "C9":  "Gana daug triukšmų, gali tikti triukšmų testavimui",
        "B10": "tag = 33330",
        "C10":  "Gana daug triukšmų, gali pretenduoti triukšmų testavimui, bus patvirtinta po anotavimo",
        "B12": "tag = 5555",
        "C12": "Ypatingi, 'keisti' atvejai",
        "B13": "tag = 9999",
        "C13": "Kurie niekam netinka",
        "B16": "Stulpelis cmt:",
        "B17": "N, J, Z",
        "C17": "Nika anotavo triukšmų intervalus, J - Jonas anotavo ekstrasistoles, Z - Žygimantas anotavo ekstrasistoles",
        "B19": "Stulpelis mark:",
        "B20": "excl",
        "C20": "eliminuojamas įrašas iš mokymo ir testinės imčių",
        "B21": "ect8",
        "C21": "sąrašas ektopinių dūžių anotavimui",
        "B22": "nz8",
        "C22": "sąrašas triukšmų žymėjimui",
        "B23": "ect8.nz8",
        "C23": "abu sąrašai",
    }

    for cell_ref, value in marks_text.items():
        ws_marks[cell_ref] = value

    for cell_ref in ("B3", "B16", "B19"):
        ws_marks[cell_ref].font = Font(bold=True)

    for row in ws_marks.iter_rows(min_row=1, max_row=30, min_col=2, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_marks.column_dimensions["A"].width = 4
    ws_marks.column_dimensions["B"].width = 18
    ws_marks.column_dimensions["C"].width = 110

    for r in (16, 20, 21, 22, 23):
        ws_marks.row_dimensions[r].height = 30

    wb.create_sheet(title="Notes")

    wb.save(out_path)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", type=Path, required=True, help="Directory containing JSON metadata and ECG files")
    ap.add_argument("--fs", type=int, default=200, help="Sampling frequency in Hz. Default: 200")
    ap.add_argument("--cfg-denoising", type=Path, required=True, help="Denoising config path")
    ap.add_argument("--unet-model-dir", type=Path, required=True, help="Model directory for denoising/motions")
    ap.add_argument("--disable-motions", action="store_true", help="Disable motions stage in denoising pipeline")
    ap.add_argument("--cfg-ectopy", type=Path, required=True, help="Ectopy config path")
    ap.add_argument("--ectopy-model-dir", type=Path, required=True, help="Model directory for ectopy detection")
    ap.add_argument("--out", type=Path, default=None, help="Output Excel path")
    args = ap.parse_args()

    src = args.dir
    if not src.exists() or not src.is_dir():
        raise FileNotFoundError(f"--dir must be an existing directory. Got: {src}")

        # PREPARE DENOISING PART 

    cfg_denoising = prepare_denoising_pipeline(
        denoising_config_path=args.cfg_denoising,
        denoising_model_dir=args.unet_model_dir,
        disable_motions=False,
        # disable_motions=args.disable_motions,
    )
    denoising_pipe = ECGDenoisingPipeline(cfg_denoising)
    
    cfg_denoising_disabled_motions = prepare_denoising_pipeline(
        denoising_config_path=args.cfg_denoising,
        denoising_model_dir=args.unet_model_dir,
        disable_motions=True,
        # disable_motions=args.disable_motions,
    )
    denoising_pipe_disabled_motions = ECGDenoisingPipeline(cfg_denoising_disabled_motions)

        # PREPARE ECTOPY DETECTION PART 

    cfg_ectopy = prepare_ectopy_pipeline(
        ectopy_config_path=args.cfg_ectopy,
        ectopy_model_dir=args.ectopy_model_dir
    )
    ectopy_pipe = ECGEctopyPipeline(cfg_ectopy)


        # CYCLE through JSON and data files and update Excel rows 

    json_paths = [p for p in src.rglob("*.json") if "__MACOSX" not in p.parts and p.is_file()]
    print(f"Source: {src}")
    print(f"Found JSON files: {len(json_paths)}")

    output_rows: List[OutputRow] = []
    processed = 0
    skipped_missing_data = 0

    for jp in sorted(json_paths):
        basename = jp.stem
        if not basename or basename.lower() == "manifest":
            continue

        data_info = find_data_file_fs(jp)
        if data_info is None:
            print(f"WARN: data file not found for JSON: {jp}")
            skipped_missing_data += 1
            continue

        data_path, data_ext = data_info
        sequence_id = infer_sequence_id_fs(jp, fallback=src.name or "dir")

        rec, rec_dict = summarize_record(
            sequenceId=sequence_id,
            basename=basename,
            json_ref=jp,
            data_ref=data_path,
            data_ext=data_ext,
            load_json_fn=load_json_fs,
            file_size_fn=file_size_fs,
            sample_count_fn=sample_count_fs,
            fs=args.fs,
        )

        noise_stats: Dict[str, Any] = {}
        ectopy_stats: Dict[str, Any] = {"ectS": 0, "ectV": 0, "ectU": 0}
        try:
            # With activated motions stage:
            x = load_ecg_npy(data_path)
            
            # denoising and getting noise stats
            res_denoising = denoising_pipe.run(x, gaps_indices=[])
            noise_stats = calc_noise_stats_from_denoised_result(res_denoising) or {}
            
            # Without detecting motions stage (for comparison/debugging):
            x = load_ecg_npy(data_path)
            
           # denoising and getting noise stats
            res_denoising_disabled_motions = denoising_pipe_disabled_motions.run(x, gaps_indices=[])
            
            # detecting ectopies and getting ectopy stats (without  detecting motions):
            res_ectopy = ectopy_pipe.run(res_denoising_disabled_motions, fs=args.fs)
            # print(f"DEBUG: Ectopy stats for {data_path.name}: {res_ectopy}")
            ectopy_stats = compute_ectopy_stats(res_ectopy)
            print(f"Processed {data_path.name}: noise_stats={noise_stats} with enabled detecting motions")
            print(f"ectopy_stats={ectopy_stats} with disabled detecting motions")
            
        except Exception as exc:
            print(f"WARN: failed denoising/stat extraction for {data_path.name}: {exc}")

        output_rows.append(
            build_output_row(
                rec=rec,
                rec_dict=rec_dict,
                filename=data_path.name,
                noise_stats=noise_stats,
                ectopy_stats=ectopy_stats,
            )
        )
        processed += 1
    
    # Optional: hide temporarely the columns if they exist
    # hide_columns_by_keys(ws, cols, "ml_nz_cnt", "ml_nz_len", "ml_nz_frac%", "mlS", "mlV", "mlU")

    out_path = args.out if args.out else src / "records_summary.xlsx"
    write_output_workbook(
        output_rows,
        out_path,
        cfg_denoising=cfg_denoising,
        denoising_model_dir=args.unet_model_dir,
        cfg_ectopy=cfg_ectopy,
        ectopy_model_dir=args.ectopy_model_dir,
    )

    print(f"Processed records: {processed}")
    print(f"Skipped due to missing data file: {skipped_missing_data}")
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
