#!/usr/bin/env python3
"""
Build an Excel summary table directly from a directory of ECG data files and JSON metadata.

This script is tailored from update_records_list_v4.py.
It keeps the RecordSummary dataclass and the existing JSON/record summarization helpers,
but changes the overall workflow from "update an existing Excel workbook" to
"scan a directory, compute/export one output workbook from scratch".

Input
-----
A directory containing JSON metadata files and accompanying ECG data files
(e.g. .npy, .bin, or files with 4-digit numeric suffixes).

Output
------
A new .xlsx file with columns in this exact order:

filename, basename, samples, tag, cmt, rpk_cnt, hN, hS, hV, hU,
ectS, ectV, ectU, h_nz_cnt, h_nz_len, h_nz_frac%, out, rdr, noi, tp%,
ml_nz_cnt, ml_nz_len, ml_nz_frac%, flags, recordingId, userId, notes

Notes
-----
- Columns tag and notes are intentionally left empty for later manual editing.
- ECG-derived denoising statistics are calculated from:
      x = load_ecg_npy(data_path)
      res_denoising = pipe.run(x, gaps_indices=[])
      noise_stats = calc_noise_stats_from_denoised_result(res_denoising) or {}
- Ectopy columns ectS/ectV/ectU currently come from a placeholder function
  returning zeros and can be replaced later with a real implementation.
"""

from __future__ import annotations

import argparse
import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple, Union

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from record_noise_stats import calc_noise_stats_from_denoised_result

try:
    from ecg_denoising_pipeline import (
        ECGDenoisingPipeline,
        DenoisingPipelineConfig,
        load_denoising_config_yaml,
        load_ecg_npy,
        resolve_model_path,
    )
except Exception as exc:
    raise ImportError(
        "Cannot import ECG denoising pipeline helpers. "
        "Update imports to match your project structure.\n"
        f"Original error: {exc}"
    ) from exc

try:
    from ecg_denoising_pipeline.steps import check_denoising_config
except Exception:
    check_denoising_config = None

JsonLikeRef = Union[str, Path]


@dataclass
class RecordSummary:
    sequenceId: str
    recordingId: Optional[str]
    basename: str
    channel_count: Optional[int]
    user_id: Optional[str]

    bin_bytes: int
    samples: int
    duration_s: Optional[float]

    flags_count: int
    flags: List[str]

    rpeaks_count: int
    h_n_count: int
    h_s_count: int
    h_v_count: int
    h_u_count: int

    ml_s_count: int
    ml_v_count: int
    ml_u_count: int

    json_ok: bool

    ml_noises_count: int
    ml_noises_samples: int
    ml_noises_fraction: Optional[float]  # percent 0..100

    h_noises_count: int
    h_noises_samples: int
    h_noises_fraction: Optional[float]  # percent 0..100

    cmt: Optional[str]
    json_keys_correct: bool
    notes: Optional[str] = None


@dataclass
class OutputRow:
    filename: Optional[str]
    basename: str
    samples: Optional[int]
    tag: Optional[str]
    cmt: Optional[str]
    rpk_cnt: Optional[int]
    hN: Optional[int]
    hS: Optional[int]
    hV: Optional[int]
    hU: Optional[int]
    ectS: int
    ectV: int
    ectU: int
    h_nz_cnt: Optional[int]
    h_nz_len: Optional[int]
    h_nz_frac_pct: Optional[float]
    out: Optional[int]
    rdr: Optional[int]
    noi: Optional[int]
    tp_pct: Optional[float]
    ml_nz_cnt: Optional[int]
    ml_nz_len: Optional[int]
    ml_nz_frac_pct: Optional[float]
    flags: Optional[str]
    recordingId: Optional[str]
    userId: Optional[str]
    notes: Optional[str]


def _flatten_flags(flags: Any) -> List[str]:
    if flags is None:
        return []
    if isinstance(flags, list):
        return [str(x) for x in flags]
    return [str(flags)]


def _pick_variant_dict(obj: Any, keys_preference: Tuple[str, ...] = ("merged", "human", "ml")) -> Any:
    if not isinstance(obj, dict):
        return None
    for k in keys_preference:
        if k in obj:
            return obj.get(k)
    for k, v in obj.items():
        if k == "__info":
            continue
        return v
    return None


def _extract_rpeaks_list(rpeaks: Any) -> List[Dict[str, Any]]:
    if isinstance(rpeaks, list):
        return [x for x in rpeaks if isinstance(x, dict)]
    if isinstance(rpeaks, dict):
        v = _pick_variant_dict(rpeaks, ("merged", "human", "ml"))
        if isinstance(v, list):
            return [x for x in v if isinstance(x, dict)]
    return []


def _extract_rpeaks_list_for_variant(rpeaks: Any, variant: str) -> List[Dict[str, Any]]:
    if isinstance(rpeaks, dict):
        v = rpeaks.get(variant)
        if isinstance(v, list):
            return [x for x in v if isinstance(x, dict)]
    return []


def _extract_mrk_counts_for_variant(h_counts: Any, variant: str) -> Dict[str, int]:
    if not isinstance(h_counts, dict):
        return {}
    v = h_counts.get(variant)
    if not isinstance(v, dict):
        return {}
    out: Dict[str, int] = {}
    for k, val in v.items():
        if k == "__info":
            continue
        if isinstance(val, int):
            out[str(k)] = int(val)
        elif isinstance(val, (float, np.floating)) and float(val).is_integer():
            out[str(k)] = int(val)
    return out


def _extract_mrk_counts(h_counts: Any) -> Dict[str, int]:
    if not isinstance(h_counts, dict):
        return {}
    if any(isinstance(v, dict) for k, v in h_counts.items() if k != "__info"):
        v = _pick_variant_dict(h_counts, ("merged", "human", "ml"))
        if isinstance(v, dict):
            out: Dict[str, int] = {}
            for k, val in v.items():
                if k == "__info":
                    continue
                if isinstance(val, int):
                    out[str(k)] = int(val)
                elif isinstance(val, (float, np.floating)) and float(val).is_integer():
                    out[str(k)] = int(val)
            return out
        return {}

    out2: Dict[str, int] = {}
    for k, v in h_counts.items():
        if k == "__info":
            continue
        if isinstance(v, int):
            out2[str(k)] = int(v)
        elif isinstance(v, (float, np.floating)) and float(v).is_integer():
            out2[str(k)] = int(v)
    return out2


def _summarize_rpeaks(rpeaks: Any) -> Tuple[int, Optional[int], Optional[int], Dict[str, int]]:
    lst = _extract_rpeaks_list(rpeaks)
    if not lst:
        return 0, None, None, {}
    idxs: List[int] = []
    ann: Dict[str, int] = {}
    for rp in lst:
        si = rp.get("sampleIndex")
        av = rp.get("annotationValue")
        if isinstance(si, int):
            idxs.append(si)
        elif isinstance(si, (float, np.floating)) and float(si).is_integer():
            idxs.append(int(si))
        if isinstance(av, str):
            ann[av] = ann.get(av, 0) + 1
        elif av is not None:
            s = str(av)
            ann[s] = ann.get(s, 0) + 1
    if not idxs:
        return len(lst), None, None, ann
    idxs.sort()
    return len(lst), idxs[0], idxs[-1], ann


def _sum_noise_samples(noises: Any, fs: Optional[int]) -> Tuple[int, int]:
    if isinstance(noises, dict):
        noises = _pick_variant_dict(noises, ("merged", "human", "ml"))
    if not isinstance(noises, list):
        return 0, 0

    cnt = 0
    total = 0
    for it in noises:
        a = b = None
        if isinstance(it, dict):
            a = it.get("startIndex", it.get("startSample"))
            b = it.get("endIndex", it.get("endSample"))
            if (a is None or b is None) and fs:
                t0 = it.get("startTime")
                t1 = it.get("endTime")
                if (
                    isinstance(t0, (int, float, np.floating))
                    and isinstance(t1, (int, float, np.floating))
                    and t1 > t0
                ):
                    a = int(round(float(t0) * fs))
                    b = int(round(float(t1) * fs))
        elif isinstance(it, (list, tuple)) and len(it) >= 2:
            a, b = it[0], it[1]

        if isinstance(a, (float, np.floating)) and float(a).is_integer():
            a = int(a)
        if isinstance(b, (float, np.floating)) and float(b).is_integer():
            b = int(b)

        if isinstance(a, int) and isinstance(b, int) and b > a:
            cnt += 1
            total += (b - a)
    return cnt, total


def load_json_fs(p: Path) -> Any:
    with open(p, "r", encoding="utf-8") as f:
        return json.load(f)


def file_size_fs(p: Path) -> int:
    return int(p.stat().st_size)


def sample_count_fs(p: Path, ext: str) -> int:
    ext = (ext or "").lower()
    if ext == ".npy":
        arr = np.load(p, allow_pickle=False)
        return int(getattr(arr, "size", 0))
    return int(file_size_fs(p) // 4)


def summarize_record(
    sequenceId: str,
    basename: str,
    json_ref: Path,
    data_ref: Path,
    data_ext: str,
    *,
    load_json_fn: Callable[[Path], Any],
    file_size_fn: Callable[[Path], int],
    sample_count_fn: Callable[[Path, str], int],
    fs: int,
) -> Tuple[RecordSummary, Dict[str, Any]]:
    try:
        meta = load_json_fn(json_ref)
        json_ok = True
    except Exception as exc:
        print(f"WARN: failed to load JSON for {sequenceId}/{basename}: {exc}")
        meta = {}
        json_ok = False

    bsz = int(file_size_fn(data_ref))
    samples = int(sample_count_fn(data_ref, data_ext))

    channel_count = meta.get("channelCount") if isinstance(meta, dict) else None
    user_id = meta.get("userId") if isinstance(meta, dict) else None
    recordingId = meta.get("recordingId") if isinstance(meta, dict) else None
    flags_list = _flatten_flags(meta.get("flags") if isinstance(meta, dict) else None)
    comment = meta.get("comment") if isinstance(meta, dict) else None

    rpeaks_any = meta.get("rpeaks") if isinstance(meta, dict) else None
    rpeaks_count, *_ = _summarize_rpeaks(rpeaks_any)

    mrk_counts_any = meta.get("rpeakAnnotationCounts") if isinstance(meta, dict) else None

    def _counts_for_h_columns() -> Tuple[Dict[str, int], bool]:
        for var in ("merged", "human"):
            d = _extract_mrk_counts_for_variant(mrk_counts_any, var)
            if d:
                return d, True

        for var in ("merged", "human"):
            rp_list = _extract_rpeaks_list_for_variant(rpeaks_any, var)
            if rp_list:
                tmp: Dict[str, int] = {}
                for rp in rp_list:
                    av = rp.get("annotationValue")
                    if isinstance(av, str):
                        tmp[av] = tmp.get(av, 0) + 1
                    elif av is not None:
                        tmp[str(av)] = tmp.get(str(av), 0) + 1
                if tmp:
                    return tmp, True

        flat = _extract_mrk_counts(mrk_counts_any)
        if flat:
            return flat, True

        return {}, False

    h_counts_primary, has_h_counts = _counts_for_h_columns()
    h_n = int(h_counts_primary.get("N", 0)) if has_h_counts else 0
    h_s = int(h_counts_primary.get("S", 0)) if has_h_counts else 0
    h_v = int(h_counts_primary.get("V", 0)) if has_h_counts else 0
    h_u = int(h_counts_primary.get("U", 0)) if has_h_counts else 0

    ml_counts = _extract_mrk_counts_for_variant(mrk_counts_any, "ml")
    if not ml_counts:
        ml_rpeaks = _extract_rpeaks_list_for_variant(rpeaks_any, "ml")
        tmp2: Dict[str, int] = {}
        for rp in ml_rpeaks:
            av = rp.get("annotationValue")
            if isinstance(av, str):
                tmp2[av] = tmp2.get(av, 0) + 1
            elif av is not None:
                tmp2[str(av)] = tmp2.get(str(av), 0) + 1
        ml_counts = tmp2

    ml_s = int(ml_counts.get("S", 0))
    ml_v = int(ml_counts.get("V", 0))
    ml_u = int(ml_counts.get("U", 0))

    h_noises = meta.get("noises_annotated") if isinstance(meta, dict) else None
    h_nz_cnt, h_nz_samples = _sum_noise_samples(h_noises, fs)

    ml_noises = meta.get("noises") if isinstance(meta, dict) else None
    ml_nz_cnt, ml_nz_samples = _sum_noise_samples(ml_noises, fs)

    duration_s = (samples / fs) if (fs and fs > 0 and samples > 0) else None
    ml_noises_fraction = (ml_nz_samples / samples) * 100.0 if samples > 0 else None
    h_noises_fraction = (h_nz_samples / samples) * 100.0 if samples > 0 else None

    allowed_keys = {
        "channelCount",
        "comment",
        "flags",
        "noises",
        "noises_annotated",
        "recordingId",
        "rpeakAnnotationCounts",
        "rpeaks",
        "userId",
    }
    meta_keys = set(meta.keys()) if isinstance(meta, dict) else set()
    json_keys_correct = meta_keys.issubset(allowed_keys) if meta_keys else False

    rec = RecordSummary(
        sequenceId=str(sequenceId),
        recordingId=str(recordingId) if isinstance(recordingId, str) else None,
        basename=str(basename),
        channel_count=int(channel_count) if isinstance(channel_count, int) else None,
        user_id=str(user_id) if isinstance(user_id, str) else None,
        bin_bytes=bsz,
        samples=samples,
        duration_s=float(duration_s) if duration_s is not None else None,
        cmt=str(comment) if isinstance(comment, str) else None,
        flags_count=len(flags_list),
        flags=flags_list,
        rpeaks_count=int(rpeaks_count),
        h_n_count=h_n,
        h_s_count=h_s,
        h_v_count=h_v,
        h_u_count=h_u,
        ml_s_count=int(ml_s),
        ml_v_count=int(ml_v),
        ml_u_count=int(ml_u),
        json_ok=bool(json_ok),
        ml_noises_count=int(ml_nz_cnt),
        ml_noises_samples=int(ml_nz_samples),
        ml_noises_fraction=float(ml_noises_fraction) if ml_noises_fraction is not None else None,
        h_noises_count=int(h_nz_cnt),
        h_noises_samples=int(h_nz_samples),
        h_noises_fraction=float(h_noises_fraction) if h_noises_fraction is not None else None,
        json_keys_correct=bool(json_keys_correct),
    )

    rec_dict = asdict(rec)
    rec_dict["_ml_noises_samples"] = int(ml_nz_samples)
    rec_dict["_h_noises_samples"] = int(h_nz_samples)
    return rec, rec_dict


def infer_sequence_id_fs(json_path: Path, fallback: str) -> str:
    parts = list(json_path.parts)
    if "recordings" in parts:
        idx = parts.index("recordings")
        if idx >= 1:
            return parts[idx - 1]
    if len(parts) >= 2:
        return parts[-2]
    return fallback


def find_data_file_fs(json_path: Path) -> Optional[Tuple[Path, str]]:
    if json_path.suffix.lower() != ".json":
        return None
    rec_dir = json_path.parent
    stem = json_path.stem

    exact = rec_dir / stem
    if exact.exists():
        suf = exact.suffix
        if suf and suf[1:].isdigit() and len(suf) == 4:
            return exact, suf

    candidates = [
        rec_dir / f"{stem}.npy",
        rec_dir / f"{stem}.bin",
        rec_dir / f"{stem}",
    ]
    for p in sorted(rec_dir.glob(f"{stem}.*")):
        suf = p.suffix
        if suf and suf[1:].isdigit() and len(suf) == 4:
            candidates.append(p)

    for c in candidates:
        if c.exists():
            return c, c.suffix
    return None


def _flags_to_cell_value(flags_list: List[str]) -> str:
    return "; ".join([f for f in flags_list if f])


def prepare_denoising_pipeline(
    config_path: Path,
    model_dir: Path,
    disable_motions: bool = False,
) -> DenoisingPipelineConfig:
    if not config_path.exists():
        raise FileNotFoundError(f"Config file not found: {config_path}")
    if not model_dir.exists():
        raise FileNotFoundError(f"Model directory not found: {model_dir}")

    cfg = load_denoising_config_yaml(str(config_path))
    if check_denoising_config is not None:
        check_denoising_config(cfg)

    cfg.motions.model_name = resolve_model_path(model_dir, cfg.motions.model_name)
    cfg.motions.enabled = not disable_motions
    return cfg


def compute_ectopy_stats_placeholder(_res_denoising: Any) -> Dict[str, int]:
    """Temporary placeholder. Replace later with the real ectopy function."""
    return {"ectS": 0, "ectV": 0, "ectU": 0}


def build_output_row(
    *,
    rec: RecordSummary,
    rec_dict: Dict[str, Any],
    filename: Optional[str],
    noise_stats: Dict[str, Any],
    ectopy_stats: Dict[str, Any],
) -> OutputRow:
    return OutputRow(
        filename=filename,
        basename=rec.basename,
        samples=int(rec.samples) if rec.samples is not None else None,
        tag=None,
        cmt=rec.cmt,
        rpk_cnt=int(rec.rpeaks_count),
        hN=int(rec.h_n_count),
        hS=int(rec.h_s_count),
        hV=int(rec.h_v_count),
        hU=int(rec.h_u_count),
        ectS=int(ectopy_stats.get("ectS", 0)),
        ectV=int(ectopy_stats.get("ectV", 0)),
        ectU=int(ectopy_stats.get("ectU", 0)),
        h_nz_cnt=int(rec.h_noises_count),
        h_nz_len=int(rec_dict.get("_h_noises_samples", 0)),
        h_nz_frac_pct=round(float(rec.h_noises_fraction), 1) if rec.h_noises_fraction is not None else None,
        out=noise_stats.get("out"),
        rdr=noise_stats.get("rdr"),
        noi=noise_stats.get("noi"),
        tp_pct=round(float(noise_stats.get("tp_pct")), 1) if noise_stats.get("tp_pct") is not None else None,
        ml_nz_cnt=int(rec.ml_noises_count),
        ml_nz_len=int(rec_dict.get("_ml_noises_samples", 0)),
        ml_nz_frac_pct=round(float(rec.ml_noises_fraction), 1) if rec.ml_noises_fraction is not None else None,
        flags=_flags_to_cell_value(rec.flags or []),
        recordingId=rec.recordingId,
        userId=rec.user_id,
        notes=None,
    )


def auto_adjust_widths(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)



def write_output_workbook(
    rows: List[OutputRow],
    out_path: Path,
    *,
    cfg_denoising: DenoisingPipelineConfig,
    model_dir: Path,
) -> None:
    # creates a new Excel workbook in memory
    wb = Workbook()

    # First sheet = records
    ws = wb.active
    ws.title = "Records"

    headers = [
        "filename",
        "basename",
        "samples",
        "tag",
        "cmt",
        "rpk_cnt",
        "hN",
        "hS",
        "hV",
        "hU",
        "ectS",
        "ectV",
        "ectU",
        "h_nz_cnt",
        "h_nz_len",
        "h_nz_frac%",
        "out",
        "rdr",
        "noi",
        "tp%",
        "ml_nz_cnt",
        "ml_nz_len",
        "ml_nz_frac%",
        "flags",
        "recordingId",
        "userId",
        "notes",
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([
            row.filename,
            row.basename,
            row.samples,
            row.tag,
            row.cmt,
            row.rpk_cnt,
            row.hN,
            row.hS,
            row.hV,
            row.hU,
            row.ectS,
            row.ectV,
            row.ectU,
            row.h_nz_cnt,
            row.h_nz_len,
            row.h_nz_frac_pct,
            row.out,
            row.rdr,
            row.noi,
            row.tp_pct,
            row.ml_nz_cnt,
            row.ml_nz_len,
            row.ml_nz_frac_pct,
            row.flags,
            row.recordingId,
            row.userId,
            row.notes,
        ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    auto_adjust_widths(ws)

    # Second sheet = Parameters
    ws_params = wb.create_sheet(title="Parameters")
    ws_params["A1"] = "Unet model dir:"
    ws_params["B1"] = str(model_dir)
    ws_params["A2"] = "Unet model:"
    ws_params["B2"] = Path(cfg_denoising.motions.model_name).name
    ws_params["A3"] = "threshold:"
    ws_params["B3"] = cfg_denoising.motions.threshold

    for cell_ref in ("A1", "A2", "A3"):
        ws_params[cell_ref].font = Font(bold=True)

    auto_adjust_widths(ws_params)

    wb.save(out_path)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", type=Path, required=True, help="Directory containing JSON metadata and ECG files")
    ap.add_argument("--fs", type=int, default=200, help="Sampling frequency in Hz. Default: 200")
    ap.add_argument("--cfg-denoising", type=Path, required=True, help="Denoising config path")
    ap.add_argument("--model-dir", type=Path, required=True, help="Model directory for denoising/motions")
    ap.add_argument("--disable-motions", action="store_true", help="Disable motions stage in denoising pipeline")
    ap.add_argument("--out", type=Path, default=None, help="Output Excel path")
    args = ap.parse_args()

    src = args.dir
    if not src.exists() or not src.is_dir():
        raise FileNotFoundError(f"--dir must be an existing directory. Got: {src}")

    cfg_denoising = prepare_denoising_pipeline(
        config_path=args.cfg_denoising,
        model_dir=args.model_dir,
        disable_motions=args.disable_motions,
    )
    pipe = ECGDenoisingPipeline(cfg_denoising)




    json_paths = [p for p in src.rglob("*.json") if "__MACOSX" not in p.parts and p.is_file()]
    print(f"Source: {src}")
    print(f"Found JSON files: {len(json_paths)}")

    output_rows: List[OutputRow] = []
    processed = 0
    skipped_missing_data = 0

    for jp in sorted(json_paths):
        basename = jp.stem
        if not basename or basename.lower() == "manifest":
            continue

        data_info = find_data_file_fs(jp)
        if data_info is None:
            print(f"WARN: data file not found for JSON: {jp}")
            skipped_missing_data += 1
            continue

        data_path, data_ext = data_info
        sequence_id = infer_sequence_id_fs(jp, fallback=src.name or "dir")

        rec, rec_dict = summarize_record(
            sequenceId=sequence_id,
            basename=basename,
            json_ref=jp,
            data_ref=data_path,
            data_ext=data_ext,
            load_json_fn=load_json_fs,
            file_size_fn=file_size_fs,
            sample_count_fn=sample_count_fs,
            fs=args.fs,
        )

        noise_stats: Dict[str, Any] = {}
        ectopy_stats: Dict[str, Any] = {"ectS": 0, "ectV": 0, "ectU": 0}
        try:
            x = load_ecg_npy(data_path)
            res_denoising = pipe.run(x, gaps_indices=[])
            noise_stats = calc_noise_stats_from_denoised_result(res_denoising) or {}
            ectopy_stats = compute_ectopy_stats_placeholder(res_denoising)
        except Exception as exc:
            print(f"WARN: failed denoising/stat extraction for {data_path.name}: {exc}")

        output_rows.append(
            build_output_row(
                rec=rec,
                rec_dict=rec_dict,
                filename=data_path.name,
                noise_stats=noise_stats,
                ectopy_stats=ectopy_stats,
            )
        )
        processed += 1

    out_path = args.out if args.out else src / "records_summary.xlsx"
    write_output_workbook(
        output_rows,
        out_path,
        cfg_denoising=cfg_denoising,
        model_dir=args.model_dir,
    )

    print(f"Processed records: {processed}")
    print(f"Skipped due to missing data file: {skipped_missing_data}")
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
