#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def pick_col(df: pd.DataFrame, *names: str) -> str:
    """Return the first existing column name from candidates."""
    for n in names:
        if n in df.columns:
            return n
    raise KeyError(f"None of these columns exist: {names}")


def to_num(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce")


def drop_unnamed_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Drop Excel artifact columns like 'Unnamed: 22'."""
    mask = ~df.columns.astype(str).str.match(r"^Unnamed:\s*\d+$")
    return df.loc[:, mask]


def sort_df_grouped(
    df: pd.DataFrame,
    user_col: str,
    hS_col: str,
    hV_col: str,
    mlS_col: str,
    mlV_col: str,
    ann_nz_col: str,
    ml_nz_col: str,
) -> pd.DataFrame:
    """
    Rules:
      - Group rows by userId.
      - Inside group:
          1) Sort by hS+hV desc, tie: h_nz_frac asc
          2) If ALL rows in group have hS+hV == 0, sort by mlS+mlV desc, tie: ml_nz_frac asc
      - Add group_no as first column (1..N by first appearance of userId in output).
      - Preserve all original columns (plus group_no).
    """
    df2 = df.copy()

    df2["_ann_sum"] = to_num(df2[hS_col]).fillna(0) + to_num(df2[hV_col]).fillna(0)
    df2["_ml_sum"] = to_num(df2[mlS_col]).fillna(0) + to_num(df2[mlV_col]).fillna(0)
    df2["_ann_nz"] = to_num(df2[ann_nz_col]).fillna(float("inf"))
    df2["_ml_nz"] = to_num(df2[ml_nz_col]).fillna(float("inf"))

    def _sort_group(g: pd.DataFrame) -> pd.DataFrame:
        if (g["_ann_sum"].fillna(0) == 0).all():
            return g.sort_values(by=["_ml_sum", "_ml_nz"], ascending=[False, True], kind="mergesort")
        return g.sort_values(by=["_ann_sum", "_ann_nz"], ascending=[False, True], kind="mergesort")

    parts = []
    # deterministic group order by userId
    for _, g in df2.groupby(user_col, sort=True):
        parts.append(_sort_group(g))

    out = pd.concat(parts, axis=0).reset_index(drop=True)

    # group_no by first appearance in output
    user_order = pd.Index(out[user_col].astype(str)).drop_duplicates()
    group_map = {u: i + 1 for i, u in enumerate(user_order)}
    out.insert(0, "group_no", out[user_col].astype(str).map(group_map))

    # Drop helper columns
    out = out.drop(columns=["_ann_sum", "_ml_sum", "_ann_nz", "_ml_nz"], errors="ignore")
    return out


def write_excel_with_formatting(
    df: pd.DataFrame,
    out_path: Path,
    sheet_name: str,
    user_col: str,
    basename_col: str | None,
    highlight_hex: str = "EAF2FF",
) -> None:
    """
    - Writes df to Excel.
    - Highlights first row of each user group (light fill).
    - Forces basename column to TEXT and ensures ',' -> '.' in values.
    - Formats h_nz_frac and ml_nz_frac as one decimal + literal % sign,
      assuming values are already in percent units (12.3 means 12.3%).
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

    wb = load_workbook(out_path)
    ws = wb[sheet_name]

    header = [cell.value for cell in ws[1]]
    col_index = {name: idx + 1 for idx, name in enumerate(header)}

    # Highlight first row of each group
    fill = PatternFill(start_color=highlight_hex, end_color=highlight_hex, fill_type="solid")
    first_rows = df.groupby(user_col, sort=False).head(1).index.tolist()  # 0-based in df

    for i0 in first_rows:
        excel_row = 2 + i0  # header row is 1
        for c in range(1, ws.max_column + 1):
            ws.cell(row=excel_row, column=c).fill = fill

    # Basename: force text + '.' decimal separator
    if basename_col and basename_col in col_index:
        bcol = col_index[basename_col]
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=bcol)
            v = "" if cell.value is None else str(cell.value).replace(",", ".")
            cell.value = v
            cell.number_format = "@"  # TEXT

    # Percent columns: one decimal + literal percent sign
    # Values are already in percent units (e.g. 12.3 == 12.3%)
    for pct_col in ["h_nz_frac", "ml_nz_frac"]:
        if pct_col in col_index:
            cidx = col_index[pct_col]
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=cidx)
                if cell.value is None or str(cell.value).strip() == "":
                    continue
                try:
                    cell.value = float(cell.value)
                except Exception:
                    # leave non-numeric as-is
                    pass
                cell.number_format = '0.0"%"'

    wb.save(out_path)


def main() -> int:
    ap = argparse.ArgumentParser(description="Sort visi_zive_irasai_atrankai._modif_v1.xlsx by userId + ann/ml rules.")
    ap.add_argument("inp", type=Path, help="Input .xlsx")
    ap.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Output .xlsx (default: input stem + _sorted_grouped_highlighted_basename_dot.xlsx)",
    )
    ap.add_argument("--sheet", type=str, default="sorted", help="Output sheet name")
    ap.add_argument("--highlight", type=str, default="FFEE99", help="Hex fill for first row of each group")
    
    # Very light blue: EAF2FF
    # Light blue: D6E8FF
    # Very light yellow: FFF7CC
    # Light yellow: FFEE99
    
    
    args = ap.parse_args()

    inp: Path = args.inp
    if args.out is None:
        args.out = inp.with_name(inp.stem + "_sorted_grouped_highlighted_basename_dot.xlsx")

    df = pd.read_excel(inp)
    df = drop_unnamed_columns(df)  # <--- exclude Unnamed: 22/23/24 (and any Unnamed:*)

    # Robust column mapping (matches your file)
    user_col = pick_col(df, "userId", "userID")
    hS_col = pick_col(df, "hS")
    hV_col = pick_col(df, "hV")
    mlS_col = pick_col(df, "mlS")
    mlV_col = pick_col(df, "mlV")
    ann_nz_col = pick_col(df, "h_nz_frac", "h_nz_frac")
    ml_nz_col = pick_col(df, "ml_nz_frac", "ml_nz_frac")

    basename_col = "basename" if "basename" in df.columns else None

    df_sorted = sort_df_grouped(
        df=df,
        user_col=user_col,
        hS_col=hS_col,
        hV_col=hV_col,
        mlS_col=mlS_col,
        mlV_col=mlV_col,
        ann_nz_col=ann_nz_col,
        ml_nz_col=ml_nz_col,
    )

    # Also ensure basename fix at dataframe level
    if basename_col:
        df_sorted[basename_col] = df_sorted[basename_col].astype(str).str.replace(",", ".", regex=False)

    write_excel_with_formatting(
        df=df_sorted,
        out_path=args.out,
        sheet_name=args.sheet,
        user_col=user_col,
        basename_col=basename_col,
        highlight_hex=args.highlight,
    )

    print(f"Wrote: {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

"""
python sort_records_modif_v3.py "visi_zive_irasai_atrankai._modif_v1.xlsx"

"""